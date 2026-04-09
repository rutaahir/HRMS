from django.shortcuts import redirect # for rediretion of urls
from django.utils import timezone
from .models import Task, Attendance  # Attendance table in models.py file 
from django.contrib.auth.models import User # for built in user model
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from accounts.models import Profile
from django.contrib.auth.models import Group
from django.db.models import Count, Q, F, Max
# Aa line upar imports ma add kari de:
from .utils import ensure_db_connection
from .models import RolePermission
from django.db import transaction
import os
from django.core.management import call_command
from django.conf import settings

from .middleware import ThreadLocal
from django.http import HttpResponse
from accounts.models import is_admin, is_hr, is_manager, is_employee, is_superadmin
from .models import Leave
from .models import Company,AttendancePolicy,AttendanceCorrection,Holiday,Shift
from django.utils import timezone
from datetime import datetime,timedelta
#from .models import Permission as Permission
import calendar
from django.utils.text import slugify
from django.shortcuts import get_object_or_404
from datetime import date
from django.contrib.auth.models import Group, Permission
from datetime import datetime, time, timedelta
from django.utils import timezone
from datetime import datetime
from .models import Notification
from django.utils.timezone import localtime
from django.contrib import messages
from django.db.models import Q
from .models import Task, TaskAttachment
from django.contrib.auth.models import User, Group, Permission

# 2. Your Models (Task અને Notification સાથે)
from .models import (
    Task,                  # Team Leader માટે
    Profile, 
    Attendance, 
    Company, 
    Shift, 
    Leave, 
    AttendanceCorrection, 
    Holiday, 
    AttendancePolicy, 
    Notification           # Notification માટે
)

import math

def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371000 # Earth radius in meters
    phi_1, phi_2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi/2.0)**2 + math.cos(phi_1) * math.cos(phi_2) * math.sin(delta_lambda/2.0)**2
    return R * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))


def home(request):
    return HttpResponse("This is Home Page")


def safe_local_dt(date_obj, time_obj):
    if not date_obj or not time_obj:
        return None

    dt = datetime.combine(date_obj, time_obj)

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt)

    return timezone.localtime(dt)

# function for attendence/check_in
# attendance/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from .models import Company, Attendance, Profile

# attendance/views.py

# attendance/views.py

from django.utils import timezone


@login_required
def check_in_view(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    today = timezone.localdate()
    
    # 🌍 GEOFENCING LOGIC
    if request.method == "POST":
        user_lat = request.POST.get('latitude')
        user_lng = request.POST.get('longitude')
        
        print("\n--- 📍 GPS CHECK-IN DEBUG ---")
        print(f"User Lat: {user_lat}, User Lng: {user_lng}")
        
        # 🔥 THE FIX: Fetch Profile and Branch from the specific COMPANY DB
        try:
            profile = Profile.objects.using(db_name).select_related('branch').get(user=request.user)
            branch = profile.branch
        except Profile.DoesNotExist:
            messages.error(request, "❌ Check-In Blocked: Profile not found in company database.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        
        # 🚨 STRICT GEOFENCING: If they have a branch, we MUST check coordinates
        if branch:
            if not branch.latitude or not branch.longitude:
                messages.error(request, f"❌ Check-In Blocked: The '{branch.name}' branch does not have GPS coordinates set up! Please contact your Admin.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
                
            print(f"Branch Lat: {branch.latitude}, Branch Lng: {branch.longitude}, Radius: {branch.radius}m")
            
            if not user_lat or not user_lng:
                messages.error(request, "⚠️ Please allow Location Access in your browser to check in.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            
            try:
                distance = calculate_distance(float(user_lat), float(user_lng), branch.latitude, branch.longitude)
                radius = branch.radius if branch.radius else 50
                print(f"📏 Calculated Distance: {int(distance)} meters")
                
                if distance > radius:
                    messages.error(request, f"❌ Access Denied: You are {int(distance)} meters away from the office. Allowed radius is {radius}m.")
                    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            except ValueError:
                messages.error(request, "Invalid GPS coordinates received.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        else:
            messages.error(request, "❌ Check-In Blocked: You are not assigned to any branch!")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    # Proceed to mark attendance
    try:
        from django.contrib.auth.models import User
        try:
            User.objects.using(db_name).get(id=request.user.id)
        except User.DoesNotExist:
            request.user.save(using=db_name)

        attendance, created = Attendance.objects.using(db_name).get_or_create(
            user_id=request.user.id,
            date=today,
            defaults={'status': 'Present'}
        )

        if attendance.check_in is None:
            attendance.check_in = timezone.localtime().time()
            attendance.status = "Present"
            attendance.save(using=db_name) 
            messages.success(request, "✅ Checked In Successfully!")
        else:
            messages.info(request, "You are already checked in for today.")
            
    except Exception as e:
        messages.error(request, f"Check-In Error: {str(e)}")
    # In your check_in view
    print(f"User: {request.user.username}")
    print(f"User Branch: {request.user.profile.branch}")
    if request.user.profile.branch:
        print(f"Branch Lat/Long: {request.user.profile.branch.latitude}, {request.user.profile.branch.longitude}")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def check_out_view(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    today = timezone.localdate()
    
    # 🌍 GEOFENCING LOGIC FOR CHECKOUT
    if request.method == "POST":
        user_lat = request.POST.get('latitude')
        user_lng = request.POST.get('longitude')
        
        # 🔥 THE FIX: Fetch Profile and Branch from the specific COMPANY DB
        try:
            profile = Profile.objects.using(db_name).select_related('branch').get(user=request.user)
            branch = profile.branch
        except Profile.DoesNotExist:
            messages.error(request, "❌ Check-Out Blocked: Profile not found.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        
        if branch:
            if not branch.latitude or not branch.longitude:
                messages.error(request, f"❌ Check-Out Blocked: The '{branch.name}' branch does not have GPS coordinates set up!")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
                
            if not user_lat or not user_lng:
                messages.error(request, "⚠️ Please allow Location Access to check out.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            
            try:
                distance = calculate_distance(float(user_lat), float(user_lng), branch.latitude, branch.longitude)
                radius = branch.radius if branch.radius else 50
                
                if distance > radius:
                    messages.error(request, f"❌ Access Denied: You are {int(distance)} meters away from the office. Allowed radius is {radius}m.")
                    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            except ValueError:
                pass
                
    # Proceed to mark checkout
    try:
        attendance = Attendance.objects.using(db_name).filter(
            user_id=request.user.id, 
            date=today
        ).first()

        if attendance:
            if attendance.check_in and attendance.check_out is None:
                attendance.check_out = timezone.localtime().time()
                attendance.save(using=db_name)
                messages.success(request, "✅ Checked Out Successfully!")
            elif attendance.check_out is not None:
                messages.info(request, "You have already checked out for today.")
            else:
                messages.warning(request, "Please Check In first.")
        else:
            messages.error(request, "No attendance record found for today.")
            
    except Exception as e:
        messages.error(request, f"Check-Out Error: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    
from datetime import datetime, timedelta  # Ensure these are imported
from django.utils import timezone

# attendance/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, time
from .models import Attendance, AttendancePolicy
from accounts.models import Profile

@login_required
def dashboard_view(request):
    user = request.user
    
    # 🔥 FIX 1: સુપરએડમિન માટે હંમેશા default DB જ રાખો
    if user.is_superuser:
        ThreadLocal.DB_NAME = 'default'
        return redirect('superadmin_dashboard')

    # બાકીના યુઝર્સ માટે કંપની મુજબ સ્વિચ કરો
    try:
        login_profile = user.profile
        company = login_profile.company
        if company:
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
        else:
            ThreadLocal.DB_NAME = 'default'
    except Profile.DoesNotExist:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile not found")

    # 🔥 2. ROLE BASED REDIRECTS (The Router)
  
    # A. Super Admin Check
    if user.is_superuser:
        return redirect('superadmin_dashboard')

    # B. Company Owner Check 
    
    if user.groups.filter(name__icontains='_owner').exists():
        if company:
        # અહીં ખાતરી કરો કે ઓનર તેના સાચા ડેશબોર્ડ પર જાય છે
            return redirect('company_owner_dashboard', company_slug=company.slug)

    # C. Company Admin Check (NEW) 🛡️
    
    if user.groups.filter(name__icontains='_admin').exists():
        # 🛑 ભૂલ અહીં હતી: તમે 'company_admin_dashboard' પર મોકલતા હતા (જે સુપર એડમિન માટે છે)
        # ✅ સાચો રસ્તો: 'admin_dashboard' (જે કંપની એડમિન માટે છે)
        return redirect('admin_dashboard')

    # D. HR Check
    if login_profile.role == 'HR' or user.groups.filter(name='HR').exists():
        return redirect('hr_dashboard_view')

    # E. Manager Check
    if login_profile.role == 'MANAGER' or user.groups.filter(name='Manager').exists():
        return redirect('manager_dashboard')

    # F. Team Leader Check (Your Logic)
    is_team_leader = request.user.groups.filter(name__icontains="team leader").exists()
    if is_team_leader:
        return redirect('tl_dashboard') # અથવા team_leader_dashboard જે URL name હોય તે

    if company:
        
        return redirect('employee_dashboard', company_slug=company.slug)
    # ---------------------------------------------------------
    # ⬇️ EMPLOYEE DASHBOARD LOGIC (YOUR ORIGINAL CODE) ⬇️
    # (જો ઉપરના કોઈ રોલ ના હોય, તો જ આ કોડ ચાલશે)
    # ---------------------------------------------------------

    # 3. Decide Target User (કોનો ડેટા જોવો છે?)
    user_id = request.GET.get("user")
    if user_id:
        target_user = get_object_or_404(User, id=user_id)
    else:
        target_user = request.user

    target_profile = target_user.profile

    # 4. Security / Permission Rules
    # જો એમ્પ્લોઈ બીજા કોઈનો ડેટા જોવા માંગતો હોય તો રોકો
    if login_profile.role == "EMPLOYEE" and target_user != request.user:
        return HttpResponse("Access Denied")

    # જો HR બીજી કંપનીના માણસને જોવા માંગતો હોય તો રોકો
    if login_profile.role == "HR" and target_profile.company != login_profile.company:
        return HttpResponse("Access Denied")

    # 5. Policy & Shift Setup
    policy = AttendancePolicy.objects.filter(company=target_profile.company).first()
    shift = target_profile.shift

    # 6. Fetch Attendance Records & Calculation (Your Logic)
    attendance_list = Attendance.objects.filter(user=target_user).order_by('-date')

    records = []

    for att in attendance_list:
        status = "Absent"
        check_in_dt = None
        check_out_dt = None
        work_hours = None
        late_minutes = 0

        # --- FIX START: Direct Use of DateTimeField ---
        if att.check_in:
            # ✅ સુરક્ષિત રીતે સમયને તારીખ સાથે જોડો
            c_in = att.check_in if isinstance(att.check_in, time) else att.check_in.time()
            check_in_dt = timezone.make_aware(datetime.combine(att.date, c_in))

            if att.check_out:
                c_out = att.check_out if isinstance(att.check_out, time) else att.check_out.time()
                check_out_dt = timezone.make_aware(datetime.combine(att.date, c_out))
                
                # કલાકો ગણાશે
                work_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
                
            
            # --- Late Calculation ---
            if shift and shift.start_time:
                current_date = att.date 
                shift_start_naive = datetime.combine(current_date, shift.start_time)
                
                # Timezone Aware બનાવો
                shift_in = timezone.make_aware(shift_start_naive, timezone.get_current_timezone())
                
                # Late check logic
                if check_in_dt > shift_in:
                    late_minutes = (check_in_dt - shift_in).total_seconds() / 60

            # --- Status Logic ---
            if not att.check_out:
                status = "Working"
            elif policy and work_hours is not None:
                if work_hours < (policy.work_hours_required / 2):
                    status = "Half Day"
                elif late_minutes > policy.late_after_minutes:
                    status = "Late"
                elif work_hours >= policy.work_hours_required:
                    status = "Present"
                else:
                    status = "Partial"
            else:
                status = "Present"
        # --- FIX END ---

        records.append({
            "date": att.date,
            "check_in": check_in_dt,   
            "check_out": check_out_dt, 
            "status": status
        })

    return render(request, 'attendance/dashboard/dashboard.html', {
        "records": records,
        "target": target_user
    })

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.contrib.auth.models import Group
from .models import Attendance, AttendancePolicy
from accounts.models import Profile

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.contrib.auth.models import Group, User
from .models import Attendance, AttendancePolicy
from accounts.models import Profile

@login_required
def admin_dashboard_view(request):
    user = request.user

    # 1. SECURITY CHECK
    is_admin = user.groups.filter(name__icontains='admin').exists()
    if not (user.is_superuser or is_admin):
        return HttpResponseForbidden("Access Denied.")

    # 2. SETUP PROFILE & DB
    try:
        profile = user.profile
        company = profile.company
        admin_branch = profile.branch 
        db_name = f"{company.slug}_db" if company else 'default'
        
        from attendance.middleware import ThreadLocal
        ThreadLocal.DB_NAME = db_name
    except Exception as e:
        return HttpResponse(f"Profile Error: {e}")

    today = timezone.localdate()
    # Fetch roles for the dropdown (Excluding the owner role)
    roles = Group.objects.using('default').filter(name__startswith=f"{company.slug}_").exclude(name__icontains='owner')
    # 3. FETCH STAFF PROFILES
    base_profiles = Profile.objects.using('default').filter(company=company, branch=admin_branch)
    base_profiles = base_profiles.exclude(user__groups__name__icontains="owner")

    # 📊 4. STATS COUNTING
    total_admin = base_profiles.filter(user__groups__name__icontains="admin").count()
    total_hr = base_profiles.filter(user__groups__name__icontains="hr").count()
    total_managers = base_profiles.filter(user__groups__name__icontains="manager").count()
    total_team_leaders = base_profiles.filter(user__groups__name__icontains="team leader").count()
    
    total_employees = base_profiles.exclude(
        user__groups__name__iregex=r'(admin|hr|manager|team leader|owner)'
    ).count()

    # 📅 5. ATTENDANCE LOGIC (ERROR FIXED HERE)
    # 🔥 અહીં લિસ્ટમાં કન્વર્ટ કરવું જરૂરી છે જેથી Cross-DB એરર ના આવે
    filtered_user_ids = list(base_profiles.values_list('user_id', flat=True))

    attendance_query = Attendance.objects.using(db_name).filter(
        date=today,
        status__in=["Present", "Working", "Late", "Half Day", "Partial"], 
        user_id__in=filtered_user_ids  # 👈 આ લિસ્ટ હવે સેફ છે
    )
        
    present_today = attendance_query.count()
    today_record = Attendance.objects.using(db_name).filter(user_id=user.id, date=today).first()

    # 6. EMPLOYEES LIST
    employees_list = base_profiles.select_related('user', 'branch').prefetch_related('user__groups').order_by('-user__date_joined')
    
    slug = company.slug.lower() if company else ""
    for emp in employees_list:
        emp.display_role = "Employee"
        for g in emp.user.groups.all():
            if slug in g.name.lower() and not 'owner' in g.name.lower():
                parts = g.name.split('_')
                emp.display_role = parts[-1].title() if len(parts) > 1 else g.name.title()
                break

    # Permissions & Alerts
    role_perms = None
    for g in user.groups.all():
        try:
            role_perms = RolePermission.objects.using(db_name).get(group__name=g.name)
            if role_perms: break
        except RolePermission.DoesNotExist: continue
            
    # base_alerts = Notification.objects.using(db_name).filter(user=user).order_by('-created_at')
    # unread_count = base_alerts.filter(is_read=False).count()
    target_user = request.user
    base_alerts = list(Notification.objects.using(db_name).filter(
        user_id=target_user.id,
        is_read=False
    ).order_by('-created_at')[:10])
    
    # Get the count of unread notifications
    unread_count = Notification.objects.using(db_name).filter(
        user_id=target_user.id, 
        is_read=False
    ).count()

    # Reset ThreadLocal
    from attendance.middleware import ThreadLocal
    ThreadLocal.DB_NAME = 'default'

    context = {
        "total_admin": total_admin,
        "total_hr": total_hr,
        "total_managers": total_managers,
        "total_team_leaders": total_team_leaders,
        "total_employees": total_employees,
        "total_staff": base_profiles.count(),
        "present_today": present_today,
        "today": today,
        "roles": roles,
        "today_record": today_record,
        "my_alerts": base_alerts[:10],
        "unread_count": unread_count,
        "admin_name": user.username,  
        "company": company, 
        "current_branch": admin_branch,
        "employees": employees_list, 
        "role_perms": role_perms,  
    }

    return render(request, "attendance/dashboard/admin_dashboard.html", context)
    
    # attendance/views.py માં આ કોડ ઉમેરો અથવા રિપ્લેસ કરો

# attendance/views.py

# attendance/views.py

# attendance/views.py

# attendance/views.py

# attendance/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.contrib.auth.models import Group
from .models import Notification, Attendance
from accounts.models import Profile

from .middleware import ThreadLocal # Pehla aa check kari lo ke import che ke nahi

@login_required
def manager_dashboard_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Multi-tenant Logic)
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન DB માંથી પ્રોફાઈલ લોડ કરો
        profile = Profile.objects.using('default').get(user=request.user)
        company = profile.company
        manager_branch = profile.branch
        
        if company:
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name # Queries during request will use tenant DB
        else:
            ThreadLocal.DB_NAME = 'default'
    except Profile.DoesNotExist:
        return redirect('login')

    # 1. SECURITY & PERMISSIONS
    user_group = request.user.groups.first()
    role_perms = getattr(user_group, 'role_permissions', None)
    if not role_perms or not role_perms.can_self_access:
        raise PermissionDenied("Access Denied.")

    # ---------------------------------------------------------
    # 2. FETCH TEAM MEMBERS (From Main DB to avoid Join Errors)
    # ---------------------------------------------------------
    # લિસ્ટિંગ હંમેશા 'default' DB થી કરો જેથી Groups/Users સાથે ફિલ્ટરિંગ કામ કરે
    team_profiles_qs = Profile.objects.using('default').filter(
        manager=request.user,
        company=company,
        branch=manager_branch 
    ).select_related("user").order_by('user__username')

    # ---------------------------------------------------------
    # 3. DATA PACKING & ROLE CHECK
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    tl_group_name = f"{slug_lower}_team leader"
    
    team_data = []
    team_users_ids = []
    total_team_leaders = 0

    for member in team_profiles_qs:
        uid = member.user.id
        team_users_ids.append(uid)
        
        # એમ્પ્લોઈનો રોલ ચેક કરો (Main DB Group check)
        is_tl = member.user.groups.filter(name__iexact=tl_group_name).exists()
        if is_tl:
            total_team_leaders += 1

        # Squad Members: આ યુઝરની નીચે કેટલા લોકો છે? (Tenant DB Context)
        squad_members = Profile.objects.using(db_name).filter(
            team_leader_id=uid,
            manager=request.user
        ).select_related('user')

        team_data.append({
            'user': member.user,
            'role': "Team Leader" if is_tl else "Employee",
            'is_team_leader': is_tl,
            'squad_list': squad_members,
            'user_id': uid 
        })

    # 📊 4. STATS CALCULATION (From Tenant DB)
    today = timezone.localdate()
    present_today = Attendance.objects.using(db_name).filter(
        user_id__in=team_users_ids,
        date=today,
        check_in__isnull=False
    ).count()

    # ---------------------------------------------------------
    # 5. FINAL CONTEXT
    # ---------------------------------------------------------
    # શિફ્ટ/ટીમ લીડર રોલ ચેક
    team_leader_role_exists = Group.objects.using('default').filter(name__iexact=tl_group_name).exists()

    return render(request, "attendance/dashboard/manager_dashboard.html", {
        "total_team": len(team_users_ids),
        "present_today": present_today,
        "total_team_leaders": total_team_leaders,
        "company": company,
        "profile": profile,
        "team": team_data, 
        "team_leader_role_exists": team_leader_role_exists,
        "role_perms": role_perms,
        "today_record": Attendance.objects.using(db_name).filter(user=request.user, date=today).first(),
        "unread_count": Notification.objects.using(db_name).filter(user=request.user, is_read=False).count(),
        "notes": Notification.objects.using(db_name).filter(user=request.user).order_by('-created_at')[:10],
    })

from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import AttendancePolicy, Company

@login_required
def company_delete_policy(request, policy_id=None, id=None):
    # Handle the fact that your URLs sometimes pass 'id' and sometimes 'policy_id'
    target_id = policy_id or id
    
    company = None
    db_name = 'default'
    
    # 1. Determine company from the logged-in user (For Owners/Admins/HR)
    try:
        if hasattr(request.user, 'profile') and request.user.profile.company:
            company = request.user.profile.company
            db_name = f"{company.slug}_db"
    except Exception:
        pass

    policy = None
    
    # 2. Look in the exact Company DB
    if company:
        policy = AttendancePolicy.objects.using(db_name).filter(id=target_id).first()
        
    # 3. If not found (and user is a Superadmin), scan all databases to find it
    if not policy and request.user.is_superuser:
        for c in Company.objects.using('default').all():
            temp_db = f"{c.slug}_db"
            try:
                policy = AttendancePolicy.objects.using(temp_db).filter(id=target_id).first()
                if policy:
                    db_name = temp_db
                    company = c
                    break
            except Exception:
                continue

    # 4. Handle if it's completely missing
    if not policy:
        messages.error(request, "Error: Policy not found. It may have already been deleted.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # 5. Safely Delete from the correct database
    try:
        policy.delete(using=db_name)
        messages.success(request, "Attendance policy deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting policy: {str(e)}")
        
    # Redirect exactly back to where the user clicked the button
    return redirect(request.META.get('HTTP_REFERER', '/'))
# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal
from datetime import datetime, time

from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
# Make sure your imports for ThreadLocal, Task, Company, Profile, Attendance etc. are here

@login_required
def employee_dashboard_view(request, company_slug):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING & SETUP
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    try:
        # પ્રોફાઈલ હંમેશા default DB માંથી લાવવી (Truth Source)
        profile = Profile.objects.using('default').select_related('company', 'branch', 'team_leader').get(user=request.user)
    except Profile.DoesNotExist:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("તમારી પ્રોફાઈલ સિસ્ટમમાં નથી.")

    # 🛡️ CURRENT TL ID મેળવો
    current_tl_id = profile.team_leader_id

    # 1. SECURITY CHECK
    if not request.user.is_superuser and profile.company != company:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 🔥 2. TASK LOGIC (CURRENT TL FILTER ONLY)
    # ---------------------------------------------------------
    today = timezone.localdate()
    target_user = request.user

    # A) આ યુઝરને અસાઇન થયા હોય
    # B) જે આપનાર (assigned_by) અત્યારે એમ્પ્લોઈનો સાચો Team Leader હોય
    task_filter = Q(assigned_to_id=target_user.id)
    if current_tl_id:
        task_filter &= Q(assigned_by_id=current_tl_id)

    # પેન્ડિંગ ટાસ્ક્સ (Force evaluation with list())
    raw_tasks = list(Task.objects.using(db_name).filter(
        task_filter,
        status__in=['Pending', 'In Progress', 'Rejected']
    ).order_by('-created_at'))

    active_tasks = []
    for t_obj in raw_tasks:
        team_members = Task.objects.using(db_name).filter(title=t_obj.title, assigned_by_id=t_obj.assigned_by_id)
        active_tasks.append({
            'obj': t_obj,
            'is_team_task': team_members.count() > 1,
            'team_names': [m.assigned_to.username.title() for m in team_members]
        })

    # 🚨 CRITICAL FIX HERE 🚨
    # કમ્પ્લીટ થયેલા ટાસ્ક્સ (Prefetch attachments & force query before DB reset)
    completed_tasks = list(Task.objects.using(db_name).filter(
        task_filter,
        status__in=['Completed', 'Verified']
    ).prefetch_related('attachments').order_by('-submitted_at'))

    # ---------------------------------------------------------
    # 3. ATTENDANCE & OTHER LOGIC (As it is)
    # ---------------------------------------------------------
    attendance = Attendance.objects.using(db_name).filter(user_id=target_user.id, date=today).first()
    
    my_alerts = list(Notification.objects.using(db_name).filter(
        user_id=target_user.id,
        is_read=False
    ).order_by('-created_at')[:10])
    
    # Get the count of unread notifications
    unread_count = Notification.objects.using(db_name).filter(
        user_id=target_user.id, 
        is_read=False
    ).count()

    # Safety reset happens AFTER all data is safely loaded into memory
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/dashboard/employee_dashboard.html", {
        "attendance": attendance,
        "profile": profile,
        "company": company,
        "completed_tasks": completed_tasks,
        "active_tasks": active_tasks,
        "is_team_leader": request.user.groups.filter(name__icontains="team leader").exists(),
        "current_branch": profile.branch,
        "my_alerts": my_alerts,
        "unread_count": unread_count,
    })    
    # attendance/views.py
# attendance/views.py
from django.utils import timezone

# views.py
# views.py
# attendance/views.py
from .middleware import ThreadLocal # Import check kari lejo

@login_required
def update_task_status(request, task_id):
    if request.method == "POST":
        # ---------------------------------------------------------
        # 0. DATABASE SWITCHING (Level 0 Logic)
        # ---------------------------------------------------------
        try:
            profile = request.user.profile
            company = profile.company
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
            print(f"DEBUG: Task update switching to {db_name}")
        except:
            ThreadLocal.DB_NAME = 'default'
            db_name = 'default'

        # ૧. ટાસ્ક શોધો (Sacha DB mathi)
        task = get_object_or_404(Task.objects.using(db_name), id=task_id, assigned_to=request.user)
        
        # 🔥 ૨. CLEANUP LOGIC (Sacha DB mathi delete thase)
        if task.status == 'Rejected' or task.status == 'Completed':
            # TaskAttachment delete karo
            task.attachments.all().delete() # using implicit connection
            
            # File cleanup
            if task.completion_file:
                task.completion_file.delete(save=False)
                task.completion_file = None

        # --- ૩. નવી ફાઇલો સેવ કરવાનું લોજિક (Using db_name) ---
        files = request.FILES.getlist('completion_file')
        if files:
            for f in files:
                # Direct creation in company DB
                TaskAttachment.objects.using(db_name).create(task=task, file=f)
        
        # ૪. ટાસ્ક અપડેટ કરો
        task.status = 'Completed' 
        task.submitted_at = timezone.now() 
        task.rejection_note = None 
        task.completion_note = request.POST.get('completion_note')
        
        # ✅ SAVE PAN DB MA J KARO
        task.save(using=db_name)

        # ૫. નોટિફિકેશન (Company DB ma jase)
        Notification.objects.using(db_name).create(
            user=task.assigned_by, 
            message=f"Employee {request.user.username} has submitted Task : {task.title}."
        )

        messages.success(request, "Task Submitted successfully!")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
    
    return redirect('dashboard')                    
        # attendance/views.py

# 1. Update existing HR Dashboard View

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.contrib.auth.models import Group, User
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives 
from django.conf import settings
from accounts.models import Profile
from attendance.models import Attendance, RolePermission, Notification

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.contrib.auth.models import User, Group
# Ensure you have your Profile, Company, Attendance, Notification, RolePermission, etc. imported here

@login_required
def hr_dashboard_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING & SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા 'default' DB માંથી પ્રોફાઇલ અને બ્રાન્ચ લોડ કરો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        hr_branch = profile.branch  # 👈 આ HR ની બ્રાન્ચ છે (ઓબ્જેક્ટ અથવા None)
        
        if company:
            db_name = f"{company.slug}_db"
            from attendance.middleware import ThreadLocal
            ThreadLocal.DB_NAME = db_name
        else:
            db_name = 'default'
    except Exception:
        return HttpResponse("Profile not found")

    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    if not (is_hr or request.user.is_superuser):
        from attendance.middleware import ThreadLocal
        ThreadLocal.DB_NAME = 'default'
        return HttpResponseForbidden("Access Denied: You do not have HR privileges.")

    today = timezone.localdate()
    slug_lower = company.slug.lower() if company else ""

    # ---------------------------------------------------------
    # 🔥 1. PROXY HACK
    # ---------------------------------------------------------
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            tenant_perms = RolePermission.objects.using(db_name).filter(group=real_group).first()
        except Exception: pass

    if request.user.is_superuser:
        class SuperPerms:
            def __getattr__(self, item): return True
        tenant_perms = SuperPerms()

    class ProxyGroup:
        def __init__(self, group_obj): self._group = group_obj; self.role_permissions = tenant_perms 
        def __getattr__(self, name): return getattr(self._group, name)
    class ProxyGroupsManager:
        def __init__(self, manager): self._manager = manager
        def first(self): obj = self._manager.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._manager, name)
    class ProxyUser:
        def __init__(self, user_obj): self._user = user_obj; self.groups = ProxyGroupsManager(user_obj.groups)
        def __getattr__(self, name): return getattr(self._user, name)
    class ProxyRequest:
        def __init__(self, req): self._req = req; self.user = ProxyUser(req.user)
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 2. FETCH STAFF (Branch Specific Filter)
    # ---------------------------------------------------------
    # 🔥 તમારી મેઈન કન્ડિશન: HR ની બ્રાન્ચ મુજબ જ સ્ટાફ ફિલ્ટર થશે
    base_profiles = Profile.objects.using('default').filter(company=company)
    
    if hr_branch:
        # જો HR બ્રાન્ચમાં હોય, તો માત્ર તે જ બ્રાન્ચનો ડેટા
        base_profiles = base_profiles.filter(branch=hr_branch)
    else:
        # જો HR બ્રાન્ચમાં ના હોય, તો એવા લોકો જે કોઈ બ્રાન્ચમાં નથી (Head Office)
        base_profiles = base_profiles.filter(branch__isnull=True) 
        
    # Exclude Owners and Admins
    base_profiles = base_profiles.exclude(user__groups__name__icontains='owner').exclude(user__groups__name__icontains='admin')
    
    employees_list = base_profiles.select_related('user', 'branch').prefetch_related('user__groups').order_by('user__first_name', 'user__username')
    staff_user_ids = [emp.user_id for emp in employees_list]

    # Attendance Fetch (Tenant DB માંથી)
    today_attendance = Attendance.objects.using(db_name).filter(
        date=today, user_id__in=staff_user_ids
    )
    att_map = {att.user_id: att for att in today_attendance}

    # ---------------------------------------------------------
    # 📊 3. THE ALL-IN-ONE CALCULATION
    # ---------------------------------------------------------
    counts = {
        'total_staff': 0,
        'employees': 0,
        'managers': 0,
        'team_leaders': 0,
        'present_today': 0
    }

    for emp in employees_list:
        counts['total_staff'] += 1
        
        # --- Role Logic ---
        user_groups = [g.name.lower() for g in emp.user.groups.all()]
        emp.display_role = "Employee" 
        
        if any(f"{slug_lower}_manager" in g for g in user_groups):
            counts['managers'] += 1
            emp.display_role = "Manager"
        elif any(f"{slug_lower}_team leader" in g for g in user_groups):
            counts['team_leaders'] += 1
            emp.display_role = "Team Leader"
        else:
            counts['employees'] += 1

        # --- Attendance Status ---
        record = att_map.get(emp.user_id)
        status = "Absent"
        if record:
            if record.status and record.status.strip().lower() not in ["", "absent", "none"]:
                status = record.status.title()
            else:
                if record.check_in and not record.check_out: status = "Working"
                elif record.check_in and record.check_out: status = "Present"
                
        if status in ["Present", "Working", "Late", "Half Day", "Partial"]:
            counts['present_today'] += 1

    absent_today = counts['total_staff'] - counts['present_today']

    # ---------------------------------------------------------
    # 📧 4. EMAIL RECIPIENTS (Branch Specific)
    # ---------------------------------------------------------
    # ઓનર અને એડમિન જે તે જ બ્રાન્ચમાં હોય તેમને જ મેઈલ જાય
    owner_qs = User.objects.using('default').filter(profile__company=company, groups__name__icontains='owner')
    admin_qs = User.objects.using('default').filter(profile__company=company, profile__branch=hr_branch, groups__name__icontains='admin')
    report_recipients = (owner_qs | admin_qs).distinct().exclude(email="")

    # (અહીં તમે તમારું EmailDispatch Logic મૂકી શકો છો જે મેં અગાઉ આપેલું છે)

    # ---------------------------------------------------------
    # 🏁 5. RENDER DASHBOARD
    # ---------------------------------------------------------
    today_record = Attendance.objects.using(db_name).filter(user_id=request.user.id, date=today).first()
    
    # 🔥 CRITICAL FIX: Evaluate notifications into memory BEFORE switching DB back to default
    my_alerts = list(Notification.objects.using(db_name).filter(
        user_id=request.user.id,
        is_read=False
    ).order_by('-created_at')[:10])
    
    unread_count = Notification.objects.using(db_name).filter(
        user_id=request.user.id, 
        is_read=False
    ).count()

    team_leader_role_exists = Group.objects.using('default').filter(name__iexact=f"{company.slug}_Team Leader").exists()

    # નોટિફિકેશન અને ડેટા મેમરીમાં લોડ થઈ ગયા પછી જ DB રિસેટ કરો
    from attendance.middleware import ThreadLocal
    ThreadLocal.DB_NAME = 'default'
    
    context = {
        "request": ProxyRequest(request),  
        "company": company,
        
        # આ આંકડા હવે બ્રાન્ચ ફિલ્ટર મુજબ હશે
        "total_staff": counts['total_staff'],
        "total_employees": counts['employees'],
        "total_managers": counts['managers'],
        "total_team_leaders": counts['team_leaders'],
        "present_today": counts['present_today'],
        "absent_today": absent_today,
        
        "team_leader_role_exists": team_leader_role_exists,
        "employees": employees_list,
        "today_record": today_record,
        
        # 🔥 Pass the forced memory variables
        "my_alerts": my_alerts,       
        "unread_count": unread_count,
        
        "current_branch": hr_branch,
    }

    return render(request, "attendance/dashboard/hr_dashboard.html", context)
    # 2. Create NEW View for TL List
# attendance/views.py

# attendance/views.py

# attendance/views.py


from django.db.models import Q  # આ લાઈન ફાઈલની શરૂઆતમાં હોવી જોઈએ

# @login_required
# def hr_team_leader_list(request):
#     # 1. Security Check
#     if not (request.user.groups.filter(name__icontains='hr').exists() or request.user.is_superuser):
#         return HttpResponse("Access Denied")

#     # 2. Context Setup
#     try:
#         profile = request.user.profile
#         company = profile.company
#         hr_branch = profile.branch  # 👈 HR ની બ્રાન્ચ
#     except:
#         return HttpResponse("Profile Error")

#     # ---------------------------------------------------------
#     # 3. Get Team Leaders
#     # ---------------------------------------------------------
#     # TL શોધવા માટે આપણે Group Check કરીશું
#     tls_profile = Profile.objects.filter(
#         company=company,
#         user__groups__name__icontains="team leader"
#     ).select_related('user', 'manager').distinct()

#     # જો HR ની બ્રાન્ચ હોય, તો ફક્ત તે બ્રાન્ચના TL જ બતાવો
#     if hr_branch:
#         tls_profile = tls_profile.filter(branch=hr_branch)

#     # ---------------------------------------------------------
#     # 4. Get Available Staff (THE GOLDEN LOGIC 🔥)
#     # ---------------------------------------------------------
    
#     # Step A: પાયાનું ફિલ્ટર (કંપની + એક્ટિવ યુઝર + જેની પાસે TL નથી)
#     base_staff_qs = Profile.objects.filter(
#         company=company,
#         team_leader__isnull=True,  # જે ફ્રી છે (TL નથી)
#         user__is_active=True       # જે એકાઉન્ટ ચાલુ છે
#     )

#     # Step B: બ્રાન્ચ લોજિક (સૌથી મહત્વનું)
#     # જો HR પાસે બ્રાન્ચ છે, તો એવા લોકો બતાવો જેની બ્રાન્ચ HR જેવી છે અથવા ખાલી (NULL) છે.
#     if hr_branch:
#         base_staff_qs = base_staff_qs.filter(
#             Q(branch=hr_branch) | Q(branch__isnull=True)
#         )

#     # Step C: બોસ લોકોને લિસ્ટમાંથી કાઢો (Negative Filtering)
#     # આપણે 'Employee' ગ્રુપ શોધતા નથી, પણ બીજા ગ્રુપને કાઢીએ છીએ.
#     # આનાથી જેનું ગ્રુપ નથી આપ્યું તે પણ લિસ્ટમાં આવી જશે.
    
#     base_staff_qs = base_staff_qs.exclude(
#         user__groups__name__icontains="admin"
#     ).exclude(
#         user__groups__name__icontains="hr"
#     ).exclude(
#         user__groups__name__icontains="manager"
#     ).exclude(
#         user__groups__name__icontains="team leader"
#     ).exclude(
#         user__is_superuser=True
#     ).select_related('user').distinct()

#     # ---------------------------------------------------------
#     # 5. Pack Data
#     # ---------------------------------------------------------
#     tl_data = []
    
#     for tl in tls_profile:
#         # A. Current Squad (જે ઓલરેડી અસાઇન છે)
#         squad_members = Profile.objects.filter(team_leader=tl.user).select_related('user')
        
#         # B. Available Staff
#         # TL પોતે લિસ્ટમાં ના આવવો જોઈએ
#         my_available_staff = base_staff_qs.exclude(user=tl.user)

#         tl_data.append({
#             'user': tl.user,
#             'manager': tl.manager,
#             'squad_count': squad_members.count(),
#             'squad_list': squad_members,
#             'available_staff': my_available_staff, # ✅ 100% Data આવશે
#             'id': tl.user.id
#         })

#     return render(request, "attendance/hr/hr_tl_list.html", {
#         "team_leaders": tl_data,
#         "company": company
#     })
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from accounts.models import Profile

@login_required
def hr_employee_list(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા 'default' DB માંથી પ્રોફાઇલ લોડ કરો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        hr_branch = profile.branch
        db_name = f"{company.slug}_db" if company else 'default'
    except Profile.DoesNotExist:
        return HttpResponse("Profile not found")

    # 1. SECURITY CHECK
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    if not (is_hr or is_admin or request.user.is_superuser):
        raise PermissionDenied("You do not have permission to manage employees.")

    # 🔥 2. THE PROXY HACK (HTML Permissions માટે)
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            from attendance.models import RolePermission
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except Exception: pass

    class ProxyGroup:
        def __init__(self, obj): self._obj = obj; self.role_permissions = tenant_perms
        def __getattr__(self, name): return getattr(self._obj, name)
    class ProxyManager:
        def __init__(self, m): self._m = m
        def first(self): obj = self._m.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._m, name)
    class ProxyUser:
        def __init__(self, u): self._user = u; self.groups = ProxyManager(u.groups)
        def __getattr__(self, name): return getattr(self._user, name)
    class ProxyRequest:
        def __init__(self, r): self._req = r; self.user = ProxyUser(r.user)
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 3. FETCH EMPLOYEES & TEAM LEADERS FROM MAIN DB
    # ---------------------------------------------------------
    slug_lower = company.slug.lower() if company else ""
    
    # 🔥 THE MAGIC FILTER: કંપની અને બ્રાન્ચ બંને એકસાથે જ ફિલ્ટર કરી દીધા
    base_profiles = Profile.objects.using('default').filter(
        company=company,
        branch=hr_branch # જો HR બ્રાન્ચમાં હોય તો બ્રાન્ચ, નહીંતર None (Head Office)
    )
    
    # 🔥 CHANGE: 'team leader' ને exclude માંથી હટાવી દીધું છે
    # હવે લિસ્ટમાં સામાન્ય Employee અને Team Leader બંને આવશે.
    # માત્ર Admin, HR, Manager અને Owner જ લિસ્ટમાં નહીં દેખાય.
    employees_qs = base_profiles.exclude(
        user__groups__name__iregex=fr'({slug_lower}_admin|{slug_lower}_hr|{slug_lower}_manager|{slug_lower}_owner)'
    ).select_related("user", "branch", "manager", "shift").order_by('-user__date_joined')

    # ---------------------------------------------------------
    # 🛠️ 4. ROLE IDENTIFICATION (Optional: HTML માં રોલ બતાવવા માટે)
    # ---------------------------------------------------------
    employees = list(employees_qs)
    for emp in employees:
        # ચેક કરો કે આ યુઝર Team Leader ગ્રુપમાં છે?
        is_tl = emp.user.groups.filter(name__icontains='team leader').exists()
        emp.is_team_leader_role = is_tl  # આ વેલ્યુ તમે HTML માં વાપરી શકશો

    return render(request, "attendance/hr/hr_employee_list.html", {
        "request": ProxyRequest(request),
        "employees": employees,
        "company": company,
        "current_branch": hr_branch
    })
    
    
from .middleware import ThreadLocal # Import hamesha check kari lejo

@login_required
def hr_pending_users(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        hr_branch = profile.branch  # HR ni branch
        
        if company:
            # Have have pachi ni badhi queries aa Company na DB ma jase
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
            print(f"DEBUG: Pending users switching to {db_name}")
        else:
            ThreadLocal.DB_NAME = 'default'
            
    except Profile.DoesNotExist:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile not found")

    # 1. PERMISSION CHECK (Have queries sacha DB ma jase)
    if not user_has_permission(request.user, "can_approve_attendance"):
        return HttpResponse("Access Denied")
    if not user_has_permission(request.user, "can_manage_users"):
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 2. DATA FETCHING (Correct Logic)
    # ---------------------------------------------------------
    # 🔥 FIX: is_approved=False hovu joie jethi 'Pending' users male
    # 🔥 FIX: branch=hr_branch add karyu jethi HR potana branch na j pending users joi shake
    
    users = Profile.objects.filter(
        is_verified=True,
        company=company,
        is_approved=False, # ✅ Pending users mate False hovu joie
        branch=hr_branch   # ✅ HR na potana branch na users
    ).select_related('user')

    return render(request, "attendance/hr/hr_pending_users.html", {
        "users": users,
        "company": company,
        "current_branch": hr_branch
    })

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def hr_approve_user(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_current = request.user.profile
        company = profile_current.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. PERMISSION CHECK
    if not user_has_permission(request.user, "can_approve_attendance") or \
       not user_has_permission(request.user, "can_manage_users"):
        return HttpResponse("Access Denied")

    # ૨. પ્રોફાઈલ મેળવો (Sacha Company DB mathi)
    profile = get_object_or_404(Profile.objects.using(db_name), user__id=user_id)

    # ૩. અપ્રૂવ કરો
    profile.is_approved = True
    profile.save(using=db_name) # ✅ SAVE PAN DB MA J KARO

    messages.success(request, f"User {profile.user.username} approved successfully.")
    return redirect("hr_pending_users")


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import User
from accounts.models import Profile

@login_required
def hr_assign_manager(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        # લોગિન થયેલ HR ની પ્રોફાઇલ મેઈન DB માંથી લાવો
        login_profile = Profile.objects.using('default').get(user=request.user)
        company = login_profile.company
        db_name = f"{company.slug}_db"
    except Profile.DoesNotExist:
        return HttpResponse("HR Profile not found")

    # Security Check
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    if not (request.user.is_superuser or is_hr):
        return HttpResponse("Access Denied")
    
    # ૧. ટાર્ગેટ એમ્પ્લોઈ મેઈન ડીબી માંથી લાવો
    employee_user = get_object_or_404(User.objects.using('default'), id=user_id)
    
    # ---------------------------------------------------------
    # 📋 ૨. FETCH MANAGERS (The Bulletproof Way)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    target_manager_group = f"{slug_lower}_manager"

    # મેઈન ડીબી માંથી આ કંપનીના મેનેજરો શોધો (જેથી ડ્રોપડાઉન ક્યારેય ખાલી ના આવે)
    managers = User.objects.using('default').filter(
        profile__company=company,
        groups__name__iexact=target_manager_group
    ).select_related('profile')

    # જો એમ્પ્લોઈ કોઈ બ્રાન્ચમાં હોય, તો મેનેજરોને પણ તે જ બ્રાન્ચના બતાવો
    employee_main_profile = Profile.objects.using('default').get(user_id=user_id)
    if employee_main_profile.branch_id:
        managers = managers.filter(profile__branch_id=employee_main_profile.branch_id)

    # ---------------------------------------------------------
    # 💾 ૩. POST HANDLING (Dual-DB Sync)
    # ---------------------------------------------------------
    if request.method == "POST":
        selected_manager_id = request.POST.get("manager")
        
        manager_user_obj = None
        if selected_manager_id:
            manager_user_obj = User.objects.using('default').get(id=selected_manager_id)

        try:
            # 🔥 STEP A: મેઈન ડેટાબેઝ (default) માં અપડેટ કરો
            # .update() વાપરવાથી ડાયરેક્ટ ID સેવ થશે અને Join Error નહીં આવે
            Profile.objects.using('default').filter(user_id=user_id).update(manager=manager_user_obj)

            # 🔥 STEP B: ટેનન્ટ ડેટાબેઝ (Company DB) માં અપડેટ કરો
            try:
                Profile.objects.using(db_name).filter(user_id=user_id).update(manager=manager_user_obj)
            except Exception:
                pass # જો કંપની ડીબીમાં હજુ પ્રોફાઈલ ના બની હોય તો વાંધો નહીં

            if manager_user_obj:
                messages.success(request, f"Manager {manager_user_obj.username} assigned successfully.")
            else:
                messages.warning(request, "Manager removed successfully.")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            
        return redirect('hr_employee_list')

    return render(request, "attendance/hr/hr_assign_manager.html", {
        "employee": employee_main_profile,
        "managers": managers
    })        
             
from .middleware import ThreadLocal # Pehla aa check kari lo ke import che ke nahi

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from accounts.models import Profile

@login_required
def hr_attendance_report(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન ડીબી માંથી પ્રોફાઈલ લાવો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        hr_branch = profile.branch
        db_name = f"{company.slug}_db" if company else 'default'
    except Profile.DoesNotExist:
        return HttpResponse("Profile not found")

    # ---------------------------------------------------------
    # 🛡️ 1. ULTIMATE SECURITY CHECK (FIX FOR 403 ERROR)
    # ---------------------------------------------------------
    # ડાયરેક્ટ ગ્રુપ ચેક અને કસ્ટમ પરમિશન ચેક બંનેને ક્લબ કરો
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    
    # જો યુઝર HR/Admin હોય અથવા તેની પાસે 'can_manage_users' પરમિશન હોય (કોઈપણ એક સાચું હોય તો ચાલે)
    has_perm = is_hr or is_admin or request.user.is_superuser or user_has_custom_permission(request.user, 'can_manage_users')
    
    if not has_perm:
        raise PermissionDenied("You do not have permission to view attendance reports.")

    # ---------------------------------------------------------
    # 🔥 2. THE PROXY HACK (HTML બટનો માટે)
    # ---------------------------------------------------------
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            from .models import RolePermission
            # કંપનીના ડીબી માંથી જ પરમિશન લાવો
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except Exception: pass

    if request.user.is_superuser:
        class SuperPerms:
            def __getattr__(self, item): return True
        tenant_perms = SuperPerms()

    class ProxyGroup:
        def __init__(self, obj): self._obj = obj; self.role_permissions = tenant_perms
        def __getattr__(self, name): return getattr(self._obj, name)
    class ProxyManager:
        def __init__(self, m): self._m = m
        def first(self): obj = self._m.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._m, name)
    class ProxyUser:
        def __init__(self, u): self._user = u; self.groups = ProxyManager(u.groups)
        def __getattr__(self, name): return getattr(self._user, name)
    class ProxyRequest:
        def __init__(self, r): self._req = r; self.user = ProxyUser(r.user)
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 3. FETCH EMPLOYEES
    # ---------------------------------------------------------
    slug_lower = company.slug.lower() if company else ""
    
    # રિપોર્ટ લિસ્ટિંગ માટે 'default' DB વાપરો (JOIN Error બચાવવા)
    base_profiles = Profile.objects.using('default').filter(company=company)
    
    if hr_branch:
        base_profiles = base_profiles.filter(branch=hr_branch)
        
    # મેનેજમેન્ટ રોલ્સને કાઢી નાખો
    employees = base_profiles.exclude(
        user__groups__name__iregex=fr'({slug_lower}_admin|{slug_lower}_hr|{slug_lower}_manager|{slug_lower}_team leader|{slug_lower}_owner)'
    ).select_related("user", "branch").order_by('user__username')

    return render(request, "attendance/hr/hr_attendance_report.html", {
        "request": ProxyRequest(request), 
        "employees": employees,
        "company": company,
        "current_branch": hr_branch
    })
    
    
def user_has_custom_permission(user, permission_name):
    """
    Checks if a user's role has a specific toggle enabled in RolePermission.
    """
    if user.is_superuser:
        return True
    
    # Get the user's first group (their primary role like 'infosys_hr')
    user_group = user.groups.first()
    if user_group:
        try:
            # Check the related RolePermission record for that group
            perm_record = user_group.role_permissions
            return getattr(perm_record, permission_name, False)
        except RolePermission.DoesNotExist:
            return False
    return False


# attendance/views.py માં manager_team_attendance શોધો અને આ કોડ મૂકો

# attendance/views.py
# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def manager_team_leader_list(request):
    try:
        # 1. Fetch current Manager's Profile and Branch
        profile = Profile.objects.using('default').get(user=request.user)
        company = profile.company
        manager_branch = profile.branch
        db_name = f"{company.slug}_db"
    except Profile.DoesNotExist: 
        return HttpResponse("Profile Error")

    slug_lower = company.slug.lower()
    
    # Exact group names for filtering
    employee_group_name = f"{slug_lower}_employee"
    tl_group_name = f"{slug_lower}_team leader"

    # --- 🔥 THE CRITICAL FILTER FIX ---
    # We filter by:
    # 1. Same Company
    # 2. Same Branch (respecting your branch-isolation rule)
    # 3. Manager MUST be the currently logged-in user
    # 4. Role MUST be 'Employee' group
    # 5. Must NOT have a Team Leader yet
    
    eligible_staff_pool = Profile.objects.using('default').filter(
        company=company,
        branch=manager_branch,
        manager=request.user,             # 👈 ONLY show my assigned employees
        user__groups__name__iexact=employee_group_name,
        team_leader__isnull=True,         # Only free staff
        is_approved=True
    ).exclude(
        user__groups__name__iexact=tl_group_name # Ensure no existing TLs show up
    ).select_related('user').distinct()

    # --- DATA FETCHING FOR TABLE ROWS ---
    
    # Get Team Leaders belonging to this manager/branch
    tls_profile_qs = Profile.objects.using('default').filter(
        company=company,
        branch=manager_branch,
        manager=request.user,
        user__groups__name__iexact=tl_group_name
    )

    tl_data = [] 
    for tl in tls_profile_qs:
        # Fetch squad members from the tenant DB
        squad_members = Profile.objects.using(db_name).filter(team_leader_id=tl.user.id)
        
        tl_data.append({
            'user': tl.user,
            'squad_count': squad_members.count(),
            'squad_list': squad_members,
            'id': tl.user.id
        })

    return render(request, "attendance/MISC/manager_tl_list.html", {
        "team_leaders": tl_data,
        "unassigned_employees": eligible_staff_pool,
        "current_branch": manager_branch 
    })

from .middleware import ThreadLocal # Import check kari lejo

from django.db import transaction

from django.db import transaction
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from accounts.models import Profile

@login_required
def manager_assign_squad(request):
    if request.method == "POST":
        leader_id = request.POST.get('leader_id')
        employee_ids = request.POST.getlist('employee_ids') 
        
        try:
            profile = request.user.profile
            db_name = f"{profile.company.slug}_db"

            # 🔥 THE FIX: માત્ર team_leader_id અપડેટ કરો. 
            # બ્રાન્ચ કે અન્ય કોઈ ફિલ્ડને અહીં લખવા જ નહીં, જેથી તે એમને એમ જ રહે.
            
            # A. મેઈન ડેટાબેઝમાં અપડેટ
            Profile.objects.using('default').filter(
                user_id__in=employee_ids, 
                company=profile.company
            ).update(team_leader_id=leader_id)

            # B. ટેનન્ટ (Company) ડેટાબેઝમાં અપડેટ
            Profile.objects.using(db_name).filter(
                user_id__in=employee_ids
            ).update(team_leader_id=leader_id)
            
            messages.success(request, "Squad members assigned. Branch settings preserved.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            
    return redirect(request.META.get('HTTP_REFERER', 'manager_dashboard'))


from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from accounts.models import Profile
from attendance.middleware import ThreadLocal

@login_required
def manager_revoke_squad(request, member_id):
    try:
        profile = request.user.profile
        db_name = f"{profile.company.slug}_db"
    except: return HttpResponse("Profile Error")

    try:
        # 🔥 THE FIX: .update(team_leader_id=None) વાપરો.
        # આનાથી તે એમ્પ્લોઈ બ્રાન્ચમાં તો રહેશે જ, પણ કોઈ ટીએલ નીચે નહીં હોય.
        
        Profile.objects.using('default').filter(user_id=member_id).update(team_leader_id=None)
        Profile.objects.using(db_name).filter(user_id=member_id).update(team_leader_id=None)

        messages.warning(request, "Member removed from squad. Branch info remains intact.")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', 'manager_dashboard'))


# @login_required
# def manager_team_attendance(request):
#     # ૧. પ્રોફાઈલ અને પરમિશન ચેક
#     profile = get_object_or_404(Profile, user=request.user)
#     # if not request.user.groups.filter(name__icontains='manager').exists():
#     #         return HttpResponse("Access Denied: You don't have permission to view team attendance.")
    
#     if not user_has_permission(request.user, "can_view_team"):
#         return HttpResponse("Access Denied: You don't have permission to view team.")
    
#     # ૨. સાચું ફિલ્ટરિંગ: 'role' ફિલ્ડ કાઢી નાખો
#     # જે લોકોના મેનેજર તમે પોતે છો તેમને ફિલ્ટર કરો
#     team = Profile.objects.filter(
#         manager=request.user,
#         company=profile.company
#     ).select_related("user")

#     return render(request, "attendance/attendance/manager_team_attendance.html", {
#         "team": team
#     })    


# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, time
from accounts.models import Profile
from .models import Attendance
from .middleware import ThreadLocal

@login_required
def manager_team_attendance(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP (Main DB Truth)
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન DB માંથી મેનેજરની વિગતો લાવો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        manager_branch = profile.branch 
        
        if company:
            db_name = f"{company.slug}_db"
            # ThreadLocal સેટ કરો જેથી અંદરના મોડલ મેથડ્સ સાચો DB વાપરે
            ThreadLocal.DB_NAME = db_name
        else:
            db_name = 'default'
            
    except Profile.DoesNotExist:
        return HttpResponse("Profile not found")

    # 1. PERMISSION CHECK
    if not user_has_custom_permission(request.user, 'can_view_team'):
        return render(request, 'attendance/MISC/access_denied.html')
    
    user_group = request.user.groups.first()
    role_perms = getattr(user_group, 'role_permissions', None)
    
    today = timezone.localdate()

    # ---------------------------------------------------------
    # 📋 2. FETCH TEAM MEMBERS (From Main DB to avoid Join Errors)
    # ---------------------------------------------------------
    # આપણે મેઈન ડીબી માંથી તે તમામ યુઝર્સના IDs લાવીશું જે આ મેનેજરના છે
    team_profiles = Profile.objects.using('default').filter(
        manager=request.user,
        company=company,
        branch=manager_branch 
    ).select_related("user")

    team_user_ids = [p.user_id for p in team_profiles]

    # ---------------------------------------------------------
    # 📅 3. FETCH ATTENDANCE FOR TODAY (From Tenant DB)
    # ---------------------------------------------------------
    # એક જ ક્વેરીમાં આખી ટીમની આજની એટેન્ડન્સ લાવો (Performance Optimization)
    attendance_records = Attendance.objects.using(db_name).filter(
        user_id__in=team_user_ids,
        date=today
    )
    
    # ઝડપથી શોધવા માટે ડિક્શનરી બનાવો: {user_id: attendance_object}
    att_dict = {a.user_id: a for a in attendance_records}

    # ---------------------------------------------------------
    # 📦 4. PREPARE DISPLAY DATA
    # ---------------------------------------------------------
    team_data = []

    for member in team_profiles:
        att = att_dict.get(member.user_id)
        
        status = "Absent"
        check_in_time = "-"
        check_out_time = "-"

        if att:
            # Status નક્કી કરો
            if att.status in ['Leave', 'Holiday', 'Week Off']:
                status = att.status
            elif att.check_in:
                status = "Present"
                
            # Format times safely
            if att.check_in:
                if isinstance(att.check_in, (datetime, time)):
                    check_in_time = att.check_in.strftime("%I:%M %p")
                else:
                    check_in_time = str(att.check_in)

            if att.check_out:
                if isinstance(att.check_out, (datetime, time)):
                    check_out_time = att.check_out.strftime("%I:%M %p")
                else:
                    check_out_time = str(att.check_out)

        team_data.append({
            'profile': member,
            'user': member.user,
            'shift': member.shift,
            'status': status,
            'check_in': check_in_time,
            'check_out': check_out_time
        })

    return render(request, "attendance/attendance/manager_team_attendance.html", {
        "team_data": team_data,
        "role_perms": role_perms,
        
        "today": today,
        "current_branch": manager_branch 
    })

    
# @login_required
# def apply_leave_view(request):
#     # if not is_employee(request.user):
#     #     return HttpResponse("only Employee can Apply for leave")
    
#     if request.method == "POST":
#         from_date = request.POST.get("from_date")
#         to_date = request.POST.get("to_date")
#         reason = request.POST.get("reason")
        
#         from datetime import timedelta

#         from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
#         to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

#         today = timezone.localdate()

# # ❌ Past date not allowed
#         if from_date < today or to_date < today:
#             return HttpResponse("Past dates are not allowed for leave")

# # ❌ Invalid date range
#         if from_date > to_date:
#             return HttpResponse("From date cannot be after To date")
        
#         # ❌ Overlapping leave check
#         existing_leave = Leave.objects.filter(
#         user=request.user,
#         status__in=["PENDING", "APPROVED"],
#         from_date__lte=to_date,
#         to_date__gte=from_date
#         ).exists()

#         if existing_leave:
#             messages.error(request,"Leave already applied for selected dates")
#             return redirect("apply_leave")

#         from .models import Holiday

#         current = from_date
#         working_day_found = False

#         while current <= to_date:

#             is_holiday = Holiday.objects.filter(date=current).exists()
#             is_weekend = current.weekday() in [5, 6]

#     # Working day = not holiday & not weekend
#             if not is_holiday and not is_weekend:
#                 working_day_found = True
#                 break

#             current += timedelta(days=1)

# # ❌ If NO working day in selected range
#         if not working_day_found:
#             messages.error(request,"Leave range must include at least one working day")
#             return redirect("apply_leave")
            

#         Leave.objects.create(user=request.user,from_date=from_date,to_date=to_date,reason=reason)

#         employee_profile = Profile.objects.get(user=request.user)
#         manager = employee_profile.manager

#         profile = Profile.objects.get(user=request.user)
#         if profile.role == "HR":
#             admin_profiles = Profile.objects.filter(
#                 role="ADMIN",
#                 company=profile.company
#             )

#             for admin in admin_profiles:
#                 create_notification(
#                     admin.user,
#                     f"HR {request.user.username} applied for leave"
#                 )


# # MANAGER → notify ADMIN
#         if profile.role.upper() == "MANAGER":

#             admin_user = Profile.objects.filter(
#                 company = profile.company,
#                 role = "ADMIN"
#             ).first()

#             if admin_user:
#                 create_notification(
#                     admin_user.user,
#                     f"Manager {request.user.username} applied for leave."
#                 )

# # EMPLOYEE → notify MANAGER (already there)
#         elif profile.role.upper() == "EMPLOYEE":

#             manager = profile.manager
#             if manager:
#                 create_notification(
#                     manager,
#                     f"{request.user.username} applied for leave."
#                 )

#             messages.success(request, "Leave applied successfully")
#             return redirect(request.path)

    
    
#     return render(request,'attendance/leave/apply_leave.html',{
#         "today":timezone.localdate()
#     })


# views.py માં apply_leave_view ફંક્શન શોધો અને આ કોડથી બદલો
from .middleware import ThreadLocal # Import check kari lejo

@login_required
def apply_leave_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Applying leave in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if request.method == "POST":
        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        reason = request.POST.get("reason")

        try:
            from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Invalid Date Format")

        today = timezone.localdate()

        # Validation
        if from_date < today or to_date < today:
            return HttpResponse("Past dates not allowed")
        
        # ✅ STEP-1: Save Leave (Using Company DB)
        Leave.objects.using(db_name).create(
            user=request.user,
            from_date=from_date,
            to_date=to_date,
            reason=reason,
            status="PENDING"
        )

        # --------------------------------------------------------
        # 🔥 STEP-2: NOTIFICATION LOGIC (Sacha DB ma save thase)
        # --------------------------------------------------------
        try:
            is_hr = request.user.groups.filter(name__icontains='hr').exists()
            receivers = []
            msg = ""

            if is_hr:
                # HR applied -> Admin ne find karo (Main DB mathi User malse)
                receivers = User.objects.filter(
                    profile__company=company,
                    groups__name__icontains="admin"
                )
                msg = f"HR {request.user.username} applied for leave."
            else:
                # Employee/Manager applied -> HR ne find karo
                receivers = User.objects.filter(
                    profile__company=company,
                    groups__name__icontains="hr"
                )
                msg = f"{request.user.username} applied for leave."

            # Create Notifications (Sacha Company DB ma fourcefully)
            for receiver in receivers:
                Notification.objects.using(db_name).create(
                    user=receiver, 
                    message=msg, 
                    is_read=False
                )

        except Exception as e:
            print(f"Notification Error: {e}")

        messages.success(request, "Leave applied successfully")
        return redirect('apply_leave')
    
    return render(request, 'attendance/leave/apply_leave.html', {"today": timezone.localdate()})



# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def manager_leave_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        # Pela profile fetch karo (Main DB mathi profile malse)
        manager_profile = request.user.profile
        company = manager_profile.company
        manager_branch = manager_profile.branch 
        
        if company:
            # Have have pachi ni badhi queries aa Company na DB ma jase
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
            print(f"DEBUG: Manager Leave Requests switching to {db_name}")
        else:
            ThreadLocal.DB_NAME = 'default'
            db_name = 'default'
            
    except Exception:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Error: Your Profile is missing. Contact Admin.")

    # 1. Permission Check (Have queries sacha DB ma jase)
    if not user_has_custom_permission(request.user, 'can_approve_leave'):
        return render(request, 'attendance/MISC/access_denied.html')

    # ---------------------------------------------------------
    # 2. Find Team Members (Sacha Company DB mathi)
    # ---------------------------------------------------------
    # Note: Badho data have g8_db (db_name) mathi avse
    team_profiles = Profile.objects.filter(
        manager=request.user,
        company=company,
        branch=manager_branch 
    ).select_related("user")
    
    # 3. Get IDs (Filtered team user ids)
    team_user_ids = team_profiles.values_list("user_id", flat=True)

    # 4. Fetch Requests (Filtered by team IDs and Company DB)
    requests = Leave.objects.filter(
        user_id__in=team_user_ids
    ).order_by("-applied_at").select_related('user')
    
    return render(request, "attendance/leave/manage_leave_request.html", {
        "requests": requests,
        "current_branch": manager_branch
    })

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def appove_leave(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_current = request.user.profile
        company = profile_current.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Approving leave in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. PERMISSION CHECK
    if not user_has_permission(request.user, "can_approve_attendance"):
        return HttpResponse("Access Denied !")
    
    # ૨. રજાની વિગત મેળવો (Using Company DB)
    leave = get_object_or_404(Leave.objects.using(db_name), id=id)

    # ૩. પ્રોફાઈલ અને રોલ ચેક (Sacha DB mathi)
    approver_profile = Profile.objects.using(db_name).get(user=request.user)
    leave_user_profile = Profile.objects.using(db_name).get(user=leave.user)

    approver_role = approver_profile.role.upper()
    leave_role = leave_user_profile.role.upper()

    # MANAGER → EMPLOYEE only
    if approver_role == "MANAGER" and leave_role != "EMPLOYEE":
        return HttpResponse("Not allowed")

    # ADMIN → MANAGER & HR only
    if approver_role == "ADMIN" and leave_role not in ["MANAGER", "HR"]:
        return HttpResponse("Not allowed")

    # ૪. અપ્રૂવ કરો
    leave.status = "APPROVED"
    leave.save(using=db_name) # ✅ SAVE IN COMPANY DB
    
    # ૫. નોટિફિકેશન (Sacha DB ma jase)
    create_notification(
        leave.user,
        "Your leave request has been approved."
    )

    messages.success(request, "Leave Approved successfully!")
    return redirect(request.META.get("HTTP_REFERER"))


@login_required
def reject_leave(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_current = request.user.profile
        company = profile_current.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_approve_attendance"):
        return HttpResponse("Access Denied !")

    # ૧. રજાની વિગત મેળવો (Using Company DB)
    leave = get_object_or_404(Leave.objects.using(db_name), id=id)

    # ૨. પ્રોફાઈલ લોજિક (Sacha DB mathi)
    approver = Profile.objects.using(db_name).get(user=request.user)
    leave_user = Profile.objects.using(db_name).get(user=leave.user)

    # Role Validation
    if approver.role.upper() == "MANAGER" and leave_user.role.upper() != "EMPLOYEE":
        return HttpResponse("Not allowed")

    if approver.role.upper() == "ADMIN" and leave_user.role.upper() not in ["MANAGER", "HR"]:
        return HttpResponse("Not allowed")

    # ૩. રિજેક્ટ કરો
    leave.status = "REJECTED"
    leave.save(using=db_name) # ✅ SAVE IN COMPANY DB

    create_notification(
        leave.user,
        "Your leave request has been rejected."
    )

    messages.warning(request, "Leave Rejected.")
    return redirect(request.META.get("HTTP_REFERER"))
from .middleware import ThreadLocal # Import check kari lejo

@login_required
def admin_pending_managers(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        # Pela Admin ni company mujab DB select karo
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Admin Pending switching to {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_manage_users"):
        return HttpResponse("Access Denied")

    # ✅ 1. FETCH PENDING MANAGERS (Sacha Company DB mathi)
    pending = Profile.objects.filter(
        is_verified=True,
        is_approved=False,
        role="MANAGER",
        company=company # Safety filter
    ).select_related('user')

    return render(request, "attendance/admin/admin_pending_managers.html", {
        "pending": pending
    })

@login_required
def admin_approve_user(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_manage_users"):
        return HttpResponse("Access Denied")

    # ✅ 1. FETCH PROFILE (Sacha Company DB mathi)
    profile = get_object_or_404(Profile.objects.using(db_name), id=id)

    # 🔥 2. GROUP LOGIC (Groups hamesha main DB ma hova joie login mate)
    if profile.role == "MANAGER" and company:
        group_name = f"{company.slug}_manager"
        
        # Default DB ma group create/get karo (Login system mate)
        mgr_group, _ = Group.objects.get_or_create(name=group_name)
        profile.user.groups.add(mgr_group)
        
        # Company DB ma pan group ensure karo (Separation mate)
        Group.objects.using(db_name).get_or_create(name=group_name)

    # ✅ 3. APPROVE & SAVE (In Company DB)
    profile.is_approved = True
    profile.save(using=db_name)

    messages.success(request, f"Manager {profile.user.username} approved successfully!")
    return redirect("/attendance/admin/pending-managers/")
from .middleware import ThreadLocal # Import hamesha check kari lejo

@login_required
def admin_reject_user(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_manage_users"):
        return HttpResponse("Access Denied")

    # ૧. પ્રોફાઇલ મેળવો અને ડિલીટ કરો (Sacha Company DB mathi)
    profile = get_object_or_404(Profile.objects.using(db_name), id=id)
    
    # Profile delete karva thi user delete nahi thay, jo user pan delete karvo hoy to:
    # user = profile.user
    profile.delete() # Company DB mathi profile jase
    # user.delete() # Jo user ne main DB mathi kadhvo hoy to aa line enable karo

    messages.warning(request, "Manager request rejected and profile removed.")
    return redirect("/attendance/admin/pending-managers/")


@login_required
def my_leave_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # ✅ ૧. FETCH LEAVES (Sacha Company DB mathi)
    leaves = Leave.objects.filter(user=request.user).order_by('-applied_at')

    return render(request, "attendance/leave/my_leave_requests.html", {
        "leaves": leaves
    })


@login_required
def admin_list_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # ✅ ૧. FETCH ADMINS (Sacha Company DB mathi)
    # Note: Filter 'admin' hamesha company wise thava joie
    users = Profile.objects.filter(
        company=company,
        user__groups__name__icontains="admin"
    ).select_related('user')

    return render(request, "attendance/admin/admin_list.html", {
        "users": users
    })

from .middleware import ThreadLocal # Import hamesha check kari lejo

@login_required
def hr_list_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        admin_branch = profile.branch
        
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: HR List switching to {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile Error")

    # ✅ 1. FETCH HRs (Sacha Company DB mathi)
    users = Profile.objects.filter(
        company=company,
        user__groups__name__icontains="hr"
    )
    
    if admin_branch:
        users = users.filter(branch=admin_branch)

    return render(request, "attendance/hr/hr_list.html", {"users": users})


@login_required
def manager_list_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile Error")

    # ✅ 1. FETCH MANAGERS (Company Filter sathe)
    # Role check karva sathe hamesha Company filter rakhvu
    users = Profile.objects.filter(
        company=company,
        role="MANAGER"
    ).select_related('user')

    return render(request, "attendance/MISC/manager_list.html", {"users": users})


@login_required
def employee_list_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile Error")

    # ✅ 1. FETCH EMPLOYEES (Company Filter sathe)
    users = Profile.objects.filter(
        company=company,
        role="EMPLOYEE"
    ).select_related('user')

    return render(request, "attendance/MISC/employee_list.html", {"users": users})
# @login_required
# def manager_team_attendance(request):

#     if not user_has_permission(request.user,"can_view_team"):
#         return HttpResponse("Access Denied")

#     team = Profile.objects.filter(manager=request.user)

#     team_users = [t.user for t in team]

#     records = Attendance.objects.filter(
#         user__in = team_users
#     ).order_by('-date')

#     return render(request,"attendance/manager_team_attendance.html",{
#         "records":records
#     })

from datetime import datetime, timedelta

# attendance/views.py માં admin_view_attendance શોધો અને આ કોડ મૂકો

from datetime import datetime

from .middleware import ThreadLocal # Import hamesha check kari lejo

@login_required
def admin_view_attendance(request, user_id):

    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        admin_profile = request.user.profile
        admin_company = admin_profile.company
        admin_branch = admin_profile.branch
        
        db_name = f"{admin_company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Viewing attendance in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. પરમિશન ચેક
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    if not (is_admin or user_has_permission(request.user, "can_view_reports")):
        return HttpResponse("Access Denied")

    # 2. ટાર્ગેટ એમ્પ્લોઈ શોધો (User main DB mathi, Profile company DB mathi)
    target_user = get_object_or_404(User, id=user_id)
    target_profile = get_object_or_404(Profile.objects.using(db_name), user=target_user)

    # Security: એમ્પ્લોઈ તે જ કંપનીનો હોવો જોઈએ
    if target_profile.company != admin_company:
        return HttpResponse("Not Allowed: Employee belongs to another company")

    # 🔥 BRANCH SECURITY CHECK
    if admin_branch and target_profile.branch != admin_branch:
        return HttpResponse("Access Denied: You cannot view attendance of another branch.")
    
    # 3. ડેટા લાવો (Sacha Company DB mathi)
    attendance_list = Attendance.objects.filter(user=target_user).order_by("-date")
    policy = AttendancePolicy.objects.filter(company=admin_company).first()
    shift = target_profile.shift

    # 4. લૂપ ફેરવીને સ્ટેટસ અને કલાકો ગણો (Original Logic maintained)
    for row in attendance_list:
        row.calculated_status = row.status 
        row.duration_str = "-"

        if row.check_in and row.check_out:
            in_dt = datetime.combine(row.date, row.check_in)
            out_dt = datetime.combine(row.date, row.check_out)
            
            hours = (out_dt - in_dt).total_seconds() / 3600
            row.duration_str = f"{int(hours)}h {int((hours*60)%60)}m"

            if policy:
                late_mins = 0
                if shift:
                    shift_start = datetime.combine(row.date, shift.start_time)
                    if in_dt > shift_start:
                        late_mins = (in_dt - shift_start).total_seconds() / 60
                
                if hours < (policy.work_hours_required / 2):
                    row.calculated_status = "Half Day"
                elif late_mins > policy.late_after_minutes:
                    row.calculated_status = "Late"
                elif hours >= policy.work_hours_required:
                    row.calculated_status = "Present"
                else:
                    row.calculated_status = "Partial"
        
        elif row.check_in:
            row.calculated_status = "Working"
        else:
            row.calculated_status = "Absent"

    return render(request, "attendance/attendance/admin_view_attendance.html", {
        "employee": target_profile,
        "attendance": attendance_list,
        "policy": policy
    })

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def admin_check_in(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_view_reports"):
        return HttpResponse("Access Denied")

    today = timezone.localdate()

    # ✅ 1. GET OR CREATE IN COMPANY DB
    attendance, created = Attendance.objects.using(db_name).get_or_create(
        user=request.user,
        date=today
    )

    if attendance.check_in is None:
        attendance.check_in = timezone.localtime().time()
        attendance.status = "Present"
        # ✅ 2. SAVE IN COMPANY DB
        attendance.save(using=db_name)

    return redirect('/attendance/dashboard/admin/')


@login_required
def admin_check_out(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_view_reports"):
        return HttpResponse("Access Denied")

    today = timezone.localdate()

    try:
        # ✅ 1. GET FROM COMPANY DB
        attendance = Attendance.objects.using(db_name).get(user=request.user, date=today)

        if attendance.check_out is None:
            attendance.check_out = timezone.localtime().time()
            # ✅ 2. SAVE IN COMPANY DB
            attendance.save(using=db_name)
            
    except Attendance.DoesNotExist:
        pass

    return redirect('/attendance/dashboard/admin/')
@transaction.non_atomic_requests
@login_required
def superadmin_dashboard_view(request):
    from accounts.models import is_superadmin
    if not request.user.is_authenticated:
        # જો લોગિન ના હોય, તો સ્લગ વગરના લોગિન પેજ પર મોકલો
        return redirect('login_no_slug')
    if not request.user.is_superuser:
        return HttpResponse("Access Denied !")
    
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        if form_type == "company":
            name = request.POST.get("company_name")
            if name:
                company_slug = slugify(name)
        
        # 1. Use 'get_or_create' so it doesn't crash if name already exists ✅
                new_company, created = Company.objects.get_or_create(
            slug=company_slug, 
            defaults={'name': name}
        )
        
        # 2. Always run setup (it will create the file if it's missing)
                try:
                    setup_company_db(new_company.slug)
            
                    if created:
                        messages.success(request, f"New company {name} created!")
                    else:
                        messages.info(request, f"Database restored/updated for existing company {name}.")
                
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
            
                return redirect('/attendance/dashboard/superadmin/')

            return redirect('/attendance/dashboard/superadmin/')
        
        if form_type == "policy":
            company_id = request.POST.get("company_id")
            work_hours = request.POST.get("work_hours")
            late_after = request.POST.get("late_after")
            grace_time = request.POST.get("grace_time")

            company = Company.objects.get(id=company_id)

            AttendancePolicy.objects.create(
                company = company,
                work_hours_required = work_hours,
                late_after_minutes = late_after,
                grace_time_minutes = grace_time
            )

        return redirect('/attendance/dashboard/superadmin/')

       
    companies = Company.objects.all()
    policies = AttendancePolicy.objects.all()
    total_companies = Company.objects.count()

    # attendance/views.py માં
    

    total_admins = User.objects.filter(groups__name__icontains="admin").count()
    total_hr = User.objects.filter(groups__name__icontains="hr").count()
    total_managers = User.objects.filter(groups__name__icontains="manager").count()
    total_employees = User.objects.filter(groups__name__icontains="employee").count()
    
    return render(request,"attendance/dashboard/superadmin_dashboard.html",{
        "companies":companies,
        "policies":policies,
        "superadmin_name": request.user.username,
        "total_companies": total_companies,
        "total_admins": total_admins,
        "total_hr": total_hr,
        "total_managers": total_managers,
        "total_employees": total_employees,
        

    })

# attendance/views.py માં import ના હોય તો ઉપર ઉમેરજો:
from django.db import transaction

# attendance/views.py

from django.db import transaction  # આ Import સૌથી ઉપર હોવું જોઈએ

from django.db import transaction
from django.contrib.auth.models import User, Group

from django.db import transaction
from .middleware import ThreadLocal # Import check kari lejo

@login_required
def delete_company(request, id):
    # 1. Security Check (Only Superuser)
    if not request.user.is_superuser:
        return HttpResponse("Access Denied!")

    company = get_object_or_404(Company, id=id)
    company_name = company.name
    company_slug = company.slug
    db_name = f"{company_slug}_db" # Company nu database name

    try:
        with transaction.atomic():
            # =========================================================
            # STEP 0: Pehla Company na DB mathi badho data delete karo
            # =========================================================
            # Aa zaruri che kem ke Attendance/Leaves sacha DB ma che
            ThreadLocal.DB_NAME = db_name
            
            # Note: Cascade delete Company DB na badha data clean kari dese
            # Pan Profile/User main DB ma hashe, etle tya switch karvu padse
            
            # =========================================================
            # STEP 1: Main DB mathi Users delete karo
            # =========================================================
            ThreadLocal.DB_NAME = 'default' # Main DB ma switch karo
            company_users_ids = Profile.objects.filter(company=company).values_list('user_id', flat=True)
            
            if company_users_ids:
                User.objects.filter(id__in=company_users_ids).exclude(is_superuser=True).delete()

            # =========================================================
            # STEP 2: Main DB mathi Groups (Roles) delete karo
            # =========================================================
            Group.objects.filter(name__startswith=f"{company_slug}_").delete()

            # =========================================================
            # STEP 3: Company delete karo
            # =========================================================
            company.delete()
            
            messages.success(request, f"Company '{company_name}' and ALL data deleted successfully.")

    except Exception as e:
        ThreadLocal.DB_NAME = 'default'
        messages.error(request, f"Error deleting company: {str(e)}")
        return redirect('superadmin_company_details', company_id=id)

    ThreadLocal.DB_NAME = 'default'
    return redirect('superadmin_dashboard')


@login_required
def edit_company(request, id):
    from accounts.models import is_superadmin

    # Security check using your custom function
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied !")
    
    # Company hamesha 'default' DB ma hoy che (Master Table)
    company = get_object_or_404(Company, id=id)

    if request.method == "POST":
        name = request.POST.get("company_name")
        if name:
            company.name = name
            company.save() # default DB ma save thase
            messages.success(request, "Company name updated successfully.")
            return redirect("/attendance/dashboard/superadmin/")
            
    return render(request, "attendance/superadmin/edit_company.html", {
        "company": company
    })    


from .middleware import ThreadLocal # Import check kari lejo

from django.http import HttpResponseForbidden

@login_required
def company_add_policy(request, company_slug):
    # 1. મેઈન ડેટાબેઝમાંથી કંપની શોધો
    company = get_object_or_404(Company.objects.using('default'), slug=company_slug)
    db_name = f"{company.slug}_db"
    
    # 🛡️ 2. SECURITY CHECK: Superadmin OR Company Owner only
    is_superuser = request.user.is_superuser
    owner_group_name = f"{company.slug}_owner"
    is_owner = request.user.groups.filter(name__iexact=owner_group_name).exists()
    
    if not (is_superuser or is_owner):
        return HttpResponseForbidden("Access Denied: Only Superadmin or Company Owner can add policies.")

    # 🛡️ 3. સુનિશ્ચિત કરો કે કનેક્શન જીવંત છે
    from attendance.utils import ensure_db_connection
    ensure_db_connection(db_name)

    if request.method == "POST":
        # 🛡️ 4. CRITICAL STEP: ફોરેન કી એરર બચાવવા માટે કંપનીને ટેનન્ટ DB માં સિંક કરો
        if not Company.objects.using(db_name).filter(id=company.id).exists():
            company.save(using=db_name)
            print(f"✅ Company {company.slug} synced to {db_name} before adding policy.")

        # 5. ડેટા મેળવો
        work_hours = request.POST.get("work_hours")
        late_after = request.POST.get("late_after")
        grace_time = request.POST.get("grace_time")

        try:
            # 6. હવે પોલિસી સેવ કરો
            AttendancePolicy.objects.using(db_name).create(
                company=company,
                work_hours_required=work_hours,
                late_after_minutes=late_after,
                grace_time_minutes=grace_time
            )
            
            # ThreadLocal સેટ કરો જેથી ડેશબોર્ડ પર રીડાયરેક્ટ થાય ત્યારે ડેટા દેખાય
            from attendance.middleware import ThreadLocal
            ThreadLocal.DB_NAME = 'default' 
            
            messages.success(request, "Policy Set Successfully")
            
            # 🔄 7. DYNAMIC REDIRECT (Superadmin vs Owner)
            if request.user.is_superuser:
                return redirect('superadmin_company_dashboard', company_id=company.id)
            else:
                return redirect('company_owner_dashboard', company_slug=company.slug)
            
        except Exception as e:
            messages.error(request, f"Error saving policy: {str(e)}")
            return redirect(request.path)
    
    return render(request, "attendance/MISC/add_policy.html", {"companies": [company]})
@login_required
def hr_policy_list(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except Profile.DoesNotExist:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile not found")

    # 1. Security Check
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    is_admin = request.user.groups.filter(name__icontains='admin').exists()

    if not (request.user.is_superuser or is_hr or is_admin or user_has_permission(request.user, "can_manage_users")):
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 2. Fetch Policies (Sacha Company DB mathi)
    # ---------------------------------------------------------
    # Note: Badho data have g8_db (db_name) mathi avse
    policies = AttendancePolicy.objects.select_related("company").filter(company=company)

    return render(request, "attendance/hr/hr_policy_list.html", {
        "policies": policies,
        "company": company
    })

    
# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def hr_today_attendance(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        
        if company:
            # Have have pachi ni badhi queries aa Company na DB ma jase
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
            print(f"DEBUG: HR Today Attendance switching to {db_name}")
        else:
            ThreadLocal.DB_NAME = 'default'
            
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile not found")

    # 1. Permission Check (Sacha DB mathi fetch thase)
    if not user_has_custom_permission(request.user, "can_view_reports"):
         return HttpResponse("Access Denied")

    today = timezone.now().date()

    # ---------------------------------------------------------
    # 2. Fetch Employees (Sacha Company DB mathi)
    # ---------------------------------------------------------
    employees = Profile.objects.filter(
        company=company,
        user__groups__name__icontains="employee"
    ).select_related('user')

    data = []

    for emp in employees:
        # ✅ Fetch Attendance from Company DB
        att = Attendance.objects.filter(user=emp.user, date=today).first()
        
        status = "Absent"
        check_in = "-"
        check_out = "-"

        if att:
            if att.status:
                status = att.status
            elif att.check_in:
                status = "Present"
            
            # Format Times
            if att.check_in:
                check_in = att.check_in.strftime("%I:%M %p") if hasattr(att.check_in, 'strftime') else str(att.check_in)
            if att.check_out:
                check_out = att.check_out.strftime("%I:%M %p") if hasattr(att.check_out, 'strftime') else str(att.check_out)

        data.append({
            "employee": emp.user.username,
            "email": emp.user.email,
            "check_in": check_in,
            "check_out": check_out,
            "status": status
        })

    return render(request, "attendance/attendance/hr_today_attendance.html", {
        "data": data,
        "today": today
    })

from .middleware import ThreadLocal # Import check kari lejo
from datetime import datetime

@login_required
def request_correction(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Requesting correction in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # ----------------------------------------------------
    # 1. POST REQUEST LOGIC
    # ----------------------------------------------------
    if request.method == "POST":
        selected_date = request.POST.get("selected_date")
        new_check_in = request.POST.get("new_check_in")
        new_check_out = request.POST.get("new_check_out")
        reason = request.POST.get("reason")

        # Future Date Validation
        try:
            request_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            today_date = timezone.localdate()

            if request_date > today_date:
                messages.error(request, "You cannot request correction for a future date.")
                return redirect('request_correction')
        except ValueError:
            messages.error(request, "Invalid Date Format")
            return redirect('request_correction')

        # Time Formatting
        formatted_in = None
        formatted_out = None
        try:
            if new_check_in:
                formatted_in = datetime.strptime(new_check_in, "%H:%M").time()
            if new_check_out:
                formatted_out = datetime.strptime(new_check_out, "%H:%M").time()
        except ValueError:
             messages.error(request, "Invalid time format")
             return redirect('request_correction')

        # ✅ STEP-1: Save Correction Request (Using Company DB)
        AttendanceCorrection.objects.using(db_name).create(
            user=request.user,
            date=selected_date,
            new_check_in=formatted_in,
            new_check_out=formatted_out,
            reason=reason,
            status="PENDING"
        )
        
        # --------------------------------------------------------
        # 🔥 STEP-2: NOTIFICATION LOGIC (Sacha DB ma save thase)
        # --------------------------------------------------------
        try:
            is_hr = request.user.groups.filter(name__icontains='hr').exists()
            receivers = []
            msg = ""

            if is_hr:
                receivers = User.objects.filter(profile__company=company, groups__name__icontains="admin")
                msg = f"HR {request.user.username} requested attendance correction."
            else:
                receivers = User.objects.filter(profile__company=company, groups__name__icontains="hr")
                msg = f"{request.user.username} requested attendance correction."

            for receiver in receivers:
                Notification.objects.using(db_name).create(
                    user=receiver, 
                    message=msg, 
                    is_read=False
                )
        except Exception as e:
            print(f"Notification Error: {e}")

        messages.success(request, "Correction Request Sent Successfully!")
        return redirect('dashboard')

    # ----------------------------------------------------
    # 2. GET REQUEST LOGIC
    # ----------------------------------------------------
    today = timezone.localdate()
    
    # Sacha DB mathi record shodho
    attendance_record = Attendance.objects.using(db_name).filter(user=request.user, date=today).first()

    return render(request, "attendance/correction/request_correction.html", {
        "today": today,
        "attendance": attendance_record
    })
from .middleware import ThreadLocal # Import check kari lejo

@login_required
def hr_approve_correction(request, pk):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Approving correction in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. SECURITY CHECK (Have queries sacha DB ma jase)
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    
    if not (request.user.is_superuser or is_hr or is_admin or user_has_permission(request.user, "can_approve_attendance")):
        return HttpResponse("Access Denied")

    # ૨. CORRECTION REQUEST મેળવો (Using Company DB)
    correction = get_object_or_404(AttendanceCorrection.objects.using(db_name), id=pk, user__profile__company=company)

    if correction.status != "PENDING":
        return redirect("/attendance/hr/correction-requests/")

    # ---------------------------------------------------
    # 3. FETCH/CREATE ATTENDANCE (Sacha Company DB ma)
    # ---------------------------------------------------
    attendance, created = Attendance.objects.using(db_name).get_or_create(
        user = correction.user,
        date = correction.date
    )

    # APPLY UPDATED TIMES
    if correction.new_check_in:
        attendance.check_in = correction.new_check_in
    if correction.new_check_out:
        attendance.check_out = correction.new_check_out
    
    # ✅ SAVE IN COMPANY DB
    attendance.save(using=db_name)

    # ---------------------------------------------------
    # 4. UPDATE REQUEST STATUS & NOTIFICATION
    # ---------------------------------------------------
    correction.status = "APPROVED"
    correction.save(using=db_name)

    # NOTIFICATION (Automatic switch thase ThreadLocal thi)
    create_notification(
        correction.user,
        "Your attendance correction request has been approved."
    )

    messages.success(request, f"Correction approved for {correction.user.username}")
    return redirect("/attendance/hr/correction-requests/")

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def hr_revoke_member(request, member_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        hr_profile = request.user.profile
        company = hr_profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: HR Revoke Member switching to {db_name}")
    except Exception:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Error: Your Profile is missing.")

    # 1. પરમિશન ચેક (ફક્ત HR અથવા એડમિન)
    if not (request.user.groups.filter(name__icontains='hr').exists() or request.user.is_superuser):
        messages.error(request, "Access Denied.")
        return redirect('hr_tl_list')

    try:
        # ૨. મેમ્બર શોધો (Sacha Company DB mathi)
        member_profile = Profile.objects.using(db_name).get(user__id=member_id)
        
        # જૂના TL નું નામ મેસેજ માટે સાચવી લો
        old_tl_name = member_profile.team_leader.username if member_profile.team_leader else "Team Leader"
        
        # ૩. રિવોક કરો (NULL સેટ કરો અને Sacha DB ma save karo)
        member_profile.team_leader = None
        member_profile.save(using=db_name)
        
        messages.warning(request, f"Member removed from {old_tl_name}'s squad.")
        
    except Profile.DoesNotExist:
        messages.error(request, "Member profile not found.")
    
    # પાછા TL લિસ્ટ પર જાઓ
    return redirect('hr_tl_list')

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def revoke_team_leader(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        requester_profile = request.user.profile
        company = requester_profile.company
        db_name = f"{company.slug}_db"
        from .middleware import ThreadLocal
        ThreadLocal.DB_NAME = db_name
    except Exception:
        return HttpResponse("Error: Profile missing.")

    # 1. SECURITY CHECK (Admin, HR, or Superuser only)
    is_management = request.user.groups.filter(name__iregex=fr'({company.slug}_admin|{company.slug}_hr)').exists()
    if not (is_management or request.user.is_superuser):
        messages.error(request, "Access Denied.")
        return redirect('dashboard')

    # Target TL User મેળવો
    target_tl = get_object_or_404(User.objects.using('default'), id=user_id)

    try:
        with transaction.atomic(using='default'):
            with transaction.atomic(using=db_name):
                
                # ---------------------------------------------------------
                # 🔥 STEP 1: DELETE ALL TASKS ASSIGNED BY THIS TL
                # ---------------------------------------------------------
                # આ TL એ જેટલા પણ ટાસ્ક આપ્યા હતા તે કંપની ડેટાબેઝમાંથી કાઢી નાખો
                deleted_tasks_count = Task.objects.using(db_name).filter(assigned_by_id=user_id).delete()
                print(f"DEBUG: Deleted {deleted_tasks_count[0]} tasks assigned by TL {target_tl.username}")

                # ---------------------------------------------------------
                # 🔥 STEP 2: RESET SQUAD MEMBERS (Team Leader = None)
                # ---------------------------------------------------------
                # સ્ક્વોડ મેમ્બર્સની બ્રાન્ચ સાચવીને માત્ર ટીએલ લિંક તોડો
                # આપણે .update() વાપરીશું જેથી બીજી કોઈ વિગત (Branch/Manager) ના ભૂંસાય
                Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader_id=None)
                Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader_id=None)

                # ---------------------------------------------------------
                # 🔥 STEP 3: REVOKE THE ROLE (Remove Group)
                # ---------------------------------------------------------
                tl_group_name = f"{company.slug.lower()}_team leader"
                tl_group = Group.objects.using('default').filter(name__iexact=tl_group_name).first()
                if tl_group:
                    target_tl.groups.remove(tl_group)
                
                # Profile status update (Optional: TL પોતે હવે એમ્પ્લોઈ ગણાય)
                Profile.objects.using('default').filter(user_id=user_id).update(role='employee')
                Profile.objects.using(db_name).filter(user_id=user_id).update(role='employee')

        messages.success(request, f"Success: {target_tl.username} is revoked. Squad cleared and all related tasks deleted.")
        
    except Exception as e:
        messages.error(request, f"Termination Error: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


from .middleware import ThreadLocal # Import check kari lejo

@login_required
def hr_reject_correction(request, pk):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Rejecting correction in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. SECURITY CHECK (Have queries sacha DB ma jase)
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    
    if not (request.user.is_superuser or is_hr or is_admin or user_has_permission(request.user, "can_approve_attendance")):
        return HttpResponse("Access Denied")

    # ૨. CORRECTION REQUEST મેળવો (Sacha Company DB mathi)
    correction = get_object_or_404(AttendanceCorrection.objects.using(db_name), id=pk, user__profile__company=company)

    # ૩. REJECT ACTION
    if correction.status == "PENDING":
        correction.status = "REJECTED"
        # ✅ SAVE IN COMPANY DB
        correction.save(using=db_name)
        
        # ૪. NOTIFICATION (Automatic switch thase ThreadLocal thi)
        create_notification(
            correction.user,
            "Your attendance correction request has been rejected."
        )
        
        messages.warning(request, f"Correction request for {correction.user.username} has been rejected.")

    return redirect("/attendance/hr/correction-requests/")


# @login_required
# def hr_correction_requests(request):
#     if not user_has_custom_permission(request.user, 'can_approve_attendance'):
#         raise PermissionDenied("You do not have permission to manage attendance corrections.")

#     profile = Profile.objects.get(user=request.user)

#     # ---------- HR ----------
#     if profile.role == "HR":
#         requests = AttendanceCorrection.objects.exclude(
#             user__profile__role = "HR"
#         ).order_by('-created_at')

#     # ---------- ADMIN ----------
#     elif profile.role == "ADMIN":
#         requests = AttendanceCorrection.objects.filter(
#             user__profile__role = "HR"
#         ).order_by('-created_at')

#     else:
#         return HttpResponse("Access Denied")

#     return render(request,"attendance/correction/hr_correction_requests.html",{
#         "requests":requests,
#         "role": profile.role
#     })

from datetime import time, timedelta, datetime
# attendance/views.py

from django.core.exceptions import PermissionDenied
from .middleware import ThreadLocal # Import check kari lejo
from datetime import datetime, time, timedelta

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from datetime import time, timedelta, datetime

from .models import Attendance, AttendancePolicy
from accounts.models import Profile

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives # 🔥 ઈમેલમાં ટેબલ મોકલવા માટે આ ફરજિયાત છે
from django.conf import settings
from datetime import time, timedelta, datetime

from .models import Attendance, AttendancePolicy
from accounts.models import Profile

@login_required
def hr_daily_attendance(request):
    # ---------------------------------------------------------
    # 0. DATABASE SETUP
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        hr_branch = profile.branch # આ None પણ હોઈ શકે અને બ્રાન્ચ પણ હોઈ શકે
        
        if company:
            db_name = f"{company.slug}_db"
        else:
            return HttpResponse("Company not found.")
    except Exception:
        return HttpResponse("Profile not found")

    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    if not (is_hr or request.user.is_superuser):
        raise PermissionDenied("Access Denied.")

    today = timezone.localdate()

    # ---------------------------------------------------------
    # 📧 1. RECIPIENTS FINDER (For Email Modal)
    # ---------------------------------------------------------
    owner_qs = User.objects.using('default').filter(profile__company=company, groups__name__icontains='owner')
    
    # 🔥 CONDITION: એડમિન પણ એ જ બ્રાન્ચનો હોવો જોઈએ જે બ્રાન્ચનો HR છે (અથવા બ્રાન્ચ વગરનો)
    admin_qs = User.objects.using('default').filter(
        profile__company=company, 
        profile__branch=hr_branch, 
        groups__name__icontains='admin'
    )
    
    report_recipients = (owner_qs | admin_qs).distinct().exclude(email="")

    # ---------------------------------------------------------
    # 👥 2. FETCH ALL USERS & OVERALL COUNTS
    # ---------------------------------------------------------
    # 🔥 CONDITION: એમ્પ્લોઈ પણ એ જ બ્રાન્ચના હોવા જોઈએ જે બ્રાન્ચનો HR છે (અથવા બ્રાન્ચ વગરના)
    base_users = User.objects.using('default').filter(
        profile__company=company,
        profile__branch=hr_branch # 👈 THE MAGIC FILTER
    ).exclude(groups__name__iregex=r'(owner|admin)')
    
    all_users = base_users.distinct().order_by('first_name', 'username')
    user_ids = list(all_users.values_list('id', flat=True))
    total = len(user_ids)

    # ---------------------------------------------------------
    # 📅 3. FETCH TODAY'S ATTENDANCE
    # ---------------------------------------------------------
    today_attendance = list(Attendance.objects.using(db_name).filter(
        date=today, user_id__in=user_ids
    ))

    present = sum(1 for att in today_attendance if att.status and att.status.lower() in ["present", "working", "late", "half day", "partial"])
    absent = total - present

    # ---------------------------------------------------------
    # 📋 4. CALCULATE TABLE DATA FIRST
    # ---------------------------------------------------------
    att_map = {att.user_id: att for att in today_attendance}
    data = []

    for user in all_users:
        record = att_map.get(user.id) 
        status = "Absent"
        check_in = None
        check_out = None

        if record:
            check_in = record.check_in
            check_out = record.check_out
            
            # 🔥 WORKING STATUS FIX
            if check_in and not check_out:
                status = "Working"
            elif check_in and check_out:
                if record.status and record.status.strip().lower() not in ["", "absent", "none"]:
                    status = record.status.title()
                else:
                    status = "Present"
            elif record.status and record.status.strip().lower() not in ["", "absent", "none"]:
                status = record.status.title()

        data.append({
            "user": user,
            "check_in": check_in,
            "check_out": check_out,
            "status": status
        })

    # ---------------------------------------------------------
    # 🚀 5. HANDLE DAILY EMAIL SEND (BULLETPROOF TABLE FIX)
    # ---------------------------------------------------------
    if request.method == "POST" and "send_daily_report" in request.POST:
        selected_emails = request.POST.getlist('target_emails')
        custom_note = request.POST.get('custom_note', '').strip()
        
        if selected_emails:
            subject = f"📊 Daily Attendance Report - {today.strftime('%d %b, %Y')} ({company.name})"
            branch_name = hr_branch.name if hr_branch else "Head Office (No Branch)"
            
            html_rows = ""
            plain_rows = "Name\t\t| In Time\t| Out Time\t| Status\n"
            plain_rows += "-" * 60 + "\n"

            # લૂપ કરીને બંને (Plain Text અને HTML) માટે ટેબલ રો બનાવો
            for row in data:
                emp_name = f"{row['user'].first_name} {row['user'].last_name}".strip() or row['user'].username.title()
                in_time = row['check_in'].strftime("%I:%M %p") if row['check_in'] else "—"
                out_time = row['check_out'].strftime("%I:%M %p") if row['check_out'] else "—"
                status = row['status']

                plain_rows += f"{emp_name[:15].ljust(15)} | {in_time.ljust(10)} | {out_time.ljust(10)} | {status}\n"

                color = "red" if status == "Absent" else "green" if status == "Present" else "blue" if status == "Working" else "orange"
                html_rows += f"""
                <tr>
                    <td style="border: 1px solid #cccccc; padding: 10px; color: #333333;"><b>{emp_name}</b></td>
                    <td style="border: 1px solid #cccccc; padding: 10px; text-align: center; color: #555555;">{in_time}</td>
                    <td style="border: 1px solid #cccccc; padding: 10px; text-align: center; color: #555555;">{out_time}</td>
                    <td style="border: 1px solid #cccccc; padding: 10px; text-align: center;"><b style="color: {color};">{status}</b></td>
                </tr>
                """

            # --- 📝 PLAIN TEXT MESSAGE (Fallback) ---
            plain_message = f"Daily Attendance Report for {branch_name} ({today.strftime('%d %b, %Y')})\n\n"
            plain_message += f"Total: {total} | Present: {present} | Absent: {absent}\n\n"
            if custom_note: plain_message += f"HR Remarks: {custom_note}\n\n"
            plain_message += plain_rows

            # --- 🎨 BEAUTIFUL HTML EMAIL MESSAGE ---
            html_message = f"""
            <html>
            <body style="font-family: Arial, sans-serif; color: #333333;">
                <h2 style="color: #540863; border-bottom: 2px solid #92487A; padding-bottom: 10px;">Daily Attendance Report</h2>
                <p><b>Branch:</b> {branch_name} | <b>Date:</b> {today.strftime('%d %b, %Y')}</p>
                
                <div style="background-color: #f8f9fa; padding: 15px; border: 1px solid #dddddd; margin-bottom: 20px;">
                    <strong>👥 Total Staff:</strong> {total} &nbsp;|&nbsp; 
                    <strong style="color: #008080;">✅ Present:</strong> {present} &nbsp;|&nbsp; 
                    <strong style="color: #e11d48;">❌ Absent:</strong> {absent}
                </div>
            """

            if custom_note:
                html_message += f"""
                <div style="background-color: #fffbeb; padding: 15px; border-left: 4px solid #f59e0b; margin-bottom: 20px;">
                    <strong>📝 HR Remarks:</strong><br>{custom_note}
                </div>
                """

            html_message += f"""
                <table width="100%" border="1" cellpadding="10" cellspacing="0" style="border-collapse: collapse; border-color: #cccccc;">
                    <thead>
                        <tr style="background-color: #f1f5f9; color: #333333;">
                            <th style="text-align: left;">Employee Name</th>
                            <th style="text-align: center;">In Time</th>
                            <th style="text-align: center;">Out Time</th>
                            <th style="text-align: center;">Status</th>
                        </tr>
                    </thead>
                    <tbody>
                        {html_rows}
                    </tbody>
                </table>
                <p style="font-size: 12px; color: #999999; margin-top: 20px;">Generated via AttendancePro System.</p>
            </body>
            </html>
            """
            
            try:
                # 🔥 ઈમેલમાં ટેબલ નહિ દેખાવાનો પ્રોબ્લેમ આનાથી સોલ્વ થશે 
                sender_email = getattr(settings, 'EMAIL_HOST_USER', 'webmaster@localhost')
                msg = EmailMultiAlternatives(subject, plain_message, sender_email, selected_emails)
                msg.attach_alternative(html_message, "text/html")
                msg.send(fail_silently=False)
                
                messages.success(request, "✅ Detailed daily report emailed successfully with Table!")
            except Exception as e:
                messages.error(request, f"⚠️ Failed to send email. Please check SMTP settings. Error: {str(e)}")
        else:
            messages.warning(request, "No recipients selected for the report.")
            
        return redirect(request.path)

    # ---------------------------------------------------------
    # 🏁 6. RENDER TEMPLATE
    # ---------------------------------------------------------
    return render(request, "attendance/attendance/hr_daily_attendance.html", {
        "today": today,
        "total": total, # જો HTML માં total વાપરતા હોવ 
        "total_staff": total, 
        "present": present, # જો HTML માં present વાપરતા હોવ
        "present_today": present, 
        "absent": absent, # જો HTML માં absent વાપરતા હોવ
        "absent_today": absent, 
        "employees": data, 
        "data": data, # જો HTML માં data વાપરતા હોવ
        "report_recipients": report_recipients, 
        "current_branch": hr_branch
    })
    
# @login_required
# def monthly_attendance(request):

#     from calendar import monthrange
#     from datetime import date, datetime, timedelta
#     from accounts.models import Profile
#     from .models import Attendance, AttendancePolicy, Holiday

#     today = date.today()

#     month = int(request.GET.get("month", today.month))
#     year = int(request.GET.get("year", today.year))

#     profile = Profile.objects.get(user=request.user)

#     # ---- Policy & Shift ----
#     policy = AttendancePolicy.objects.filter(company=profile.company).first()
#     shift = profile.shift

#     # ---- Get no. of days in month ----
#     days_in_month = monthrange(year, month)[1]

#     # ---- Attendance Records ----
#     attendance_map = {
#         a.date: a
#         for a in Attendance.objects.filter(
#             user=request.user,
#             date__year=year,
#             date__month=month
#         )
#     }

#     # ---- Holidays ----
#     holidays = {h.date: h.name for h in Holiday.objects.all()}

#     days = []

#     for i in range(1, days_in_month + 1):

#         d = date(year, month, i)

#         att = attendance_map.get(d)

#         day_info = {
#             "date": d,
#             "check_in": att.check_in if att else None,
#             "check_out": att.check_out if att else None,
#             "status": None,
#             "is_holiday": False,
#             "holiday_name": None,
#             "is_sunday": d.weekday() == 6,
#             "is_saturday": d.weekday() == 5,
#         }

#         # ----------------------------
#         # RULE-1 → Holiday
#         # ----------------------------
#         if d in holidays:
#             day_info["is_holiday"] = True
#             day_info["holiday_name"] = holidays[d]
#             day_info["status"] = holidays[d]

#         # ----------------------------
#         # RULE-2 → Attendance exists
#         # ----------------------------
#         elif att:

#             # only check-in done → WORKING
#             if att.check_in and not att.check_out:
#                 day_info["status"] = "Working"

#             # check-in & check-out → calculate
#             elif att.check_in and att.check_out:

#                 check_in_dt = datetime.combine(d, att.check_in)
#                 check_out_dt = datetime.combine(d, att.check_out)

#                 work_hours = (check_out_dt - check_in_dt).total_seconds() / 3600

#                 # Late calc
#                 late_minutes = 0
#                 if shift:
#                     shift_in = datetime.combine(d, shift.start_time)
#                     late_minutes = (check_in_dt - shift_in).total_seconds() / 60

#                 if policy:

#                     if work_hours < (policy.work_hours_required / 2):
#                         day_info["status"] = "Half Day"

#                     elif late_minutes > policy.late_after_minutes:
#                         day_info["status"] = "Late"

#                     elif work_hours >= policy.work_hours_required:
#                         day_info["status"] = "Present"

#                     else:
#                         day_info["status"] = "Partial"

#                 else:
#                     day_info["status"] = "Present"

#             else:
#                 day_info["status"] = "-"

#         # ----------------------------
#         # RULE-3 → Past date — no attendance → ABSENT
#         # ----------------------------
#         elif d < today:
#             day_info["status"] = "Absent"

#         # ----------------------------
#         # RULE-4 → Future date — blank
#         # ----------------------------
#         else:
#             day_info["status"] = "-"

#         days.append(day_info)


#     return render(request,"attendance/monthly.html",{
#         "days":days,
#         "month":month,
#         "year":year,
#     })

# @login_required
# def monthly_attendance(request):

#     from calendar import monthrange
    

#     def to_local(dt):
#         if not dt:
#             return None
#         if timezone.is_naive(dt):
#             return dt
#         return timezone.localtime(dt)

#     today = date.today()

#     # ---------- Logged-in user ----------
#     login_profile = Profile.objects.get(user=request.user)
#     # if not user_has_permission(request.user , "can_view_reports"):
#     #     return HttpResponse("Access Denied")
#     # ---------- Target user ----------
#     user_id = request.GET.get("user")

#     if user_id:
#         target_user = User.objects.get(id=user_id)
#     else:
#         target_user = request.user

#     target_profile = Profile.objects.get(user=target_user)

#     # ---------- SECURITY ----------

#     # EMPLOYEE → only self
#     if login_profile.role == "EMPLOYEE" and target_user != request.user:
#         return HttpResponse("Access Denied")

#     # HR → company restricted
#     if login_profile.role == "HR":
#         if target_profile.company != login_profile.company:
#             return HttpResponse("Access Denied")


#     # ---------- Month / Year ----------
#     month = int(request.GET.get("month", today.month))
#     year  = int(request.GET.get("year", today.year))


#     # ---------- Policy & Shift (EMPLOYEE BASED) ----------
#     policy = AttendancePolicy.objects.filter(
#         company = target_profile.company
#     ).first()

#     shift = target_profile.shift


#     # ---------- Attendance ----------
#     days_in_month = monthrange(year, month)[1]

#     records = Attendance.objects.filter(
#         user = target_user,
#         date__year = year,
#         date__month = month
#     )
#     # ---------- APPROVED LEAVES (NEW) ----------


#     approved_leaves = Leave.objects.filter(
#         user = target_user,
#         status = "APPROVED",
#         from_date__lte = date(year, month, days_in_month),
#         to_date__gte = date(year, month, 1)
#     )

#     leave_dates = set()

#     for leave in approved_leaves:
#         current = leave.from_date
#         while current <= leave.to_date:
#             leave_dates.add(current)
#             current += timedelta(days=1)


#     attendance_map = {a.date: a for a in records}

#     holidays = {h.date: h.name for h in Holiday.objects.all()}


#     # ---------- Counters ----------
#     days = []
#     working_days = present_count = absent_count = 0


#     for i in range(1, days_in_month + 1):

#         d = date(year, month, i)
#         att = attendance_map.get(d)

#         info = {
#             "date": d,
#             "check_in": to_local(att.check_in)if att else None,
#             "check_out": to_local(att.check_out)if att else None,

#             "status": None,
#             "is_holiday": d in holidays,
#             "holiday_name": holidays.get(d),
#             "is_saturday": d.weekday() == 5,
#             "is_sunday": d.weekday() == 6,
#         }


#         # -------- HOLIDAY --------
#         if info["is_holiday"]:
#             info["status"] = info["holiday_name"]
#         # -------- APPROVED LEAVE (NEW) --------
#         elif d in leave_dates:
#             info["status"] = "Leave"



#         # -------- WEEKEND --------
#         elif info["is_saturday"] or info["is_sunday"]:
#             info["status"] = "-"


#         # -------- ATTENDANCE EXISTS --------
#         elif att:

#             # only checked-in
#             if att.check_in and not att.check_out:
#                 info["status"] = "Working"

#             elif att.check_in and att.check_out:

#                 # make aware datetime
#                 cin  = datetime.combine(d, att.check_in)
#                 cout = datetime.combine(d, att.check_out)


#                 hours = (cout - cin).total_seconds() / 3600


#                 # -------- LATE CALC (SAFE) --------
#                 late = 0
#                 if shift and shift.start_time:
#                     shift_in = datetime.combine(d, shift.start_time)
#                     late = (cin - shift_in).total_seconds() / 60


#                 # -------- STATUS LOGIC --------
#                 if policy:

#                     if hours < (policy.work_hours_required / 2):
#                         info["status"] = "Half Day"

#                     elif late > policy.late_after_minutes:
#                         info["status"] = "Late"

#                     elif hours >= policy.work_hours_required:
#                         info["status"] = "Present"

#                     else:
#                         info["status"] = "Partial"

#                 else:
#                     info["status"] = "Present"


#         # -------- PAST — NO ATTENDANCE --------
#         elif d < today:
#             info["status"] = "Absent"


#         # -------- FUTURE --------
#         else:
#             info["status"] = "-"


#         # ---------- COUNTERS ----------
#         # ---------- COUNTERS (UPDATED) ----------
#         if (
#             not info["is_holiday"]
#             and not info["is_saturday"]
#             and not info["is_sunday"]
#             and info["status"] != "Leave"
#         ):
#             working_days += 1

#         if info["status"] in ["Present", "Late", "Half Day", "Partial"]:
#             present_count += 1

#         if info["status"] == "Absent":
#             absent_count += 1


#         days.append(info)


#     return render(request, "attendance/attendance/monthly_attendance.html", {
#         "days": days,
#         "month": month,
#         "year": year,
#         "working_days": working_days,
#         "present_count": present_count,
#         "absent_count": absent_count,
#         "target": target_user
#     })



# attendance/views.py

# attendance/views.py

# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo
from calendar import monthrange
from datetime import date, datetime, timedelta
from django.db.models import Q

@login_required
def monthly_attendance(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    # Target user pehla fetch karo (Main DB mathi profile malse)
    target_user_id = request.GET.get("user")
    if target_user_id:
        target_user = get_object_or_404(User, id=target_user_id)
    else:
        target_user = request.user

    try:
        profile = target_user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Monthly Report switching to {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. PERMISSION LOGIC (Sacha DB na context ma check thase)
    if target_user_id:
        if not request.user.is_superuser and request.user != target_user:
            if request.user.profile.company != target_user.profile.company:
                return HttpResponse("Access Denied: Different Company")

            is_squad_member = False
            try:
                if profile.team_leader == request.user:
                    is_squad_member = True
            except AttributeError:
                pass 

            if request.user.profile.role == "EMPLOYEE" and not is_squad_member:
                return HttpResponse("Access Denied")

    # 2. Date & Policy Setup (Fetch from Company DB)
    today = date.today()
    try:
        month = int(request.GET.get("month", today.month))
        year = int(request.GET.get("year", today.year))
    except ValueError:
        month = today.month
        year = today.year

    # ✅ AA BADHI QUERIES HAVE COMPANY DB MA JASE
    policy = AttendancePolicy.objects.filter(company=company).first()
    week_off_list = []  
    if policy and policy.week_off_days:
        week_off_list = [int(day) for day in policy.week_off_days.split(',') if day.isdigit()]
    else:
        week_off_list = [6] # Default Sunday

    shift = profile.shift
    days_in_month = monthrange(year, month)[1]

    # Attendance Records (From Company DB)
    attendance_map = {
        a.date: a
        for a in Attendance.objects.filter(
            user=target_user,
            date__year=year,
            date__month=month
        )
    }

    # Holidays (From Company DB)
    holiday_qs = Holiday.objects.filter(Q(company=company) | Q(company__isnull=True))
    holidays = {h.date: h.name for h in holiday_qs}

    # Approved Leaves (From Company DB)
    start_date = date(year, month, 1)
    end_date = date(year, month, days_in_month)
    approved_leaves = Leave.objects.filter(
        user=target_user,
        status='APPROVED',
        from_date__lte=end_date,
        to_date__gte=start_date
    )

    days = []
    # 3. Days Loop (Logic preserved as per your code)
    for i in range(1, days_in_month + 1):
        d = date(year, month, i)
        att = attendance_map.get(d)
        is_week_off_day = d.weekday() in week_off_list

        is_on_leave = False
        for leave in approved_leaves:
            if leave.from_date <= d <= leave.to_date:
                is_on_leave = True
                break

        day_info = {
            "date": d,
            "check_in": att.check_in if att else None,
            "check_out": att.check_out if att else None,
            "status": None,
            "is_holiday": d in holidays,
            "holiday_name": holidays.get(d),
            "is_week_off": is_week_off_day,
            "is_sunday": d.weekday() == 6,
            "is_saturday": d.weekday() == 5,
        }

        # Priority status logic (Maintained from original)
        if att:
            if att.check_in and not att.check_out:
                day_info["status"] = "Working"
            elif att.check_in and att.check_out:
                check_in_dt = datetime.combine(d, att.check_in)
                check_out_dt = datetime.combine(d, att.check_out)
                work_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
                late_minutes = 0
                if shift:
                    shift_in = datetime.combine(d, shift.start_time)
                    late_minutes = (check_in_dt - shift_in).total_seconds() / 60

                if policy:
                    if work_hours < (policy.work_hours_required / 2):
                        day_info["status"] = "Half Day"
                    elif late_minutes > policy.late_after_minutes:
                        day_info["status"] = "Late"
                    elif work_hours >= policy.work_hours_required:
                        day_info["status"] = "Present"
                    else:
                        day_info["status"] = "Partial"
                else:
                    day_info["status"] = "Present"
        elif d in holidays:
            day_info["status"] = "Holiday"
        elif is_on_leave:
            day_info["status"] = "Leave"
        else:
            if d > today:
                day_info["status"] = "-"
            elif is_week_off_day: 
                 day_info["status"] = "Week Off"
            else:
                 day_info["status"] = "Absent"

        days.append(day_info)

    # 4. Accurate Counters Calculation
    total_present = sum(1 for d in days if d["status"] in ["Present", "Late", "Half Day", "Partial", "Working"])
    total_absent = sum(1 for d in days if d["status"] == "Absent") 
    total_holidays = sum(1 for d in days if d["is_holiday"])
    total_leaves = sum(1 for d in days if d["status"] == "Leave")
    
    working_days = sum(1 for d in days if not d["is_week_off"] and not d["is_holiday"] and d["status"] != "Leave")

    context = {
        "days": days,
        "month": month,
        "year": year,
        "target_user": target_user,
        "target": target_user,
        "present_count": total_present,
        "absent_count": total_absent,
        "working_days": working_days,
        "total_holidays": total_holidays,
        "total_leaves": total_leaves
    }

    return render(request, "attendance/attendance/monthly_attendance.html", context)


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import Notification

from .middleware import ThreadLocal # Import check kari lejo

@login_required
@require_POST
def mark_notifications_read(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0)
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        db_name = 'default'

    # Badha unread notifications ne sacha DB ma 'read' karo
    Notification.objects.using(db_name).filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'status': 'success'})


@login_required
def holiday_list(request):
    # Superadmin hamesha 'default' DB mathi badhi company na holidays jove
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")

    holidays = Holiday.objects.all().order_by('date')
    return render(request, "attendance/MISC/holiday_list.html", {"holidays": holidays})


@login_required
def company_add_holiday(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company_slug}_db"
    
    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":
        # ✅ SAVE IN COMPANY DB
        Holiday.objects.using(db_name).create(
            company=company,
            date=request.POST.get("date"),
            name=request.POST.get("name")
        )
        messages.success(request, "Holiday Added Successfully")
        return get_dashboard_redirect(request.user, company)
    
    return render(request, "attendance/MISC/add_holiday.html", {"company": company})


from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import Holiday, Company

@login_required
def company_delete_holiday(request, id):
    company = None
    db_name = 'default'
    
    # 1. Try to determine company from the logged-in user (For Owners)
    try:
        if hasattr(request.user, 'profile') and request.user.profile.company:
            company = request.user.profile.company
            db_name = f"{company.slug}_db"
    except Exception:
        pass

    holiday = None
    
    # 2. Look in the exact Company DB
    if company:
        holiday = Holiday.objects.using(db_name).filter(id=id).first()
        
    # 3. If not found (and user is a Superadmin), scan all databases to find it
    if not holiday and request.user.is_superuser:
        for c in Company.objects.using('default').all():
            temp_db = f"{c.slug}_db"
            try:
                holiday = Holiday.objects.using(temp_db).filter(id=id).first()
                if holiday:
                    db_name = temp_db
                    company = c
                    break
            except Exception:
                continue

    # 4. Handle if it's completely missing
    if not holiday:
        messages.error(request, "Error: Holiday not found. It may have already been deleted.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # 5. Security check
    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")
    
    # 6. Safely Delete from the correct database
    try:
        holiday.delete(using=db_name)
        messages.success(request, "Holiday Deleted Successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting holiday: {str(e)}")
        
    return get_dashboard_redirect(request.user, company)

@login_required
def shift_list(request):
    # HR mate potana company DB mathi shifts fetch karo
    if request.user.groups.filter(name__icontains='hr').exists() or request.user.is_superuser:
        company = request.user.profile.company
        db_name = f"{company.slug}_db"
        
        shifts = Shift.objects.using(db_name).filter(company=company)
        return render(request, 'attendance/MISC/shift_list.html', {'shifts': shifts}) 
    
    return render(request, 'attendance/MISC/access_denied.html')


@login_required
def company_add_shift(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company_slug}_db"

    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":
        # ✅ CREATE IN COMPANY DB
        Shift.objects.using(db_name).create(
            company=company,
            name=request.POST.get("name"),
            start_time=request.POST.get("start_time"),
            end_time=request.POST.get("end_time"),
            break_minutes=request.POST.get("break_minutes")
        )
        messages.success(request, "Shift Added Successfully")
        return get_dashboard_redirect(request.user, company)

    return render(request, "attendance/MISC/add_shift.html", {"company": company})


# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def company_edit_shift(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    # Pehla default mathi fetch karo jya company info che
    shift = get_object_or_404(Shift, id=id)
    company = shift.company
    db_name = f"{company.slug}_db"
    
    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")
    
    if request.method == "POST":
        shift.name = request.POST.get("name")
        shift.start_time = request.POST.get("start_time")
        shift.end_time = request.POST.get("end_time")
        shift.break_minutes = request.POST.get("break_minutes")
        
        # ✅ UPDATE IN COMPANY DB
        shift.save(using=db_name)
        
        messages.success(request, "Shift Updated Successfully")
        return get_dashboard_redirect(request.user, company)
        
    return render(request, "attendance/MISC/edit_shift.html", {"shift": shift})


@login_required
def company_edit_holiday(request, id):
    holiday = get_object_or_404(Holiday, id=id)
    company = holiday.company
    db_name = f"{company.slug}_db"

    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")
    
    if request.method == "POST":
        holiday.date = request.POST.get("date")
        holiday.name = request.POST.get("name")
        
        # ✅ UPDATE IN COMPANY DB
        holiday.save(using=db_name)
        
        messages.success(request, "Holiday Updated Successfully")
        return get_dashboard_redirect(request.user, company)
        
    return render(request, "attendance/MISC/edit_holiday.html", {"holiday": holiday})


from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from .models import AttendancePolicy, Company

@login_required
def company_edit_policy(request, id):
    company = None
    db_name = 'default'
    
    # 1. Try to determine company from the logged-in user (For Owners, Admins, HR)
    try:
        if hasattr(request.user, 'profile') and request.user.profile.company:
            company = request.user.profile.company
            db_name = f"{company.slug}_db"
    except Exception:
        pass

    policy = None
    
    # 2. Look in the exact Company DB
    if company:
        policy = AttendancePolicy.objects.using(db_name).filter(id=id).first()
        
    # 3. If not found (and user is a Superadmin), scan all databases to find it
    if not policy and request.user.is_superuser:
        for c in Company.objects.using('default').all():
            temp_db = f"{c.slug}_db"
            try:
                policy = AttendancePolicy.objects.using(temp_db).filter(id=id).first()
                if policy:
                    db_name = temp_db
                    company = c
                    break
            except Exception:
                continue

    # 4. Handle if it's completely missing
    if not policy:
        messages.error(request, "Error: Policy not found.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # 5. Security Check
    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")
    
    # 6. Handle Form Submission
    if request.method == "POST":
        policy.work_hours_required = request.POST.get("work_hours")
        policy.late_after_minutes = request.POST.get("late_after")
        policy.grace_time_minutes = request.POST.get("grace_time")
        
        # ✅ UPDATE SAFELY IN COMPANY DB
        policy.save(using=db_name)
        
        messages.success(request, "Policy Updated Successfully")
        return get_dashboard_redirect(request.user, company)
        
    return render(request, "attendance/MISC/edit_policy.html", {"policy": policy, "companies": [company]})
    
    
    
# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def company_delete_shift(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    shift = get_object_or_404(Shift, id=id)
    company = shift.company
    db_name = f"{company.slug}_db"

    if not is_owner_or_superuser(request.user, company):
        return HttpResponseForbidden("Access Denied")
    
    # ✅ DELETE FROM COMPANY DB
    shift.delete(using=db_name)
    
    messages.success(request, "Shift Deleted Successfully")
    return get_dashboard_redirect(request.user, company)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import User
from accounts.models import Profile
from attendance.models import Shift, Company
from attendance.middleware import ThreadLocal

@login_required
def hr_assign_shift(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        hr_profile = Profile.objects.using('default').get(user=request.user)
        company = hr_profile.company
        db_name = f"{company.slug}_db"
    except Exception:
        return HttpResponse("HR Profile Error")

    # 1. Target User/Profile fetch (Main DB mathi)
    target_user = get_object_or_404(User.objects.using('default'), id=user_id)
    target_profile_main = get_object_or_404(Profile.objects.using('default'), user_id=user_id)

    if request.method == "POST":
        shift_id_raw = request.POST.get("shift_id")
        
        try:
            if shift_id_raw:
                # 🔥 STEP 1: શિફ્ટને ટેનન્ટ ડીબી માંથી લાવો
                shift_tenant = Shift.objects.using(db_name).get(id=shift_id_raw)

                # 🔥 STEP 2: આ જ શિફ્ટને મેઈન (default) ડીબીમાં સિંક કરો (Error 1452 રોકવા માટે)
                # આનાથી જો મેઈન ડીબીમાં એ શિફ્ટ નહીં હોય તો તે બની જશે
                shift_main, created = Shift.objects.using('default').get_or_create(
                    id=shift_tenant.id,
                    defaults={
                        'company': company,
                        'name': shift_tenant.name,
                        'start_time': shift_tenant.start_time,
                        'end_time': shift_tenant.end_time,
                        'break_minutes': shift_tenant.break_minutes
                    }
                )
                
                # જો નામ કે સમય બદલાયો હોય તો અપડેટ કરો
                if not created:
                    shift_main.name = shift_tenant.name
                    shift_main.start_time = shift_tenant.start_time
                    shift_main.end_time = shift_tenant.end_time
                    shift_main.save(using='default')

                new_shift_id = shift_main.id
            else:
                new_shift_id = None

            # ---------------------------------------------------------
            # ✅ STEP 3: હવે બંને ડેટાબેઝમાં પ્રોફાઈલ અપડેટ કરો
            # ---------------------------------------------------------
            # A. Update in Company DB
            Profile.objects.using(db_name).filter(user_id=user_id).update(shift_id=new_shift_id)

            # B. Update in Main DB (હવે એરર નહીં આવે કારણ કે શિફ્ટ ત્યાં હાજર છે!)
            Profile.objects.using('default').filter(user_id=user_id).update(shift_id=new_shift_id)

            messages.success(request, f"Shift successfully assigned to {target_user.username}")

        except Exception as e:
            messages.error(request, f"Shift Assignment Failed: {str(e)}")
            
        return redirect('hr_employee_list')

    # 📋 FETCH SHIFTS (From Company DB for dropdown display)
    shifts = Shift.objects.using(db_name).filter(company=company)
    
    return render(request, "attendance/hr/hr_assign_shift.html", {
        "target_user": target_user,
        "shifts": shifts,
        "current_shift": target_profile_main.shift,
    })


from .middleware import ThreadLocal # Import check kari lejo

# 1. ADMIN LIST
@login_required
def superadmin_admin_list(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    # Superadmin hamesha 'default' mathi profiles jove che (Master Profiles)
    # Pan sacho data fetch karva 'using' logic use thase jo profiles split hoy to
    admins = Profile.objects.filter(user__groups__name__icontains="admin").select_related('user', 'company')

    return render(request, "attendance/superadmin_admin_list.html", {
        "users": admins,
        "title": "Admin Users"
    })

# 2. HR LIST
@login_required
def superadmin_hr_list(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    hr_users = Profile.objects.filter(user__groups__name__icontains="HR").select_related('user', 'company')

    return render(request, "attendance/superadmin_hr_list.html", {
        "users": hr_users,
        "title": "HR Users"
    })

# 3. MANAGER LIST
@login_required
def superadmin_manager_list(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    # Note: 'company' filter mate URL mathi company_id levu pade jo specific list jovanu hoy
    managers = Profile.objects.filter(user__groups__name__icontains="MANAGER").select_related('user', 'company')

    return render(request, "attendance/superadmin/superadmin_manager_list.html", {
        "users": managers,
        "title": "Managers"
    })

# 4. EMPLOYEE LIST
@login_required
def superadmin_employee_list(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    employees = Profile.objects.filter(user__groups__name__icontains="EMPLOYEE").select_related('user', 'company')

    return render(request, "attendance/superadmin_employee_list.html", {
        "users": employees,
        "title": "Employees"
    })

# 5. MANAGER TEAM (Strictly Multi-tenant)
@login_required
def superadmin_manager_team(request, manager_id):
    # Target manager main DB mathi fetch karo
    manager_user = get_object_or_404(User, id=manager_id)
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        # Manager ni company mujab DB select karo
        company = manager_user.profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    if not user_has_permission(request.user, "can_view_team"):
        return HttpResponse("Access Denied")

    # ✅ FETCH TEAM FROM COMPANY DB
    team = Profile.objects.using(db_name).filter(manager=manager_user)

    return render(request, "attendance/superadmin/superadmin_manager_team.html", {
        "manager": manager_user,
        "team": team
    })
from .middleware import ThreadLocal
from datetime import date

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.models import User
# Profile 'accounts' માં છે અને Company 'attendance' માં છે
from accounts.models import Profile 
from .models import Company 

@login_required
def superadmin_company_overview(request):
    # 1. SECURITY CHECK
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied: SuperAdmin Access Required.")

    # 2. FETCH ALL COMPANIES (attendance.models માંથી)
    companies = Company.objects.using('default').all()
    data = []

    for c in companies:
        slug_lower = c.slug.lower()
        
        # પાયાની ક્વેરી: આ કંપનીના તમામ પ્રોફાઇલ્સ (Main DB માંથી)
        company_profiles = Profile.objects.using('default').filter(company=c)
        
        stats = {
            "company": c,
            "admins": 0,
            "hr": 0,
            "managers": 0,
            "team_leaders": 0,
            "employees_only": 0,
            "total_staff": company_profiles.count(),
        }

        # 🛡️ ROLE COUNTING logic
        for p in company_profiles.select_related('user').prefetch_related('user__groups'):
            user_groups = [g.name.lower() for g in p.user.groups.all()]
            
            # Owner ને સ્કીપ કરો
            if f"{slug_lower}_owner" in user_groups:
                continue

            if f"{slug_lower}_admin" in user_groups:
                stats["admins"] += 1
            elif f"{slug_lower}_hr" in user_groups:
                stats["hr"] += 1
            elif f"{slug_lower}_manager" in user_groups:
                stats["managers"] += 1
            elif f"{slug_lower}_team leader" in user_groups:
                stats["team_leaders"] += 1
            else:
                stats["employees_only"] += 1

        data.append(stats)

    return render(request, "attendance/superadmin/company_overview.html", {
        "data": data,
        "total_companies": companies.count()
    })
@login_required
def superadmin_company_details(request, company_id):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    # Company info main DB mathi
    company = get_object_or_404(Company, id=company_id)
    db_name = f"{company.slug}_db"
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    ThreadLocal.DB_NAME = db_name

    # Fetch Profiles from Company DB
    admins = Profile.objects.using(db_name).filter(role="ADMIN")
    hr_users = Profile.objects.using(db_name).filter(role="HR")
    managers = Profile.objects.using(db_name).filter(role="MANAGER")
    employees = Profile.objects.using(db_name).filter(role="EMPLOYEE")

    # --- BASIC ANALYTICS (Today) ---
    today = date.today()

    # ✅ FETCH ATTENDANCE FROM COMPANY DB
    # Attendance model company na isolated DB ma hoy che
    present = Attendance.objects.using(db_name).filter(
        date=today
    ).count()

    total_emp = employees.count()
    absent = max(0, total_emp - present) # Negative value na ave te mate safety

    data = {
        "company": company,
        "admins": admins,
        "hr_users": hr_users,
        "managers": managers,
        "employees": employees,
        "total_emp": total_emp,
        "present": present,
        "absent": absent
    }

    # Context reset for next request
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/superadmin/company_details.html", data)


from django.contrib.auth.models import Group
from .middleware import ThreadLocal

@login_required
def superadmin_role_list(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    # Groups hamesha main DB mathi j fetch thase
    roles = Group.objects.all()

    return render(request, "attendance/MISC/role_list.html", {
        "roles": roles
    })

@login_required
def superadmin_add_role(request):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    if request.method == "POST":
        name = request.POST.get("role_name")
        if name:
            # Main DB ma group banao (Login/Permissions mate)
            Group.objects.get_or_create(name=name.upper())

    return redirect("superadmin_role_list")

@login_required
def superadmin_delete_role(request, id):
    if not is_superadmin(request.user):
        return HttpResponse("Access Denied")

    # Main DB mathi delete karo
    Group.objects.filter(id=id).delete()

    return redirect("superadmin_role_list")

@login_required
def superadmin_permissions(request):
    # Security Check
    try:
        profile = request.user.profile
        if profile.role.upper() != "SUPERADMIN":
            return HttpResponse("Access Denied")
        
        # Superadmin potani company (Default) na permissions manage kare che
        # Pan jo specific company na karva hoy to URL mathi slug levu pade
        db_name = 'default' 
    except:
        return HttpResponse("Profile Error")

    roles = Group.objects.all()

    if request.method == "POST":
        for r in roles:
            # ✅ Save Permission in Specific DB
            rp, created = RolePermission.objects.using(db_name).get_or_create(group=r)

            rp.can_manage_users = request.POST.get(f"manage_{r.id}") == "on"
            rp.can_approve_attendance = request.POST.get(f"approve_{r.id}") == "on"
            rp.can_view_team = request.POST.get(f"team_{r.id}") == "on"
            rp.can_view_reports = request.POST.get(f"report_{r.id}") == "on"
            rp.can_self_access = request.POST.get(f"self_{r.id}") == "on"

            rp.save(using=db_name)

        messages.success(request, "Global Permissions Updated Successfully")
        return redirect("superadmin_dashboard")

    data = []
    for r in roles:
        # Load from Specific DB
        rp, created = RolePermission.objects.using(db_name).get_or_create(group=r)
        data.append((r, rp))

    return render(request, "attendance/MISC/permissions.html", {
        "data": data
    })


# views.py માં user_has_permission ફંક્શન આ મુજબ બદલો
# attendance/views.py માં આ ફંક્શન શોધો અને રિપ્લેસ કરો

def user_has_permission(user, perm_name):
    if user.is_superuser:
        return True
    
    # અહીં જૂનો કોડ 'from .models import Permission' કાઢી નાખ્યો છે
    # હવે આપણે ગ્લોબલ ઈમ્પોર્ટ કરેલ 'RolePermission' વાપરીશું
    
    for group in user.groups.all():
        try:
            # અહીં Permission ની જગ્યાએ RolePermission વાપરો
            rp = RolePermission.objects.get(group=group)
            if getattr(rp, perm_name, False):
                return True
        except RolePermission.DoesNotExist:
            continue
            
    return False



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import Group, User
from accounts.models import Profile

from .middleware import ThreadLocal # Import check kari lejo

@login_required
def admin_assign_role(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        admin_branch = profile_admin.branch 
        
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
        print(f"DEBUG: Assigning role in {db_name}")
    except:
        ThreadLocal.DB_NAME = 'default'
        db_name = 'default'

    # 1. PERMISSION CHECK
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    if not (request.user.is_superuser or is_admin or user_has_permission(request.user, "can_manage_users")):
        return HttpResponse("Access Denied: You do not have permission to assign roles.")

    # 3. GET TARGET USER & SECURITY CHECKS
    # User object hamesha main/default mathi j ave (Auth purpose mate)
    target_user = get_object_or_404(User, id=user_id)
    # Profile fetch from Company DB
    target_profile = get_object_or_404(Profile.objects.using(db_name), user=target_user)

    # Security: Company Check
    if target_profile.company != company:
        return HttpResponse("Access Denied: User belongs to another company.")

    # Branch Security
    if admin_branch and target_profile.branch != admin_branch:
        return HttpResponse(f"Access Denied: You can only assign roles to employees in {admin_branch.name}.")

    # 4. FETCH ROLES (Groups main DB mathi ave)
    roles = Group.objects.filter(name__startswith=f"{company.slug}_")

    # 5. ASSIGN ROLE LOGIC
    if request.method == "POST":
        group_id = request.POST.get("role")

        if group_id:
            try:
                selected_group = Group.objects.get(id=group_id)
                
                if not selected_group.name.startswith(f"{company.slug}_"):
                    messages.error(request, "Invalid Role Selection.")
                    return redirect("admin_employee_list")

                # 🔥 STEP 1: CLEANUP (Cleanup function sacha DB context ma chalse)
                cleanup_user_roles_and_links(target_user)
                
                # 🔥 STEP 2: ASSIGN GROUP (Main DB operation)
                target_user.groups.add(selected_group)
                
                # ✅ STEP 3: UPDATE PROFILE IN COMPANY DB
                clean_role_name = selected_group.name.split('_')[-1].capitalize()
                target_profile.role = clean_role_name
                target_profile.save(using=db_name) # Fourcefully company DB ma save karo

                messages.success(request, f"Role updated to '{clean_role_name}' successfully.")

            except Group.DoesNotExist:
                messages.error(request, "Role not found.")
        
        return redirect("admin_employee_list")

    return render(request, "attendance/admin/admin_assign_role.html", {
        "profile": target_profile,
        "roles": roles,
        "current_role": target_user.groups.first() 
    })


from .middleware import ThreadLocal
from django.contrib import messages

@login_required
def superadmin_unassigned_users(request):
    # Security check
    if request.user.profile.role.upper() != "SUPERADMIN":
        return HttpResponse("Access Denied")

    # Unassigned users hamesha 'default' database ma j hashe kem ke haji emni koi company nathi
    users = Profile.objects.filter(
        is_verified=True,
        company__isnull=True  # 🔥 Only users without a company
    )

    return render(request, "attendance/superadmin/superadmin_unassigned_users.html", {
        "users": users
    })

@login_required
def superadmin_assign_company(request, user_id):
    if request.user.profile.role.upper() != "SUPERADMIN":
        return HttpResponse("Access Denied")

    # Target profile main DB mathi fetch karo
    profile = get_object_or_404(Profile, user__id=user_id)
    companies = Company.objects.all()

    if request.method == "POST":
        company_id = request.POST.get("company")

        if company_id:
            company = get_object_or_404(Company, id=company_id)
            
            # Step 1: Assign Company in Main DB
            profile.company = company
            profile.save() # default DB ma save thase
            
            # Step 2: (Advanced Level 0) Profile ni copy company database ma pan create thavi joie
            # db_name = f"{company.slug}_db"
            # profile.save(using=db_name)

            messages.success(request, f"User {profile.user.username} assigned to {company.name}")
            return redirect("superadmin_unassigned_users")

    return render(request, "attendance/superadmin/superadmin_assign_company.html", {
        "profile": profile,
        "companies": companies
    })


@login_required
def manage_role_permissions(request, company_slug, group_id):
    # 1. Superuser check
    if not request.user.is_superuser:
        return HttpResponse("Access Denied: Only Superuser allowed.")

    company = get_object_or_404(Company, slug=company_slug)
    role_group = get_object_or_404(Group, id=group_id)
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name # Queries have aa company na DB ma jase

    # 2. Security Check: Company-Role logic
    if not role_group.name.startswith(f"{company.slug}_"):
        return HttpResponse("Invalid Role for this Company")

    # 3. Get or Create Permission in the COMPANY DATABASE
    perm, created = RolePermission.objects.using(db_name).get_or_create(group=role_group)

    if request.method == "POST":
        perm.can_manage_users = request.POST.get("can_manage_users") == "on"
        perm.can_approve_attendance = request.POST.get("can_approve_attendance") == "on"
        perm.can_view_team = request.POST.get("can_view_team") == "on"
        perm.can_view_reports = request.POST.get("can_view_reports") == "on"
        perm.can_self_access = request.POST.get("can_self_access") == "on"
        perm.can_view_policy = request.POST.get("can_view_policy") == "on"
        
        # ✅ Save in Company DB
        perm.save(using=db_name)
        
        messages.success(request, f"Permissions updated for {role_group.name} in {company.name}")
        return redirect('superadmin_company_dashboard', company_id=company.id)

    # 4. Data for Template
    data = [(role_group, perm)]

    return render(request, "attendance/MISC/permissions.html", {
        "data": data,
        "company": company
    })
# attendance/views.py
# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import Group, User
from accounts.models import Profile

@login_required
def manager_total_staff_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન ડીબી માંથી મેનેજરની પ્રોફાઈલ લાવો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        manager_branch = profile.branch
        db_name = f"{company.slug}_db"
    except Profile.DoesNotExist:
        return redirect('login')

    # 1. Security Check
    if not request.user.groups.filter(name__icontains='manager').exists():
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 📋 2. FETCH STAFF LOGIC (Using Main DB for 100% Visibility)
    # ---------------------------------------------------------
    # આપણે 'default' DB વાપરીશું જેથી User અને Groups સાથેના ફિલ્ટર્સ કામ કરે
    staff_profiles = Profile.objects.using('default').filter(
        manager=request.user,
        company=company,
        branch=manager_branch 
    ).select_related('user').order_by('user__username')

    # રોલ ચેક કરવા માટે ગ્રુપનું નામ સેટ કરો
    slug_lower = company.slug.lower()
    tl_group_name = f"{slug_lower}_team leader"

    staff_data = []
    for member in staff_profiles:
        # મેઈન ડીબી માં ગ્રુપ ચેક કરો (Bulletproof Logic)
        is_tl = member.user.groups.filter(name__iexact=tl_group_name).exists()
        
        staff_data.append({
            'user': member.user,
            'role': "Team Leader" if is_tl else "Employee",
            'is_tl': is_tl,
            'email': member.user.email,
            'status': 'Active' if member.user.is_active else 'Inactive',
            'profile': member # બ્રાન્ચ કે અન્ય વિગતો માટે
        })

    # ---------------------------------------------------------
    # 3. DB RESET & RENDER
    # ---------------------------------------------------------
    # રિકવેસ્ટ પૂરી થતા પહેલા ખાતરી કરો કે DB સાચો સેટ છે
    from attendance.middleware import ThreadLocal
    ThreadLocal.DB_NAME = db_name # Template માં જો કોઈ ડેટા કંપની DB માંથી જોઈતો હોય તો

    return render(request, "attendance/MISC/staff_list.html", {
        "page_title": "Total Staff Directory",
        "staff_list": staff_data,
        "is_attendance_view": False,
        "current_branch": manager_branch
    })
@login_required
def manager_present_today_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # Get Manager's Profile from Main DB
        profile = Profile.objects.using('default').get(user=request.user)
        company = profile.company
        manager_branch = profile.branch
        db_name = f"{company.slug}_db"
        
        # Ensure ThreadLocal is set for the Company DB
        ThreadLocal.DB_NAME = db_name
    except Profile.DoesNotExist:
        return redirect('login')

    if not request.user.groups.filter(name__icontains='manager').exists():
        return HttpResponse("Access Denied")

    today = timezone.localdate()

    # ---------------------------------------------------------
    # 🔥 STEP 1: GET TEAM IDs FROM DEFAULT DB (The Truth Source)
    # ---------------------------------------------------------
    # We find all employees where the manager is the logged-in user.
    # We use 'default' DB because the manager-employee relationship is stored there.
    team_user_ids = list(Profile.objects.using('default').filter(
        manager=request.user,
        company=company,
        branch=manager_branch
    ).values_list('user_id', flat=True))

    # Debug print to your console to see if IDs are being found
    print(f"DEBUG: Found Team IDs for Manager {request.user.username}: {team_user_ids}")

    # ---------------------------------------------------------
    # 📅 STEP 2: FETCH ATTENDANCE FROM COMPANY DB (Using the ID list)
    # ---------------------------------------------------------
    # Now we query the specific company database for today's entries.
    # We EXCLUDE the manager (request.user.id) so it only shows employees.
    present_records = Attendance.objects.using(db_name).filter(
        date=today,
        user_id__in=team_user_ids,
        check_in__isnull=False
    ).exclude(user_id=request.user.id).select_related('user').order_by('check_in')

    # ---------------------------------------------------------
    # 📦 STEP 3: PREPARE DATA FOR TEMPLATE
    # ---------------------------------------------------------
    staff_data = []
    slug_lower = company.slug.lower()
    tl_group_name = f"{slug_lower}_team leader"

    for record in present_records:
        # Time formatting
        c_in = record.check_in.strftime("%I:%M %p") if record.check_in else "-"
        
        # Role Check (Are they a TL or just an Employee?)
        # Groups live in the 'default' DB
        is_tl = record.user.groups.filter(name__iexact=tl_group_name).exists()
        
        staff_data.append({
            'user': record.user,
            'role': "Team Leader" if is_tl else "Employee",
            'check_in_time': c_in,
            'email': record.user.email,
            'status': record.status or "Present"
        })

    # Safety Reset for database context
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/MISC/manager_present_today.html", {
        "page_title": f"Present Staff Today",
        "staff_list": staff_data,
        "is_attendance_view": True,
        "current_branch": manager_branch
    })    # 


# attendance/views.py


@login_required
def superadmin_company_dashboard(request, company_id):
    # 🛡️ Security Check
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied.")

    # 🏢 Company Setup
    company = get_object_or_404(Company.objects.using('default'), id=company_id)
    db_name = f"{company.slug}_db"
    slug_lower = company.slug.lower()
    
    # 🔄 Database Context Switching
    from attendance.middleware import ThreadLocal
    ThreadLocal.DB_NAME = db_name 

    # 1. FETCH ALL PROFILES FROM MAIN DB (Exclude Owners)
    # આપણે અહીં જ માલિકને કાઢી નાખીશું જેથી ગણતરીમાં ભૂલ ના થાય
    all_staff_profiles = Profile.objects.using('default').filter(
        company=company
    ).exclude(user__groups__name__icontains='owner').select_related('user', 'branch').prefetch_related('user__groups')

    # 🔥 TOTAL STAFF COUNT (બ્રાન્ચ + હેડ ઓફિસ બધું જ આવી ગયું)
    total_staff_count = all_staff_profiles.count()

    # 2. ROLE WISE & BRANCH WISE SEGREGATION
    admins_count = 0
    hr_count = 0
    managers_count = 0
    ho_employees = [] # Head Office (No Branch)

    for p in all_staff_profiles:
        user_groups = [g.name.lower() for g in p.user.groups.all()]
        
        # ગ્રાફ માટે રોલ મુજબ ગણતરી
        if f"{slug_lower}_admin" in user_groups:
            admins_count += 1
        elif f"{slug_lower}_hr" in user_groups:
            hr_count += 1
        elif f"{slug_lower}_manager" in user_groups:
            managers_count += 1
        else:
            # જો કોઈ મેનેજમેન્ટ રોલ નથી અને બ્રાન્ચ પણ નથી, તો HO Directory માં એડ કરો
            if not p.branch_id:
                ho_employees.append(p)

    # 3. BRANCH TABLE LOGIC (Fixed Count)
    from .models import Branch
    branches_qs = Branch.objects.using(db_name).all()
    branches_with_count = []
    
    for branch in branches_qs:
        # આ પર્ટીક્યુલર બ્રાન્ચના પ્રોફાઈલ્સ ગણો (Main DB માંથી)
        b_count = all_staff_profiles.filter(branch_id=branch.id).count()
        branch.manual_staff_count = b_count
        branches_with_count.append(branch)

    # 4. OPERATIONAL DATA
    from .models import Shift, Holiday, AttendancePolicy
    from django.contrib.auth.models import Group
    shifts = Shift.objects.using(db_name).all()
    holidays = Holiday.objects.using(db_name).all()
    policies = AttendancePolicy.objects.using(db_name).all()
    roles = Group.objects.using('default').filter(name__startswith=f"{company.slug}_").exclude(name__icontains='owner')

    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/dashboard/superadmin_company_dashboard.html", {
        "company": company,
        "company_owner": Profile.objects.using('default').filter(company=company, user__groups__name__icontains='owner').first(),
        "admins_count": admins_count,
        "hr_count": hr_count,
        "managers_count": managers_count,
        "ho_emp_count": len(ho_employees),
        "branches": branches_with_count, 
        "employees": ho_employees, 
        "total_staff_count": total_staff_count,
        "policies": policies,
        "shifts": shifts,
        "holidays": holidays,
        "roles": roles,
    })    
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.models import Group
from .models import Company, RolePermission
from .utils import ensure_db_connection

@login_required
def manage_all_role_permissions(request, company_slug):
    """કંપનીના તમામ રોલ્સની પરમિશન એકસાથે મેનેજ કરવા માટે (Optimized Level 0)"""
    
    # ૧. કંપની અને ડેટાબેઝ સેટઅપ
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    
    # 🛡️ રિફ્રેશ અને કનેક્શન સેફ્ટી
    ensure_db_connection(db_name)

    # ૨. મેઈન ડેટાબેઝમાંથી આ કંપનીના ગ્રુપ્સ લાવો
    groups = Group.objects.filter(name__startswith=f"{company.slug}_")

    if request.method == "POST":
        try:
            # 🛡️ ટ્રાન્ઝેક્શન ટેનન્ટ DB માં - ડેટા ઇન્ટિગ્રિટી માટે
            with transaction.atomic(using=db_name):
                for group in groups:
                    # ૩. ટેનન્ટ DB માં ગ્રુપની હાજરી ખાતરી કરો (Constraint Fix)
                    tenant_group, _ = Group.objects.using(db_name).get_or_create(
                        name=group.name
                    )

                    # ૪. પરમિશન ઓબ્જેક્ટ મેળવો (group ફિલ્ડ સાથે)
                    perm, _ = RolePermission.objects.using(db_name).get_or_create(
                        group=tenant_group
                    )
                    
                    # ૫. ડાયનેમિકલી પરમિશન અપડેટ (Checkbox Logic)
                    perm.can_manage_users = request.POST.get(f'can_manage_users_{group.id}') == 'on'
                    perm.can_approve_attendance = request.POST.get(f'can_approve_attendance_{group.id}') == 'on'
                    perm.can_approve_leave = request.POST.get(f'can_approve_leave_{group.id}') == 'on'
                    perm.can_manage_shifts = request.POST.get(f'can_manage_shifts_{group.id}') == 'on'
                    perm.can_view_reports = request.POST.get(f'can_view_reports_{group.id}') == 'on'
                    perm.can_view_team = request.POST.get(f'can_view_team_{group.id}') == 'on'
                    perm.can_self_access = request.POST.get(f'can_self_access_{group.id}') == 'on'
                    
                    # ટેનન્ટ DB માં સેવ કરો
                    perm.save(using=db_name)
            
            messages.success(request, f"✅ Permissions for all roles in {company.name} updated successfully.")
            return redirect('superadmin_company_dashboard', company_id=company.id)

        except Exception as e:
            messages.error(request, f"❌ Error updating permissions: {str(e)}")
            return redirect('superadmin_company_dashboard', company_id=company.id)

    # ૬. GET રિક્વેસ્ટ: ટેમ્પલેટ માટે ડેટા તૈયાર કરો
    data = []
    for g in groups:
        # ખાતરી કરો કે ડિસ્પ્લે વખતે પણ ટેનન્ટ DB માં રેકોર્ડ હોય
        tg, _ = Group.objects.using(db_name).get_or_create(name=g.name)
        p, _ = RolePermission.objects.using(db_name).get_or_create(group=tg)
        data.append((g, p))

    return render(request, 'attendance/MISC/permissions.html', {
        'company': company, 
        'data': data
    })

# @login_required
# def company_delete_policy(request, policy_id):
#     # Fetch from Main DB context initially
#     policy = get_object_or_404(AttendancePolicy, id=policy_id)
#     company = policy.company
#     db_name = f"{company.slug}_db"

#     # ✅ Force delete from specific company DB
#     policy.delete(using=db_name)
    
#     messages.success(request, "Attendance policy deleted successfully.")
#     return redirect('superadmin_company_dashboard', company_id=company.id)


from .middleware import ThreadLocal

@login_required
def company_admin_pending_users(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Admin Profile Error")

    if profile_admin.role != "ADMIN":
        return HttpResponse("Access Denied")

    # Fetch users from Company DB
    users = Profile.objects.using(db_name).filter(
        company=company,
        is_verified=True,
        is_approved=False
    )

    return render(request, "attendance/admin/company_admin_pending_users.html", {"users": users})


@login_required
def company_admin_approve_user(request, user_id):
    profile_admin = request.user.profile
    if profile_admin.role != "ADMIN":
        return HttpResponse("Access Denied")

    company = profile_admin.company
    db_name = f"{company.slug}_db"

    # ✅ STEP 1: Update in Main/Default DB
    profile = get_object_or_404(Profile, user__id=user_id)
    profile.is_approved = True
    profile.save()

    # ✅ STEP 2: Sync to Company DB (Level 0)
    # Aa jaroori che kem ke Dashboard Company DB mathi status check kare che
    profile.save(using=db_name)

    messages.success(request, f"User {profile.user.username} approved successfully.")
    return redirect("company_admin_pending_users")


@login_required
def company_admin_reject_user(request, user_id):
    profile_admin = request.user.profile
    if profile_admin.role != "ADMIN":
        return HttpResponse("Access Denied")

    company = profile_admin.company
    db_name = f"{company.slug}_db"

    # Get profile from main DB
    profile = get_object_or_404(Profile, user__id=user_id)

    # ✅ STEP 1: Update in Main DB (Remove link)
    profile.company = None
    profile.is_approved = False
    profile.save()

    # ✅ STEP 2: Also remove from Company DB if exists
    try:
        profile_in_company_db = Profile.objects.using(db_name).get(user__id=user_id)
        profile_in_company_db.delete()
    except:
        pass

    messages.warning(request, "User request rejected and removed from company list.")
    return redirect("company_admin_pending_users")


@login_required
def superadmin_make_admin(request, user_id):
    if request.user.profile.role.upper() != "SUPERADMIN":
        return HttpResponse("Access Denied")

    # Main DB mathi user fetch karo
    profile = get_object_or_404(Profile, user__id=user_id)
    company = profile.company
    
    if company:
        db_name = f"{company.slug}_db"
        
        # 1. Group Logic (Hamesha Main DB ma thashe)
        profile.user.groups.clear()
        admin_group, _ = Group.objects.get_or_create(name=f"{company.slug}_Admin")
        profile.user.groups.add(admin_group)

        # 2. Update Profile Role
        profile.role = "ADMIN"
        profile.is_approved = True
        
        # ✅ Save in BOTH Databases
        profile.save() # Main DB
        profile.save(using=db_name) # Company DB for Dashboard Consistency

    messages.success(request, f"{profile.user.username} is now Admin of {company.name}")
    return redirect("superadmin_unassigned_users")

# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import Group, User
from accounts.models import Profile

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import Group, User
from accounts.models import Profile
from .middleware import ThreadLocal

# --- ૧. માસ્ટર ક્લીનઅપ ફંક્શન (બધું જ ફ્રી કરવા માટે) ---
def cleanup_user_roles_and_links(user_obj, db_name):
    """યુઝરનો રોલ બદલાતા પહેલા તેના જૂના લિંક્સ સાફ કરવા, બ્રાન્ચ સાચવીને."""
    user_id = user_obj.id

    # ૧. જો આ યુઝર TL હતો, તો તેની ટીમને ફ્રી કરો
    Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader_id=None)
    Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader_id=None)

    # ૨. જો આ યુઝર મેનેજર હતો, તો તેના એમ્પ્લોઈઝને ફ્રી કરો
    Profile.objects.using('default').filter(manager_id=user_id).update(manager_id=None)
    Profile.objects.using(db_name).filter(manager_id=user_id).update(manager_id=None)

    # ૩. આ યુઝર પોતે જો કોઈની નીચે હોય, તો તેને ત્યાંથી હટાવો (બ્રાન્ચને ટચ કર્યા વગર)
    # 🔥 અહીં .update() વાપરવાથી branch_id એમને એમ જ રહેશે
    Profile.objects.using('default').filter(user_id=user_id).update(team_leader_id=None, manager_id=None)
    Profile.objects.using(db_name).filter(user_id=user_id).update(team_leader_id=None, manager_id=None)

    # ૪. જૂના ગ્રુપ્સ ક્લિયર કરો
    user_obj.groups.clear()

# --- ૨. મેઈન વ્યુ ફંક્શન ---
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import Group, User
from django.contrib import messages
from accounts.models import Profile
from .middleware import ThreadLocal
# તમારી હેલ્પર મેથડ અહી ઇમ્પોર્ટ હોવી જોઈએ
# from .utils import cleanup_user_roles_and_links 

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import Group, User
from django.contrib import messages
from accounts.models import Profile
from .middleware import ThreadLocal

@login_required
def admin_manage_roles(request):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        profile_admin = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile_admin.company
        admin_branch = profile_admin.branch # For rendering in the template
        admin_branch_id = profile_admin.branch_id # 🔥 SAFE ID FOR COMPARISON
        
        if company:
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
        else:
            db_name = 'default'
    except Exception:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Admin Profile Error: Profile detail not found.")

    # 1. SECURITY CHECK
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    if not (request.user.is_superuser or is_admin):
        return HttpResponse("Access Denied.")

    # ---------------------------------------------------------
    # 💾 2. POST REQUEST (Role Update & Cleanup)
    # ---------------------------------------------------------
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        group_id = request.POST.get("role") or request.POST.get("role_id") 
        
        user_to_update = get_object_or_404(User.objects.using('default'), id=user_id)
        target_profile_main = get_object_or_404(Profile.objects.using('default'), user_id=user_id)

        # 🛡️ Branch Security Check
        if target_profile_main.branch_id != admin_branch_id:
            messages.error(request, "⛔ Access Denied: You cannot change the role of a user from another branch.")
            return redirect("admin_manage_roles")

        if group_id:
            try:
                new_group = Group.objects.using('default').get(id=group_id)
                
                # 🔥 FIX 1: Clear old groups and add the new one
                user_to_update.groups.clear()  
                user_to_update.groups.add(new_group)
                
                # Extract clean role name for the success message
                clean_role_name = new_group.name.split('_')[-1].upper()

                # 🔥 FIX 2: Free up any subordinates
                Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader_id=None)
                Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader_id=None)
                Profile.objects.using('default').filter(manager_id=user_id).update(manager_id=None)
                Profile.objects.using(db_name).filter(manager_id=user_id).update(manager_id=None)

                # 🔥 FIX 3: REMOVED .role assignment since it's a read-only @property!
                # Update Main DB (Only save standard database fields)
                target_profile_main.manager = None
                target_profile_main.team_leader = None
                target_profile_main.is_approved = True
                target_profile_main.save(using='default')

                # Update Tenant DB (Company DB)
                try:
                    target_profile_tenant = Profile.objects.using(db_name).get(user_id=user_id)
                    target_profile_tenant.manager = None
                    target_profile_tenant.team_leader = None
                    target_profile_tenant.is_approved = True
                    target_profile_tenant.save(using=db_name)
                except Profile.DoesNotExist:
                    pass

                messages.success(request, f"✅ Role successfully updated to '{clean_role_name}' for {user_to_update.username}")
                
            except Group.DoesNotExist:
                messages.error(request, "Selected role not found.")
        else:
            messages.error(request, "Please select a role from the dropdown.")
        
        return redirect("admin_manage_roles")
    
    # ---------------------------------------------------------
    # 📋 3. DATA FETCHING (Branch Specific Logic)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    owner_group_name = f"{slug_lower}_owner"
    
    # Remove owners from the list
    owner_ids = list(User.objects.using('default').filter(groups__name__iexact=owner_group_name).values_list('id', flat=True))

    # 🔥 Filter by branch ID safely
    users_qs = Profile.objects.using('default').filter(
        company=company,
        branch_id=admin_branch_id
    ).exclude(user_id=request.user.id).exclude(user_id__in=owner_ids)

    users_list = list(users_qs.select_related('user', 'branch').order_by('user__username'))

    # Role Mapping Logic (Display Purpose)
    staff_user_ids = [u.user_id for u in users_list]
    main_users = User.objects.using('default').filter(id__in=staff_user_ids).prefetch_related('groups')
    
    role_map = {}
    for u in main_users:
        u_role = "Employee"
        for g in u.groups.all():
            g_name = g.name.lower()
            if slug_lower in g_name and 'owner' not in g_name:
                parts = g.name.split('_')
                u_role = parts[-1].title() if len(parts) > 1 else g.name.title()
                break
        role_map[u.id] = u_role

    for u in users_list:
        u.display_role = role_map.get(u.user_id, "Employee")

    # Roles for dropdown
    roles = Group.objects.using('default').filter(name__startswith=f"{company.slug}_").exclude(name__icontains='owner')

    ThreadLocal.DB_NAME = 'default' # Safety reset

    return render(request, "attendance/admin/admin_manage_roles.html", {
        "users": users_list, 
        "roles": roles,
        "current_branch": admin_branch,
        "company": company
    })

from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import User
from accounts.models import Profile
from .middleware import ThreadLocal

@login_required
def admin_employee_list(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન DB માંથી એડમિનની વિગતો લાવો
        profile_admin = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile_admin.company
        admin_branch = profile_admin.branch 
        db_name = f"{company.slug}_db"
        
        # ThreadLocal સેટ કરો
        ThreadLocal.DB_NAME = db_name
    except Profile.DoesNotExist:
        return HttpResponse("Admin Profile Error")

    # 1. PERMISSION CHECK (Main DB check)
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    is_hr = request.user.groups.filter(name__icontains='hr').exists()
    
    if not (request.user.is_superuser or is_admin or is_hr):
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 🔥 2. FETCH ELIGIBLE EMPLOYEE IDs (Avoid Cross-DB JOIN)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    employee_group_name = f"{slug_lower}_employee"
    
    # એવા એમ્પ્લોઈઝ શોધો જે:
    # A) આ જ કંપનીના હોય
    # B) જે 'employee' ગ્રુપમાં હોય
    # C) જો એડમિન બ્રાન્ચ લેવલ પર હોય, તો તેની બ્રાન્ચના જ
    
    base_query = Profile.objects.using('default').filter(
        company=company,
        user__groups__name__iexact=employee_group_name
    )

    if admin_branch:
        base_query = base_query.filter(branch=admin_branch)

    # IDs ની લિસ્ટ બનાવો
    eligible_user_ids = list(base_query.values_list('user_id', flat=True))

    # ---------------------------------------------------------
    # 📥 3. FETCH FULL PROFILES FROM TENANT DB
    # ---------------------------------------------------------
    # હવે ટેનન્ટ ડેટાબેઝમાંથી ફાઇનલ લિસ્ટ લાવો (ID-based હોવાથી એરર નહીં આવે)
    users = []
    if eligible_user_ids:
        users = Profile.objects.using(db_name).filter(
            user_id__in=eligible_user_ids
        ).select_related('user', 'shift', 'branch').order_by('user__username')

    return render(request, "attendance/admin/admin_employee_list.html", {
        "users": users,
        "company": company,
        "current_branch": admin_branch
    })
    
# @login_required
# def admin_view_attendance(request, user_id):

#     if not user_has_permission(request.user , "can_view_reports"):
#         return HttpResponse("Access Denied")

#     admin_company = request.user.profile.company
#     employee = Profile.objects.get(user__id=user_id)

#     # security: employee must belong to same company
#     if employee.company != admin_company:
#         return HttpResponse("Not Allowed")

#     from attendance.models import Attendance, AttendancePolicy

#     attendance = Attendance.objects.filter(user__id=user_id).order_by("-date")

#     policy = AttendancePolicy.objects.filter(company=admin_company).first()

#     # calculate status dynamically
#     for a in attendance:

#         if a.total_hours is None:
#             a.result = "No Record"
#             continue

#         if policy:

#             if a.total_hours >= policy.work_hours_required:
#                 a.result = "Present"

#             elif a.total_hours >= (policy.work_hours_required / 2):
#                 a.result = "Half Day"

#             else:
#                 a.result = "Absent"

#             if a.late_minutes and a.late_minutes > policy.late_after_minutes:
#                 a.result = "Late"

#         else:
#             a.result = a.status

#     return render(request, "attendance/admin_view_attendance.html", {
#         "employee": employee,
#         "attendance": attendance,
#         "policy": policy
#     })
from .middleware import ThreadLocal # Import check kari lejo

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import User
from accounts.models import Profile

@login_required
def admin_user_list_by_role(request, role):
    # ---------------------------------------------------------
    # 0. PROFILE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        admin_branch = profile_admin.branch # આ બ્રાન્ચ ઓબ્જેક્ટ અથવા None હોઈ શકે
    except Exception:
        return HttpResponse("Admin Profile Error")

    # Template માં Attendance ના ડેટા માટે DB context સેટ કરવું
    if company:
        from attendance.middleware import ThreadLocal
        ThreadLocal.DB_NAME = f"{company.slug}_db"

    role_clean = role.lower().strip()
    slug_lower = company.slug.lower()

    # ---------------------------------------------------------
    # 📋 1. FETCH BASE PROFILES (Branch Specific Filtering)
    # ---------------------------------------------------------
    # 🔥 તમારી મેઈન કન્ડિશન: લોગિન એડમિનની બ્રાન્ચ મુજબ જ પ્રોફાઈલ્સ ફિલ્ટર થશે
    # જો admin_branch = None હશે, તો Django 'branch IS NULL' (Head Office) ફિલ્ટર કરશે.
    base_profiles = Profile.objects.using('default').filter(
        company=company,
        branch=admin_branch
    )

    # પાયથન લેવલ પર રોલ ફિલ્ટર કરવા માટે યુઝર્સ લાવો
    base_user_ids = list(base_profiles.values_list('user_id', flat=True))
    
    company_users = User.objects.using('default').filter(
        id__in=base_user_ids
    ).prefetch_related('groups')

    filtered_user_ids = []

    # 🛡️ 2. PYTHON-LEVEL ROLE FILTERING
    for u in company_users:
        user_groups = [g.name.lower() for g in u.groups.all()]
        
        if role_clean == 'employee':
            # જો કોઈ મેનેજમેન્ટ ગ્રુપ ના હોય, તો તે એમ્પ્લોઈ છે
            has_management = any(
                (f"{slug_lower}_admin" in g) or 
                (f"{slug_lower}_hr" in g) or 
                (f"{slug_lower}_manager" in g) or 
                (f"{slug_lower}_team leader" in g) or 
                (f"{slug_lower}_owner" in g) 
                for g in user_groups
            )
            if not has_management:
                filtered_user_ids.append(u.id)
        else:
            # જો પર્ટીક્યુલર રોલ (HR, Manager, etc.) પર ક્લિક કર્યું હોય
            target_group = f"{slug_lower}_{role_clean}"
            if any(target_group in g for g in user_groups): 
                filtered_user_ids.append(u.id)

    # ---------------------------------------------------------
    # 📋 3. FETCH FINAL USERS FOR TEMPLATE
    # ---------------------------------------------------------
    users = Profile.objects.using('default').filter(
        user_id__in=filtered_user_ids
    ).select_related('user', 'branch').order_by('user__username')

    # display_role સેટ કરો
    for u in users:
        u.display_role = role.title()

    # ---------------------------------------------------------
    # 👥 4. TEAM LEADER SQUAD LOGIC (Branch Consistent)
    # ---------------------------------------------------------
    if role_clean == "team leader":
        for u in users:
            # સ્ક્વોડ લિસ્ટમાં પણ એ જ બ્રાન્ચ લોજિક લાગશે જે એડમિનનું છે
            squad_query = Profile.objects.using('default').filter(
                team_leader_id=u.user_id,
                company=company,
                branch=admin_branch # 👈 સ્ક્વોડ પણ બ્રાન્ચ વાઈઝ જ હોવી જોઈએ
            )
                
            squad_members = squad_query.select_related('user')
            u.squad_count = squad_members.count()
            u.squad_members = squad_members

    return render(request, "attendance/admin/admin_user_list_by_role.html", {
        "users": users,
        "role": role.title(), 
        "role_name": role.upper(),
        "company": company,
        "current_branch": admin_branch
    })

@login_required
def superadmin_user_company_list(request, type):
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")

    heading = ""
    records = None
    mode = "" 

    # ---------------------------------------------------------
    # GLOBAL VIEW LOGIC
    # ---------------------------------------------------------
    # Superadmin hamesha 'default' DB mathi badho master data jove che
    if type == "companies":
        heading = "All Companies"
        mode = "company"
        records = Company.objects.all()

    elif type in ["admins", "hr", "managers", "employees"]:
        mode = "user"
        heading = f"{type.capitalize()} Users"
        
        # User roles filters (Hamesha main DB mathi profiles filter thase)
        role_map = {
            "admins": "ADMIN",
            "hr": "HR",
            "managers": "MANAGER",
            "employees": "EMPLOYEE"
        }
        records = Profile.objects.filter(user__groups__name__icontains=role_map[type]).select_related('user', 'company')

    else:
        return HttpResponse("Invalid Type")

    return render(request, "attendance/superadmin/superadmin_dynamic_list.html", {
        "heading": heading,
        "records": records,
        "mode": mode
    })# attendance/views.py

from .middleware import ThreadLocal # ThreadLocal ઇમ્પોર્ટ કરી લેજો

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import Group, User
from accounts.models import Profile
from attendance.models import RolePermission

@login_required
def hr_managers_list(request):
    # ---------------------------------------------------------
    # 0. DATABASE & PROFILE SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા 'default' DB માંથી HR ની પ્રોફાઈલ લાવો
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        hr_branch = profile.branch  # આ બ્રાન્ચ ઓબ્જેક્ટ અથવા None હોઈ શકે
        db_name = f"{company.slug}_db" if company else 'default'
    except Profile.DoesNotExist:
        return HttpResponse("Profile not found")

    # 1. SECURITY CHECK
    if not (request.user.groups.filter(name__icontains='hr').exists() or request.user.is_superuser):
        return HttpResponse("Access Denied")

    # 🔥 2. THE PROXY HACK (HTML બટનો અને પરમિશન માટે)
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except Exception: pass

    class ProxyGroup:
        def __init__(self, obj): self._obj = obj; self.role_permissions = tenant_perms
        def __getattr__(self, name): return getattr(self._obj, name)
    class ProxyManager:
        def __init__(self, m): self._m = m
        def first(self): obj = self._m.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._m, name)
    class ProxyUser:
        def __init__(self, u): self._u = u; self.groups = ProxyManager(u.groups)
        def __getattr__(self, name): return getattr(self._u, name)
    class ProxyRequest:
        def __init__(self, r): self._req = r; self.user = ProxyUser(r.user)
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 3. FETCH MANAGERS (Branch Specific Filtering)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    target_manager_group = f"{slug_lower}_manager"

    # 🔥 તમારી કન્ડિશન: HR ની બ્રાન્ચ મુજબ જ મેનેજરો ફિલ્ટર થશે
    manager_users_qs = User.objects.using('default').filter(
        profile__company=company,
        profile__branch=hr_branch, # જો HR બ્રાન્ચમાં હશે તો બ્રાન્ચ, નહીંતર None
        groups__name__iexact=target_manager_group
    ).select_related('profile', 'profile__branch')

    managers_data = []
    for mgr_user in manager_users_qs:
        mgr_profile = mgr_user.profile
        
        # મેનેજરની અંડરમાં રહેલી ટીમ (Main DB માંથી જેથી JOIN એરર ના આવે)
        team_members = Profile.objects.using('default').filter(
            manager_id=mgr_user.id,
            company=company
        ).select_related('user', 'branch')

        team_list = []
        for member in team_members:
            # રોલ ચેક (Team Leader છે કે નહીં?)
            is_tl = member.user.groups.filter(name__icontains='team leader').exists()
            team_list.append({
                'user': member.user,
                'profile': member,
                'is_team_leader': is_tl
            })

        managers_data.append({
            'manager_user': mgr_user,
            'profile': mgr_profile,
            'email': mgr_user.email,
            'team_count': team_members.count(),
            'team_list': team_list,
            'id': mgr_user.id
        })

    # TL રોલ અસ્તિત્વમાં છે કે નહીં તે ચેક કરવા
    tl_group_exists = Group.objects.using('default').filter(name__icontains=f"{company.slug}_Team Leader").exists()

    return render(request, "attendance/hr/hr_managers_list.html", {
        "request": ProxyRequest(request), 
        "managers": managers_data,
        "team_leader_role_exists": tl_group_exists,
        "company": company,
        "current_branch": hr_branch
    })
    
@login_required
def manager_view_corrections(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        db_name = f"{profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Access Denied")

    if not request.user.groups.filter(name__icontains='manager').exists():
        return HttpResponse("Access Denied")

    # ✅ કરેક્શન રિક્વેસ્ટ કંપની ડેટાબેઝમાંથી લાવો
    requests = AttendanceCorrection.objects.using(db_name).filter(
        user=request.user
    ).order_by("-created_at")

    return render(request, "attendance/correction/manager_corrections.html", {
        "requests": requests
    })

@login_required
def hr_view_corrections(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        db_name = f"{profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Access Denied")

    if not request.user.groups.filter(name__icontains='hr').exists():
        return HttpResponse("Access Denied")

    # ✅ HR પોતાની કંપનીના મેનેજરોની રિક્વેસ્ટ કંપની ડેટાબેઝમાંથી જોશે
    requests = AttendanceCorrection.objects.using(db_name).filter(
        user__profile__company=profile.company,
        user__profile__role="MANAGER"
    ).order_by("-created_at")

    return render(request, "attendance/correction/hr_corrections.html", {
        "requests": requests
    })
from .middleware import ThreadLocal

@login_required
def approve_correction(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Profile Error")

    correction = get_object_or_404(AttendanceCorrection.objects.using(db_name), id=id)
    
    # Security Checks
    if request.user.profile.company != correction.user.profile.company:
        return HttpResponse("Unauthorized")
        
    if not (user_has_permission(request.user, 'can_approve_attendance') or request.user == correction.user.profile.manager):
         return HttpResponse("Access Denied")
         
    # 1. Status Update (Save in Company DB)
    correction.status = "APPROVED"
    correction.save(using=db_name)
    
    # 2. Actual Attendance Update (In Company DB)
    attendance, created = Attendance.objects.using(db_name).get_or_create(
        user=correction.user,
        date=correction.date
    )
    
    # Backup & Update Logic
    if correction.new_check_in:
        attendance.old_check_in = attendance.check_in
        attendance.check_in = correction.new_check_in
        
    if correction.new_check_out:
        attendance.old_check_out = attendance.check_out
        attendance.check_out = correction.new_check_out
        
    attendance.status = "Present"
    attendance.save(using=db_name)
    
    # Notification & Message
    create_notification(correction.user, f"Correction for {correction.date} Approved.")
    messages.success(request, "Correction Approved & Attendance Updated.")
    
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def reject_correction(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    db_name = f"{request.user.profile.company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    if not user_has_permission(request.user, "can_approve_attendance"):
        return HttpResponse("Access Denied !")
    
    correction = get_object_or_404(AttendanceCorrection.objects.using(db_name), id=id)

    # Role specific security
    req_role = (correction.user.profile.role or "").upper()
    if req_role == "HR":
        if request.user.profile.role != "ADMIN":
            return HttpResponse("Only Admin Can Reject HR Requests")
    else:
        if not request.user.groups.filter(name__icontains='hr').exists():
            return HttpResponse("Only HR Can Reject This")

    correction.status = "REJECTED"
    correction.save(using=db_name)
    
    create_notification(correction.user, "Your attendance correction request has been rejected.")
    return redirect("/attendance/hr/correction-requests/")


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import datetime
from .models import Attendance, AttendanceCorrection
from .middleware import ThreadLocal

@login_required
def manager_request_correction(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        db_name = f"{profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except Exception:
        return HttpResponse("Profile Error")

    if request.method == "POST":
        selected_date_str = request.POST.get("selected_date")
        new_in_raw = request.POST.get("new_check_in")
        new_out_raw = request.POST.get("new_check_out")
        reason = request.POST.get("reason")

        # 🔥 FIX: વેરીએબલ્સને પહેલા None સેટ કરો જેથી 'NameError' ના આવે
        new_in = None
        new_out = None

        try:
            # ૧. તારીખ પાર્સિંગ
            if selected_date_str:
                selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            else:
                messages.error(request, "Please select a date.")
                return redirect(request.path)

            # ૨. ટાઇમ પાર્સિંગ (નામમાં સુધારો: HTML માંથી પકડીને new_in માં સ્ટોર)
            if new_in_raw:
                new_in = datetime.strptime(new_in_raw, "%H:%M").time()
            
            if new_out_raw:
                new_out = datetime.strptime(new_out_raw, "%H:%M").time()

            # ૩. જૂની એટેન્ડન્સ ફેચ કરો (Tenant DB mathi)
            actual_att = Attendance.objects.using(db_name).filter(
                user=request.user, 
                date=selected_date
            ).first()

            # ---------------------------------------------------------
            # 💾 ૪. SAVE REQUEST (Using tenant DB context)
            # ---------------------------------------------------------
            # હવે 'new_in' અને 'new_out' ૧૦૦% ડિફાઇન્ડ છે.
            AttendanceCorrection.objects.using(db_name).create(
                user=request.user,
                date=selected_date,
                old_check_in=actual_att.check_in if actual_att else None,
                old_check_out=actual_att.check_out if actual_att else None,
                new_check_in=new_in,   # 👈 No more NameError
                new_check_out=new_out, 
                reason=reason,
                status="PENDING"
            )

            messages.success(request, "Correction request submitted successfully!")
            return redirect('/attendance/dashboard/manager/')

        except ValueError:
            messages.error(request, "Invalid Date or Time format. Please use HH:MM.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return render(request, "attendance/correction/request_correction.html")

from .middleware import ThreadLocal
from datetime import datetime

@login_required
def hr_request_correction(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except Profile.DoesNotExist:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile not found")

    # 1. Security Check
    if not request.user.groups.filter(name__icontains='hr').exists():
        return HttpResponse("Access Denied")

    if request.method == "POST":
        selected_date_raw = request.POST.get("selected_date")
        selected_date = datetime.strptime(selected_date_raw, "%Y-%m-%d").date()

        # ✅ Fetch Attendance from Company DB
        attendance = Attendance.objects.using(db_name).filter(
            user=request.user,
            date=selected_date
        ).first()

        reason = request.POST.get("reason")
        new_in_raw = request.POST.get("new_check_in")
        new_out_raw = request.POST.get("new_check_out")

        new_in = new_out = None
        if new_in_raw:
            new_in = datetime.strptime(new_in_raw, "%Y-%m-%dT%H:%M").time()
        if new_out_raw:
            new_out = datetime.strptime(new_out_raw, "%Y-%m-%dT%H:%M").time()

        # ✅ STEP 1: Save Correction Request in Company DB
        AttendanceCorrection.objects.using(db_name).create(
            user=request.user,
            date=selected_date,
            old_check_in=attendance.check_in if attendance else None,
            old_check_out=attendance.check_out if attendance else None,
            new_check_in=new_in,
            new_check_out=new_out,
            reason=reason,
            status="PENDING"
        )

        # ✅ STEP 2: Find ADMIN of same company (From Company DB)
        # Note: Role check hamesha text field and group banne par rakho safety mate
        admin_profile = Profile.objects.using(db_name).filter(
            company=company,
            user__groups__name__icontains="admin"
        ).first()

        # Send notification to Admin
        if admin_profile:
            create_notification(
                admin_profile.user,
                f"HR {request.user.username} requested an attendance correction."
            )

        messages.success(request, "Correction request sent to Admin successfully.")
        return redirect("/attendance/dashboard/hr/")

    return render(request, "attendance/request_correction.html")
# attendance/views.py

@login_required
def export_monthly_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font
    from calendar import monthrange
    from datetime import date, datetime, timedelta
    from django.db.models import Q  # Q object import karvu jaruri che

    # 1. Fetch Request Data
    user_id = request.GET.get("user")
    month = int(request.GET.get("month"))
    year = int(request.GET.get("year"))

    user = get_object_or_404(User, id=user_id)
    profile = user.profile
    company = profile.company

    # 2. Setup Data
    days_in_month = monthrange(year, month)[1]
    today = date.today()
    
    # Policy & Shift (For Logic Calculation)
    policy = AttendancePolicy.objects.filter(company=company).first()
    shift = profile.shift

    # Fetch Attendance Records
    attendance_map = {
        a.date: a
        for a in Attendance.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        )
    }

    # Fetch Holidays (Company Specific)
    holiday_qs = Holiday.objects.filter(Q(company=company) | Q(company__isnull=True))
    holidays = {h.date: h.name for h in holiday_qs}

    # Fetch Approved Leaves
    start_date = date(year, month, 1)
    end_date = date(year, month, days_in_month)
    
    approved_leaves = Leave.objects.filter(
        user=user,
        status="APPROVED",
        from_date__lte=end_date,
        to_date__gte=start_date
    )
    
    # 3. Excel Workbook Setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    # Header Row
    ws.append(["Date", "Check-In", "Check-Out", "Status"])
    for c in ws["1:1"]:
        c.font = Font(bold=True)

    # 4. Loop & Logic (Matching 'monthly_attendance' View)
    working_days = 0
    present_count = 0
    absent_count = 0
    
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        att = attendance_map.get(d)
        
        # Check Leave
        is_on_leave = False
        for leave in approved_leaves:
            if leave.from_date <= d <= leave.to_date:
                is_on_leave = True
                break
        
        # Check Weekend & Holiday
        is_weekend = d.weekday() == 5 or d.weekday() == 6 # Sat or Sun
        is_holiday = d in holidays

        cin = ""
        cout = ""
        status = "-" # Default

        # --- Priority 1: Attendance (હાજરી હોય તો ગણતરી કરો) ---
        if att:
            if att.check_in:
                cin = att.check_in.strftime("%I:%M %p")
            
            if att.check_out:
                cout = att.check_out.strftime("%I:%M %p")

            if att.check_in and not att.check_out:
                status = "Working"
            elif att.check_in and att.check_out:
                # Calculate Status based on policy (Late/Half Day etc.)
                check_in_dt = datetime.combine(d, att.check_in)
                check_out_dt = datetime.combine(d, att.check_out)
                work_hours = (check_out_dt - check_in_dt).total_seconds() / 3600
                
                late_minutes = 0
                if shift:
                    shift_in = datetime.combine(d, shift.start_time)
                    late_minutes = (check_in_dt - shift_in).total_seconds() / 60
                
                if policy:
                    if work_hours < (policy.work_hours_required / 2):
                        status = "Half Day"
                    elif late_minutes > policy.late_after_minutes:
                        status = "Late"
                    elif work_hours >= policy.work_hours_required:
                        status = "Present"
                    else:
                        status = "Partial"
                else:
                    status = "Present"
        
        # --- Priority 2: Holiday ---
        elif is_holiday:
            status = holidays[d] # Holiday Name (e.g. Diwali)
            
        # --- Priority 3: Leave ---
        elif is_on_leave:
            status = "Approved Leave"
            
        # --- Priority 4: Others ---
        else:
            if d > today:
                status = "-"
            elif is_weekend:
                status = "Week Off"
            else:
                status = "Absent"

        # --- Counters Calculation ---
        
        # Present Logic: (Any kind of presence counts)
        if status in ["Present", "Late", "Half Day", "Partial", "Working"]:
            present_count += 1
        
        # Absent Logic:
        if status == "Absent":
            absent_count += 1
            
        # 🔥 Working Days Logic (Updated as per requirement) 🔥
        # Working Day ત્યારે જ ગણાય જો: Weekend ના હોય, Holiday ના હોય અને Leave ના હોય.
        if not is_weekend and not is_holiday and not is_on_leave:
            working_days += 1

        # Append Row to Excel
        ws.append([
            d.strftime("%d-%b-%Y"),
            cin,
            cout,
            status
        ])

    # 5. Summary Sheet (Total Counts)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Monthly Attendance Summary"])
    ws2["A1"].font = Font(size=14, bold=True)
    ws2.append(["Employee", user.username])
    ws2.append(["Month", f"{month}-{year}"])
    ws2.append([])
    
    # Final Counts
    ws2.append(["Total Working Days", working_days])
    ws2.append(["Present Days", present_count])
    ws2.append(["Absent Days", absent_count])

    for c in ws2["A2:A6"]:
        c[0].font = Font(bold=True)

    # 6. Return File
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="attendance_{user.username}_{month}-{year}.xlsx"'
    
    wb.save(response)
    return response
@login_required
def export_yearly_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from calendar import monthrange
    from datetime import date, datetime, timedelta, time
    from django.db.models import Q

    user_id = request.GET.get("user")
    year = int(request.GET.get("year", datetime.now().year))
    user = get_object_or_404(User, id=user_id)
    profile = user.profile
    company = profile.company
    
    # --- ૧. પોલિસી અને શિફ્ટ ડેટા મેળવો ---
    policy = AttendancePolicy.objects.filter(company=company).first()
    shift = profile.shift
    
    # વીક-ઓફ લોજિક
    week_off_indices = [int(x) for x in policy.week_off_days.split(',') if x.isdigit()] if policy and policy.week_off_days else [6]

    wb = openpyxl.Workbook()
    
    # --- ૨. SUMMARY SHEET ---
    ws_sum = wb.active
    ws_sum.title = "Yearly Summary"
    ws_sum.append(["Month", "Total Working Days", "Present", "Half Day", "Late", "Partial", "Absent", "Leave", "Holidays"])
    
    header_fill = PatternFill(start_color="540863", end_color="540863", fill_type="solid")
    for cell in ws_sum[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- ૩. DETAILED DATA SHEET ---
    ws_data = wb.create_sheet(title="Daily Attendance Details")
    ws_data.append(["Month", "Date", "Check-In", "Check-Out", "Status"])
    for cell in ws_data[1]:
        cell.font = Font(bold=True)

    # --- ૪. DATA PROCESSING LOOP ---
    for m in range(1, 13):
        days_in_month = monthrange(year, m)[1]
        m_stats = {'working_days': 0, 'present': 0, 'half_day': 0, 'late': 0, 'partial': 0, 'absent': 0, 'leave': 0, 'holiday': 0}

        attendance_map = {a.date: a for a in Attendance.objects.filter(user=user, date__year=year, date__month=m)}
        holidays = {h.date: h.name for h in Holiday.objects.filter(Q(company=company) | Q(company__isnull=True), date__year=year, date__month=m)}
        
        # Leaves logic
        leaves = Leave.objects.filter(user=user, status='APPROVED', from_date__lte=date(year, m, days_in_month), to_date__gte=date(year, m, 1))
        leave_dates = set()
        for l in leaves:
            curr = l.from_date
            while curr <= l.to_date:
                if curr.month == m: leave_dates.add(curr)
                curr += timedelta(days=1)

        for day in range(1, days_in_month + 1):
            d = date(year, m, day)
            att = attendance_map.get(d)
            is_holiday = d in holidays
            is_leave = d in leave_dates
            is_week_off = d.weekday() in week_off_indices
            
            status = "-"
            cin_str = att.check_in.strftime("%I:%M %p") if att and att.check_in else ""
            cout_str = att.check_out.strftime("%I:%M %p") if att and att.check_out else ""

            # --- 🔥 CORE STATUS LOGIC 🔥 ---
            if is_leave:
                status = "Approved Leave"
                m_stats['leave'] += 1
            elif is_holiday:
                status = holidays[d]
                m_stats['holiday'] += 1
            elif att:
                if att.check_in and att.check_out:
                    # કલાકો અને મોડા સમયની ગણતરી
                    in_dt = datetime.combine(d, att.check_in)
                    out_dt = datetime.combine(d, att.check_out)
                    work_hours = (out_dt - in_dt).total_seconds() / 3600
                    
                    late_mins = 0
                    if shift and shift.start_time:
                        shift_in = datetime.combine(d, shift.start_time)
                        if in_dt > shift_in:
                            late_mins = (in_dt - shift_in).total_seconds() / 60

                    # પોલિસી મુજબ સ્ટેટસ નક્કી કરો
                    if policy:
                        if work_hours < (policy.work_hours_required / 2):
                            status = "Half Day"
                            m_stats['half_day'] += 1
                        elif late_mins > policy.late_after_minutes:
                            status = "Late"
                            m_stats['late'] += 1
                        elif work_hours >= policy.work_hours_required:
                            status = "Present"
                            m_stats['present'] += 1
                        else:
                            status = "Partial"
                            m_stats['partial'] += 1
                    else:
                        status = "Present"
                        m_stats['present'] += 1
                elif att.check_in:
                    status = "Working"
                    m_stats['present'] += 1 # Working ને હાજરમાં ગણવું
                else:
                    status = "Absent"
                    m_stats['absent'] += 1
            elif is_week_off:
                status = "Week Off"
            elif d < date.today():
                status = "Absent"
                m_stats['absent'] += 1

            if not is_week_off and not is_holiday and not is_leave:
                m_stats['working_days'] += 1

            ws_data.append([d.strftime("%B"), d.strftime("%d-%b-%Y"), cin_str, cout_str, status])

        # Summary Row Append
        ws_sum.append([
            date(year, m, 1).strftime("%B"), 
            m_stats['working_days'], m_stats['present'], 
            m_stats['half_day'], m_stats['late'], 
            m_stats['partial'], m_stats['absent'], 
            m_stats['leave'], m_stats['holiday']
        ])

    # Final Auto-Formatting
    for ws in [ws_sum, ws_data]:
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 3

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Yearly_Report_{user.username}_{year}.xlsx"'
    wb.save(response)
    return response
# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal # ThreadLocal import check kari lejo

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from .models import Leave, RolePermission
from accounts.models import Profile

@login_required
def admin_manager_leave_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        # હંમેશા મેઈન ડીબી માંથી પ્રોફાઈલ લાવો
        hr_profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = hr_profile.company
        hr_branch = hr_profile.branch # આ બ્રાન્ચ ઓબ્જેક્ટ અથવા None હોઈ શકે
        db_name = f"{company.slug}_db"
    except Exception:
        return HttpResponse("Error: User Profile not found.")

    user_groups = [g.name.lower() for g in request.user.groups.all()]
    slug_lower = company.slug.lower()

    # 🔥 1. THE PROXY HACK (HTML બટનો માટે)
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except: pass

    class ProxyGroup:
        def __init__(self, obj): self._obj = obj; self.role_permissions = tenant_perms
        def __getattr__(self, name): return getattr(self._obj, name)
    class ProxyManager:
        def __init__(self, m): self._m = m
        def first(self): obj = self._m.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._m, name)
    class ProxyUser:
        def __init__(self, u): self._user = u; self.groups = ProxyManager(u.groups)
        def __getattr__(self, name): return getattr(self._user, name)
    class ProxyRequest:
        def __init__(self, r): self._req = r; self.user = ProxyUser(r.user)
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 2. GET TARGET USER IDs (Branch Specific Filter)
    # ---------------------------------------------------------
    # 🔥 તમારી મેઈન કન્ડિશન: લોગિન કરનારની બ્રાન્ચ મુજબ જ સ્ટાફ ફિલ્ટર થશે
    base_user_qs = User.objects.using('default').filter(
        profile__company=company,
        profile__branch=hr_branch # જો hr_branch=None હશે તો Head Office ના જ લોકો આવશે
    ).exclude(id=request.user.id)

    # રોલ મુજબ ફિલ્ટરિંગ (Main DB Groups logic)
    target_user_ids = []
    
    if request.user.is_superuser or any('admin' in g for g in user_groups):
        # Admin પોતાની બ્રાન્ચના બધાને જોઈ શકે
        target_user_ids = list(base_user_qs.values_list('id', flat=True))
        
    elif any('hr' in g for g in user_groups):
        # HR પોતાની બ્રાન્ચના એડમિન સિવાયના બધાને જોઈ શકે
        target_user_ids = list(base_user_qs.exclude(
            groups__name__iexact=f"{slug_lower}_admin"
        ).values_list('id', flat=True))
        
    else: # Manager
        # મેનેજર માત્ર પોતાની બ્રાન્ચમાં પોતાની અંડર આવતા લોકોને જોઈ શકે
        target_user_ids = list(base_user_qs.filter(
            profile__manager_id=request.user.id
        ).values_list('id', flat=True))

    # ---------------------------------------------------------
    # 📥 3. FETCH LEAVES FROM TENANT DB
    # --- FILTERS LOGIC ---
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    # 📥 FETCH LEAVES FROM TENANT DB
    leave_requests_list = Leave.objects.using(db_name).filter(
        user_id__in=target_user_ids
    ).order_by('-id').select_related('user')

    # ૧. નામથી ફિલ્ટર
    if search_query:
        leave_requests_list = leave_requests_list.filter(
            user__first_name__icontains=search_query
        ) | leave_requests_list.filter(
            user__last_name__icontains=search_query
        )

    # ૨. સ્ટેટસથી ફિલ્ટર
    if status_filter:
        leave_requests_list = leave_requests_list.filter(status=status_filter)

    total_count = leave_requests_list.count()

    # ---------------------------------------------------------
    # 📄 4. PAGINATION LOGIC (10 items per page)
    # ---------------------------------------------------------
    paginator = Paginator(leave_requests_list, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "attendance/leave/manage_leave_request.html", {
        "request": ProxyRequest(request), 
        "requests": page_obj,                   
        "total_requests": total_count,    
        "current_branch": hr_branch # HTML માં બતાવવા માટે
    })

@login_required
def admin_approve_leave(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    db_name = f"{request.user.profile.company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    # 1. Security Check
    user_groups = [g.name.lower() for g in request.user.groups.all()]
    if not (request.user.is_superuser or any(x in g for g in user_groups for x in ['admin', 'hr', 'manager'])):
        return HttpResponse("Access Denied.")

    # 2. Main Logic (Sacha DB mathi fetch ane update karo)
    leave = get_object_or_404(Leave.objects.using(db_name), id=id)
    leave.status = "APPROVED"
    leave.save(using=db_name) # Force save in Company DB
    
    # Notification automatically company DB ma jase ThreadLocal thi
    create_notification(leave.user, f"Your leave request for {leave.from_date} has been APPROVED.")

    messages.success(request, "Leave Approved Successfully!")
    return redirect(request.META.get("HTTP_REFERER", '/'))


@login_required
def admin_reject_leave(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    db_name = f"{request.user.profile.company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    # 1. Security Check
    user_groups = [g.name.lower() for g in request.user.groups.all()]
    if not (request.user.is_superuser or any(x in g for g in user_groups for x in ['admin', 'hr', 'manager'])):
        return HttpResponse("Access Denied.")

    # 2. Main Logic (Sacha DB mathi update)
    leave = get_object_or_404(Leave.objects.using(db_name), id=id)
    leave.status = "REJECTED"
    leave.save(using=db_name)

    create_notification(leave.user, f"Your leave request for {leave.from_date} has been REJECTED.")

    messages.error(request, "Leave Rejected.")
    return redirect(request.META.get("HTTP_REFERER", '/'))

from .middleware import ThreadLocal # ThreadLocal check kari lejo

@login_required
def manager_my_leaves(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        db_name = f"{profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        ThreadLocal.DB_NAME = 'default'
        return HttpResponse("Profile Error")

    # Permission Check
    if profile.role.upper() != "SUPERADMIN":
        if not user_has_permission(request.user, "can_self_access"):
            return HttpResponse("Access Denied")

    # ✅ Fetch ONLY manager’s own leave requests from COMPANY DB
    leaves = Leave.objects.using(db_name).filter(
        user=request.user
    ).order_by("-applied_at")

    return render(request, "attendance/leave/manager_my_leaves.html", {
        "leaves": leaves
    })

@login_required
def my_correction_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        db_name = 'default'

    # 1. Fetch from Company DB
    corrections = AttendanceCorrection.objects.using(db_name).filter(
        user=request.user
    ).order_by('-date')

    # ---------------------------------------------------------
    # 🔥 OLD TIME LOGIC (Using Company DB context)
    # ---------------------------------------------------------
    for c in corrections:
        actual_att = Attendance.objects.using(db_name).filter(user=request.user, date=c.date).first()
        if actual_att:
            c.old_check_in = actual_att.check_in
            c.old_check_out = actual_att.check_out
        else:
            c.old_check_in = None
            c.old_check_out = None

    return render(request, "attendance/correction/my_correction_requests.html", {
        "corrections": corrections
    })


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from .models import AttendanceCorrection, Branch, Attendance
from accounts.models import Profile
from .middleware import ThreadLocal

@login_required
def admin_correction_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        admin_branch = profile.branch 
        admin_branch_id = profile.branch_id 
        
        if not company:
            return HttpResponse("તમારી પ્રોફાઇલમાં કંપની અસાઇન કરેલી નથી.")
            
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except Profile.DoesNotExist:
        return HttpResponse("Profile Error: Profile not found.")

    # 1. SECURITY CHECK
    is_admin = request.user.groups.filter(name__icontains='admin').exists()
    if not (is_admin or request.user.is_superuser):
        return HttpResponse("Access Denied: Admin privileges required.")

    # ---------------------------------------------------------
    # 🔍 1. FILTERS & SEARCH (Getting Query Params)
    # ---------------------------------------------------------
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    # ---------------------------------------------------------
    # 📋 2. FETCH ELIGIBLE HR USERS (Branch vs Global Logic)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    hr_group_name = f"{slug_lower}_hr"
    
    # પાયાની ક્વેરી: આ કંપનીના HR શોધો
    hr_profiles_qs = Profile.objects.using('default').filter(
        company=company,
        user__groups__name__iexact=hr_group_name
    )

    # ✅ તમારી શરત: બ્રાન્ચ મુજબ ફિલ્ટર
    if admin_branch_id:
        hr_profiles_qs = hr_profiles_qs.filter(branch_id=admin_branch_id)
    else:
        hr_profiles_qs = hr_profiles_qs.filter(branch__isnull=True)

    hr_user_ids = list(hr_profiles_qs.values_list('user_id', flat=True))

    # ---------------------------------------------------------
    # 📥 3. QUERY CORRECTIONS WITH FILTERS
    # ---------------------------------------------------------
    if not hr_user_ids:
        corrections_list = AttendanceCorrection.objects.none()
    else:
        # બેઝ ક્વેરી
        corrections_list = AttendanceCorrection.objects.using(db_name).filter(
            user_id__in=hr_user_ids
        ).order_by("-id").select_related('user')

        # 🔥 નામ કે યુઝરનેમ થી ફિલ્ટર
        if search_query:
            corrections_list = corrections_list.filter(
                user__first_name__icontains=search_query
            ) | corrections_list.filter(
                user__last_name__icontains=search_query
            ) | corrections_list.filter(
                user__username__icontains=search_query
            )
        
        # 🔥 સ્ટેટસ થી ફિલ્ટર
        if status_filter:
            corrections_list = corrections_list.filter(status=status_filter)

    # ---------------------------------------------------------
    # 🕒 4. FETCH OLD ATTENDANCE DATA (For Comparison)
    # ---------------------------------------------------------
    for corr in corrections_list:
        old_record = Attendance.objects.using(db_name).filter(
            user_id=corr.user_id, 
            date__year=corr.date.year,
            date__month=corr.date.month,
            date__day=corr.date.day
        ).first()
        corr.old_attendance = old_record 

    # ---------------------------------------------------------
    # 📄 5. PAGINATION LOGIC
    # ---------------------------------------------------------
    total_count = corrections_list.count()
    paginator = Paginator(corrections_list, 10) 
    page_obj = paginator.get_page(request.GET.get('page'))

    # 🏢 6. RENDER
    # ThreadLocal Reset for safety
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/correction/hr_correction_requests.html", {
        "corrections": page_obj, 
        "total_requests": total_count,
        "role": "ADMIN",
        "current_branch": admin_branch,
        "company": company
    })
@login_required
def hr_correction_requests(request):
    # ---------------------------------------------------------
    # 0. DATABASE & CONTEXT SETUP
    # ---------------------------------------------------------
    try:
        profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = profile.company
        hr_branch = profile.branch 
        db_name = f"{company.slug}_db"
    except Exception:
        return HttpResponse("Profile not found")

    user_groups = [g.name.lower() for g in request.user.groups.all()]
    slug_lower = company.slug.lower()
    
    # Proxy Hack (Buttons માટે)
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            from .models import RolePermission
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except: pass

    class ProxyRequest:
        def __init__(self, r): self.user = r.user; self._req = r
        def __getattr__(self, name): return getattr(self._req, name)

    # ---------------------------------------------------------
    # 📋 1. BRANCH FILTERING
    # ---------------------------------------------------------
    base_users = User.objects.using('default').filter(
        profile__company=company,
        profile__branch=hr_branch
    )

    if any('hr' in g for g in user_groups):
        allowed_user_ids = list(base_users.exclude(
            groups__name__iregex=fr'({slug_lower}_admin|{slug_lower}_hr|{slug_lower}_owner|admin|hr|owner)'
        ).values_list('id', flat=True))
    elif any('admin' in g for g in user_groups) or request.user.is_superuser:
        allowed_user_ids = list(base_users.filter(groups__name__iexact=f"{slug_lower}_hr").values_list('id', flat=True))
    else:
        return HttpResponse("Access Denied")

    # ---------------------------------------------------------
    # 📋 2. FETCH DATA (Error Fixed Here)
    # ---------------------------------------------------------
    # 🔥 અહીંથી 'attendance' કાઢી નાખ્યું છે જેથી FieldError ના આવે
    corrections_list = AttendanceCorrection.objects.using(db_name).filter(
        user_id__in=allowed_user_ids
    ).order_by("-id").select_related('user')
    
    # ---------------------------------------------------------
    # 📋 2. FETCH DATA (Stronger Manual Lookup)
    # ---------------------------------------------------------
    from .models import Attendance
    corrections_list = AttendanceCorrection.objects.using(db_name).filter(
        user_id__in=allowed_user_ids
    ).order_by("-id").select_related('user')
    
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    corrections_list = AttendanceCorrection.objects.using(db_name).filter(
        user_id__in=allowed_user_ids
    ).order_by("-id").select_related('user')

    # 🔥 જો સર્ચમાં નામ લખ્યું હોય તો ફિલ્ટર કરો
    if search_query:
        corrections_list = corrections_list.filter(
            user__first_name__icontains=search_query
        ) | corrections_list.filter(
            user__last_name__icontains=search_query
        ) | corrections_list.filter(
            user__username__icontains=search_query
        )
    if status_filter:
        corrections_list = corrections_list.filter(status=status_filter)
    # --- OLD TIME FETCHING LOGIC ---
    from .models import Attendance
    for corr in corrections_list:
        # 🎯 પ્રોબ્લેમ અહીં હતો: તારીખને સ્ટ્રિંગમાં કન્વર્ટ કરીને મેચ કરો અથવા __date વાપરો
        # આપણે અહીં __date ફિલ્ટર વાપરીશું જે DateTime માંથી માત્ર Date ને જ ચેક કરશે
        old_record = Attendance.objects.using(db_name).filter(
            user_id=corr.user_id, 
            date__year=corr.date.year,
            date__month=corr.date.month,
            date__day=corr.date.day
        ).first()

        # જો હજુ પણ ના મળે, તો એકવાર સીધું ફિલ્ટર ટ્રાય કરો
        if not old_record:
            old_record = Attendance.objects.using(db_name).filter(
                user_id=corr.user_id, 
                date=corr.date
            ).first()
        
        # આ ઓબ્જેક્ટને રિક્વેસ્ટ સાથે એટેચ કરો
        corr.old_attendance = old_record
        # 🔍 Debugging માટે (જો હજુ ખાલી આવે તો console માં દેખાશે)
        # print(f"User: {corr.user_id}, Date: {corr.date}, Old Record: {old_record}")

    # ---------------------------------------------------------
    # 📄 3. PAGINATION
    # ---------------------------------------------------------
    paginator = Paginator(corrections_list, 10) 
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, "attendance/correction/hr_correction_requests.html", {
        "request": ProxyRequest(request), 
        "corrections": page_obj,          
        "total_requests": corrections_list.count(), 
        "current_branch": hr_branch
    })

def create_notification(user, message):
    Notification.objects.create(
        user = user,
        message = message
    )
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Notification
from accounts.models import Profile
from .middleware import ThreadLocal

@login_required
def notifications_page(request):
    """બધા નોટિફિકેશન્સ બતાવવા અને તેને Read માર્ક કરવા માટેનું વ્યુ"""
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        # પ્રોફાઇલ હંમેશા 'default' DB માંથી જ લાવો
        profile = Profile.objects.using('default').get(user=request.user)
        company = profile.company
        
        if company:
            db_name = f"{company.slug}_db"
            ThreadLocal.DB_NAME = db_name
        else:
            db_name = 'default'
            ThreadLocal.DB_NAME = 'default'
            
    except Exception as e:
        print(f"DEBUG: Profile fetching error: {e}")
        db_name = 'default'
        ThreadLocal.DB_NAME = 'default'

    # ---------------------------------------------------------
    # 📥 1. FETCH NOTIFICATIONS (Force into Memory)
    # ---------------------------------------------------------
    # 🔥 FIX: Wrap the query in list() to force database evaluation immediately!
    notes = list(Notification.objects.using(db_name).filter(
        user_id=request.user.id
    ).order_by('-created_at'))

    print(f"DEBUG: Fetching from {db_name}, Found: {len(notes)} notes for User ID {request.user.id}")

    # ---------------------------------------------------------
    # ✅ 2. MARK AS READ (Bulk Update)
    # ---------------------------------------------------------
    # Now that we have safely stored the notes in a python list, we can update the DB
    if len(notes) > 0:
        Notification.objects.using(db_name).filter(
            user_id=request.user.id,
            is_read=False
        ).update(is_read=True)

    # ---------------------------------------------------------
    # 3. SAFETY RESET & RENDER
    # ---------------------------------------------------------
    # Reset ThreadLocal so we don't leak database connections to other users
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/MISC/notifications.html", {
        "notes": notes, 
        "db_info": db_name
    })    
    from .utils import validate_leave_dates   # <-- SAME validation as apply leave
from .views import create_notification    # or import correctly if in utils


# @login_required
# def update_leave(request, id):

#     # 🔒 Only owner can update
#     leave = Leave.objects.get(id=id, user=request.user)

#     if request.method == "POST":

#         # ----------------------------------
#         # STEP-1 : parse new dates
#         # ----------------------------------
#         new_from = date.fromisoformat(request.POST.get("from_date"))
#         new_to   = date.fromisoformat(request.POST.get("to_date"))
#         reason   = request.POST.get("reason")

#         # ----------------------------------
#         # STEP-2 : VALIDATION (reuse logic)
#         # ----------------------------------
#         error = validate_leave_dates(request.user, new_from, new_to)

#         if error:
#             messages.error(request, error)
#             return redirect("update_leave", id=leave.id)

#         # ----------------------------------
#         # STEP-3 : save OLD dates (history)
#         # ----------------------------------
#         leave.old_from_date = leave.from_date
#         leave.old_to_date   = leave.to_date

#         # ----------------------------------
#         # STEP-4 : apply NEW values
#         # ----------------------------------
#         leave.from_date = new_from
#         leave.to_date   = new_to
#         leave.reason    = reason

#         # ----------------------------------
#         # STEP-5 : reset approval state
#         # ----------------------------------
#         leave.status     = "PENDING"
#         leave.is_updated = True
#         leave.updated_at = timezone.now()

#         leave.save()

#         # ----------------------------------
#         # STEP-6 : NOTIFICATIONS
#         # ----------------------------------
#         profile = request.user.profile

#         # EMPLOYEE → MANAGER
#         if profile.role.upper() == "EMPLOYEE":
#             if profile.manager:
#                 create_notification(
#                     profile.manager,
#                     f"{request.user.username} updated leave request"
#                 )

#         # MANAGER / HR → ADMIN
#         else:
#             admin = Profile.objects.filter(
#                 company=profile.company,
#                 role="ADMIN"
#             ).first()

#             if admin:
#                 create_notification(
#                     admin.user,
#                     f"{request.user.username} updated leave request"
#                 )

#         # ----------------------------------
#         # STEP-7 : success message
#         # ----------------------------------
#         messages.success(
#             request,
#             "Leave updated successfully and sent for approval"
#         )

#         return redirect("my_leave_requests")

#     # ----------------------------------
#     # GET REQUEST
#     # ----------------------------------
#     return render(request, "attendance/leave/update_leave.html", {
#         "leave": leave,
#         "today": date.today()
#     })


# views.py માં 'update_leave' ફંક્શન શોધો અને આ મુજબ અપડેટ કરો
from .middleware import ThreadLocal
from datetime import datetime

@login_required
def update_leave(request, id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        db_name = 'default'

    # ૧. માલિકી ચેક કરો (Sacha DB mathi fetch karo)
    leave = get_object_or_404(Leave.objects.using(db_name), id=id, user=request.user)

    if request.method == "POST":
        new_from = request.POST.get("from_date")
        new_to = request.POST.get("to_date")
        reason = request.POST.get("reason")

        leave.from_date = datetime.strptime(new_from, "%Y-%m-%d").date()
        leave.to_date = datetime.strptime(new_to, "%Y-%m-%d").date()
        leave.reason = reason
        leave.status = "PENDING" 
        
        # ✅ Save in Company DB
        leave.save(using=db_name)

        # ૨. Group Check (Main DB operation)
        is_employee = request.user.groups.filter(name__icontains='employee').exists()
        
        try:
            profile = request.user.profile # standard related_name vapro
            
            # ૩. નોટિફિકેશન લોજિક (create_notification handles routing)
            if is_employee:
                if profile.manager:
                    create_notification(profile.manager, f"{request.user.username} updated leave request")
            else:
                admin = Profile.objects.using(db_name).filter(user__groups__name__icontains='admin').first()
                if admin:
                    create_notification(admin.user, f"{request.user.username} updated leave request")
        except Exception as e:
            print(f"Notification Error: {e}")

        messages.success(request, "Leave updated successfully and sent for approval")
        return redirect("my_leave_requests")

    return render(request, "attendance/leave/update_leave.html", {"leave": leave})


@login_required
def profile_view(request):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        db_name = 'default'

    # Profile fetch from Company DB
    profile = get_object_or_404(Profile.objects.using(db_name), user=request.user)
    
    if request.method == "POST":
        new_username = request.POST.get("username")
        new_phone = request.POST.get("phone")
        
        # A. USERNAME UPDATE (Hamesha DEFAULT/MAIN DB ma thashe)
        if new_username and new_username != request.user.username:
            if User.objects.filter(username=new_username).exists():
                messages.error(request, "Username already taken!")
            else:
                request.user.username = new_username
                request.user.save() # default DB save
                messages.success(request, "Username updated successfully!")

        # B. PHONE UPDATE (COMPANY DB MA SAVE KARO)
        if new_phone:
            if len(new_phone) == 10 and new_phone.isdigit():
                # Check uniqueness in Company DB
                if Profile.objects.using(db_name).filter(phone=new_phone).exclude(user=request.user).exists():
                    messages.error(request, "This phone number is already registered.")
                else:
                    profile.phone = new_phone
                    profile.save(using=db_name) # ✅ Sync to Company DB
                    
                    # Optional: Profile copy jo main DB ma hoy to tya pan update karo
                    # profile.save(using='default')
                    
                    messages.success(request, "Phone number updated successfully!")
            else:
                messages.error(request, "Invalid phone number! Please enter 10 digits.")

        return redirect('profile')
    
    return render(request, 'attendance/MISC/profile.html', {
        'profile': profile,
        'user': request.user
    })
from django.utils import timezone
from .middleware import ThreadLocal
import random
from django.core.mail import send_mail
from django.conf import settings

@login_required
def profile_change_password_otp(request):
    """પાસવર્ડ બદલવા માટે OTP મોકલવાનું ફંક્શન"""
    
    # Generate 6 digit OTP
    otp = str(random.randint(100000, 999999))
    
    # Store in session
    request.session['password_reset_otp'] = otp
    request.session['otp_timestamp'] = str(timezone.now())

    # Send Email
    subject = "OTP for Password Change - AttendancePro"
    message = f"Hello {request.user.username},\n\nYour OTP for changing password is {otp}.\nValid for 5 minutes.\n\nIf you didn't request this, please secure your account."
    email_from = settings.EMAIL_HOST_USER
    recipient_list = [request.user.email]
    
    try:
        send_mail(subject, message, email_from, recipient_list)
        messages.info(request, "OTP has been sent to your registered email.")
    except Exception as e:
        messages.error(request, "Failed to send email. Please check your connection.")
        return redirect('profile')

    return render(request, 'attendance/MISC/profile_verify_otp.html')



@login_required
def profile_verify_password_otp(request):
    """OTP વેરિફિકેશન અને નવો પાસવર્ડ સેટ કરવાનું લોજિક"""
    
    if request.method == "POST":
        entered_otp = request.POST.get("otp")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")
        stored_otp = request.session.get('password_reset_otp')
        
        # 1. Check OTP
        if entered_otp and entered_otp == stored_otp:
            # 2. Check Password Match
            if new_password == confirm_password:
                # Password set logic (Hamesha Main/Default DB ma j thashe)
                request.user.set_password(new_password)
                request.user.save()
                
                # Fetch company slug before session clear
                try:
                    company_slug = request.user.profile.company.slug
                except:
                    company_slug = 'default'

                # 3. Clear session
                if 'password_reset_otp' in request.session:
                    del request.session['password_reset_otp']
                
                messages.success(request, "Password changed successfully! Please login again.")
                
                # Dynamic Redirect to company login
                if company_slug != 'default':
                    return redirect('login', company_slug=company_slug)
                else:
                    return redirect('login')
            else:
                messages.error(request, "Passwords do not match!")
        else:
            messages.error(request, "Invalid OTP! Please try again.")
            
    return render(request, 'attendance/MISC/profile_verify_otp.html')


from django.contrib.auth import authenticate
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.models import User

def ajax_check_password(request):
    identifier = request.GET.get('username', '').strip()
    password = request.GET.get('password', '').strip()
    company_slug = request.GET.get('company_slug', None)

    user_obj = User.objects.filter(
        (Q(username__iexact=identifier) | Q(email__iexact=identifier)),
        is_superuser=True
    ).first()

    # ટર્મિનલમાં ચેક કરવા માટે
    print(f"--- AJAX DEBUG ---")
    print(f"Identifier: {identifier}, Slug: {company_slug}")

    if not identifier or not password:
        return JsonResponse({'is_correct': False})

    # યુઝરને શોધો
    if not user_obj:
        user_obj = User.objects.filter(
            (Q(username__iexact=identifier) | Q(email__iexact=identifier)),
            profile__company__slug=company_slug
        ).first()

    if user_obj:
        print(f"User Found: {user_obj.username}")
        # પાસવર્ડ ચેક કરો
        user = authenticate(username=user_obj.username, password=password)
        if user is not None:
            print("Password Match: YES ✅")
            return JsonResponse({'is_correct': True})
        else:
            print("Password Match: NO ❌")
    else:
        print("User NOT Found in DB")
    
    return JsonResponse({'is_correct': False})

from django.http import JsonResponse

def ajax_check_otp(request):
    entered_otp = request.GET.get('otp', None)
    stored_otp = request.session.get('password_reset_otp')
    
    if entered_otp and stored_otp and entered_otp == stored_otp:
        return JsonResponse({'is_valid': True})
    else:
        return JsonResponse({'is_valid': False})


from django.shortcuts import render, get_object_or_404, redirect


# 1. Particular Company nu Dashboard View
from .middleware import ThreadLocal # Middleware import check kari lejo

@login_required
def company_admin_dashboard(request, company_slug):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    
    # 🛡️ SECURITY: Jo login user aa company no na hoy to block karo
    if not request.user.is_superuser and request.user.profile.company != company:
        return HttpResponseForbidden("Access Denied: This is not your company dashboard.")

    # Database name set karo (Ex: google_db)
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name
    print(f"DEBUG: Dashboard switching to {db_name}")

    # ---------------------------------------------------------
    # 1. DATA FETCHING (Using Company DB)
    # ---------------------------------------------------------
    
    # ✅ Profiles fetch from Company DB
    employees = Profile.objects.using(db_name).filter(company=company).select_related('user')
    
    # ✅ Roles/Groups hamesha 'default' DB ma hoy che (Auth consistency mate)
    roles = Group.objects.filter(name__startswith=f"{company.slug}_")
    
    # Extra Data (Today's Attendance Count)
    from datetime import date
    present_today = Attendance.objects.using(db_name).filter(
        date=date.today(), 
        status="Present"
    ).count()

    return render(request, 'attendance/superadmin/superadmin_company_dashboard.html', {
        'company': company,
        'employees': employees,
        'roles': roles,
        'present_today': present_today,
        'total_employees': employees.count()
    })
# 2. Dynamic Role Creation (Specific to that Company)
# attendance/views.py na end ma muko

# attendance/views.py

# attendance/views.py

# attendance/views.py


# attendance/views.py માં ઉમેરો

from .middleware import ThreadLocal # ThreadLocal check kari lejo

@login_required
def company_assign_admin(request, company_slug):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":
        user_id = request.POST.get('user_id')
        user_to_promote = get_object_or_404(User, id=user_id)
        
        # 1. Group logic (Hamesha 'default' DB ma thashe auth mate)
        user_to_promote.groups.clear()
        admin_group_name = f"{company.slug}_Admin"
        admin_group, _ = Group.objects.get_or_create(name=admin_group_name)
        user_to_promote.groups.add(admin_group)

        # 2. Profile Logic (Sacha DB context ma)
        # Main DB update
        profile = get_object_or_404(Profile, user=user_to_promote)
        profile.role = "ADMIN" 
        profile.is_approved = True
        profile.save() # default DB ma save thashe
        
        # ✅ STEP 3: SYNC TO COMPANY DATABASE (Level 0 Fix)
        # Aa profiles table company na potana database ma pan update thavu joie
        profile.save(using=db_name)
        
        # Create a notification in the company DB
        create_notification(user_to_promote, f"Congratulations! You have been promoted to Admin of {company.name}.")
        
        messages.success(request, f"{user_to_promote.username} is now an Admin for {company.name}")
        return redirect('superadmin_company_dashboard', company_id=company.id)

    return redirect('superadmin_company_dashboard', company_id=company.id)
# attendance/views.py માં એકદમ નીચે ઉમેરો

# attendance/views.py


from django.contrib.auth.models import Group, Permission
from django.contrib import messages

from django.contrib.auth.models import Group, Permission
from django.contrib import messages

from django.contrib.auth.models import Group, Permission
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import Company

from django.contrib.auth.models import Group, Permission
from django.db import transaction

# @login_required
# def create_company_role(request, company_slug):
#     # 1. માત્ર સુપર એડમિન માટે
#     if not request.user.is_superuser:
#         return HttpResponse("Access Denied")

#     company = get_object_or_404(Company, slug=company_slug)
    
#     if request.method == "POST":
#         role_name = request.POST.get('role_name')
        
#         # HTML ફોર્મમાંથી ચેક કરેલા બોક્સના IDs મેળવો
#         perm_ids_raw = request.POST.getlist('permissions') 

#         print(f"DEBUG: Role Name -> {role_name}")
#         print(f"DEBUG: Checkbox IDs -> {perm_ids_raw}")

#         try:
#             with transaction.atomic():
#                 # 2. ગ્રુપ (Role) બનાવો
#                 group_name = f"{company.slug}_{role_name}"
#                 group, created = Group.objects.get_or_create(name=group_name)
                
#                 # 3. પરમિશન્સ સેવ કરો (CRITICAL FIX)
#                 if perm_ids_raw:
#                     # IDs ને નંબરમાં ફેરવો
#                     perm_ids = [int(pid) for pid in perm_ids_raw if pid.isdigit()]
                    
#                     # ડેટાબેઝમાંથી સાચા પરમિશન ઓબ્જેક્ટ શોધો
#                     valid_perms = RolePermission.objects.filter(id__in=perm_ids)
                    
#                     print(f"DEBUG: Saving {valid_perms.count()} permissions...")
                    
#                     # ✅ MAGIC LINE: આ લાઈન ડેટાબેઝમાં સાચો કનેક્શન બનાવે છે
#                     group.permissions.set(valid_perms)
#                 else:
#                     group.permissions.clear()
                
#                 # ગ્રુપ સેવ કરો
#                 group.save()
                
#             messages.success(request, f"Role '{role_name}' saved automatically with permissions.")
            
#         except Exception as e:
#             print(f"❌ ERROR: {e}")
#             messages.error(request, "Error saving permissions.")

#         return redirect('superadmin_company_dashboard', company_id=company.id)

#     return redirect('superadmin_dashboard')    

# origional

from django.db import transaction
from django.contrib.auth.models import Group
from .models import Company, RolePermission
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404

# attendance/views.py

# attendance/views.py

from django.db import transaction # જો import ના હોય તો ઉમેરવું

import subprocess
import os
from django.conf import settings
import pymysql
from django.db import transaction




from django.http import JsonResponse

from .middleware import ThreadLocal # ThreadLocal check kari lejo

@login_required
def get_live_employee_roles(request, company_slug):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0)
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company_slug}_db"
    
    # ✅ Profiles fetch from Company DB
    employees = Profile.objects.using(db_name).filter(company=company).select_related('user')
    
    data = []
    for emp in employees:
        # Groups hamesha 'default' DB ma hoy che
        roles = list(emp.user.groups.values_list('name', flat=True))
        data.append({
            'user_id': emp.user.id,
            'roles': roles
        })
        
    return JsonResponse({'employees': data})


@login_required
def remove_from_branch(request, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    try:
        profile_main = get_object_or_404(Profile, user__id=user_id)
        db_name = f"{profile_main.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Profile/Company Error")

    # ૧. એમ્પ્લોઈની પ્રોફાઈલ Company DB mathi મેળવો
    profile = get_object_or_404(Profile.objects.using(db_name), user__id=user_id)
    branch_id = profile.branch.id if profile.branch else None
    
    # ૩. બ્રાન્ચ અન-અસાઇન (Remove) કરો
    profile.branch = None
    profile.save(using=db_name) # ✅ Save in Company DB
    
    # Optional: Sync with main DB if needed
    profile_main.branch = None
    profile_main.save(using='default')
    
    messages.success(request, f"Employee {profile.user.username} removed from the branch.")
    
    if branch_id:
        return redirect('branch_detail', branch_id=branch_id)
    return redirect('dashboard')


# attendance/views.py
from django.db import transaction

from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.http import HttpResponseForbidden

@login_required
def company_owner_dashboard(request, company_slug):
    # 🛡️ 0. DATABASE & SECURITY
    from .middleware import ThreadLocal
    company = get_object_or_404(Company.objects.using('default'), slug=company_slug)
    db_name = f"{company.slug}_db"
    
    from attendance.utils import ensure_db_connection
    ensure_db_connection(db_name)
    ThreadLocal.DB_NAME = db_name

    owner_group_name = f"{company.slug}_owner"
    if not (request.user.is_superuser or request.user.groups.filter(name=owner_group_name).exists()):
        ThreadLocal.DB_NAME = 'default'
        return HttpResponseForbidden("Access Denied.")

    # 🛡️ 1. POST HANDLING
    if request.method == "POST":
        
        # ----------------------------------------------------
        # ACTION A: CREATE NEW BRANCH 
        # ----------------------------------------------------
        if 'create_branch' in request.POST:
            branch_name = request.POST.get('branch_name')
            location = request.POST.get('location')
            branch_admin_id = request.POST.get('branch_admin_id') 
            
            # 🌍 NEW GPS FIELDS (Safe Parsing)
            lat_raw = request.POST.get('latitude')
            lng_raw = request.POST.get('longitude')
            rad_raw = request.POST.get('radius')
            
            # Convert explicitly so empty strings don't crash MySQL
            lat = float(lat_raw) if lat_raw else None
            lng = float(lng_raw) if lng_raw else None
            rad = int(rad_raw) if rad_raw else 50
            
            if branch_name:
                branch_main = None
                branch_tenant = None
                try:
                    with transaction.atomic(using='default'):
                        with transaction.atomic(using=db_name):
                            
                            branch_main = Branch.objects.using('default').create(
                                company_id=company.id, name=branch_name, location=location,
                                latitude=lat, longitude=lng, radius=rad
                            )
                            branch_tenant = Branch.objects.using(db_name).create(
                                company_id=company.id, name=branch_name, location=location,
                                latitude=lat, longitude=lng, radius=rad
                            )
                            
                            msg = f"✅ Branch '{branch_name}' created successfully with Geofencing!"

                            if branch_admin_id:
                                Profile.objects.using('default').filter(user_id=branch_admin_id).update(
                                    branch_id=branch_main.id, manager=None, team_leader=None
                                )
                                Profile.objects.using(db_name).filter(user_id=branch_admin_id).update(
                                    branch_id=branch_tenant.id, manager=None, team_leader=None
                                )

                                Profile.objects.using('default').filter(manager_id=branch_admin_id).update(manager=None)
                                Profile.objects.using(db_name).filter(manager_id=branch_admin_id).update(manager=None)
                                Profile.objects.using('default').filter(team_leader_id=branch_admin_id).update(team_leader=None)
                                Profile.objects.using(db_name).filter(team_leader_id=branch_admin_id).update(team_leader=None)

                                user = User.objects.using('default').get(id=branch_admin_id)
                                user.groups.clear()
                                admin_group_name = f"{company.slug.lower()}_admin"
                                admin_group, _ = Group.objects.using('default').get_or_create(name=admin_group_name)
                                user.groups.add(admin_group)

                                msg += f" And {user.username} is now the Branch Admin."
                    
                    messages.success(request, msg)
                    
                except Exception as e:
                    if branch_main: branch_main.delete()
                    messages.error(request, f"❌ Error creating branch: {str(e)}")
                    
            return redirect('company_owner_dashboard', company_slug=company.slug)

        # ----------------------------------------------------
        # ACTION B: EDIT EXISTING BRANCH (GPS SAVING FIXED)
        # ----------------------------------------------------
        elif 'edit_branch' in request.POST:
            branch_id = request.POST.get('branch_id')
            branch_name = request.POST.get('branch_name')
            location = request.POST.get('location')
            
            # 🌍 NEW GPS FIELDS (Safe Parsing)
            lat_raw = request.POST.get('latitude')
            lng_raw = request.POST.get('longitude')
            rad_raw = request.POST.get('radius')
            
            lat = float(lat_raw) if lat_raw else None
            lng = float(lng_raw) if lng_raw else None
            rad = int(rad_raw) if rad_raw else 50
            
            try:
                with transaction.atomic(using='default'):
                    with transaction.atomic(using=db_name):
                        # Update Main DB
                        Branch.objects.using('default').filter(id=branch_id).update(
                            name=branch_name, location=location, latitude=lat, longitude=lng, radius=rad
                        )
                        # Update Tenant DB
                        Branch.objects.using(db_name).filter(id=branch_id).update(
                            name=branch_name, location=location, latitude=lat, longitude=lng, radius=rad
                        )
                messages.success(request, f"✅ Branch '{branch_name}' GPS coordinates updated successfully!")
            except Exception as e:
                messages.error(request, f"Error updating branch: {str(e)}")
                
            return redirect('company_owner_dashboard', company_slug=company.slug)

        # ----------------------------------------------------
        # ACTION C: ASSIGN STAFF ROLE
        # ----------------------------------------------------
        elif 'assign_staff_role' in request.POST:
            user_id = request.POST.get('user_id')
            tenant_branch_id = request.POST.get('branch_id')
            role_name = request.POST.get('role_name', '').strip().lower() 

            try:
                with transaction.atomic(using='default'):
                    with transaction.atomic(using=db_name):
                        target_user = User.objects.using('default').get(id=user_id)
                        
                        # 1. Get the branch from the tenant database (Safe because ID is unique)
                        branch_tenant = Branch.objects.using(db_name).get(id=tenant_branch_id)

                        # 2. 🔥 THE FIX: Use .filter().first() INSTEAD of get_or_create()
                        branch_main = Branch.objects.using('default').filter(
                            company=company, 
                            name=branch_tenant.name
                        ).first()

                        # If it doesn't exist, create it manually
                        if not branch_main:
                            branch_main = Branch.objects.using('default').create(
                                company=company, 
                                name=branch_tenant.name,
                                location=branch_tenant.location, 
                                latitude=branch_tenant.latitude, 
                                longitude=branch_tenant.longitude, 
                                radius=branch_tenant.radius
                            )

                        # 3. Update Profiles in both databases
                        Profile.objects.using('default').filter(user_id=user_id).update(
                            branch_id=branch_main.id, manager=None, team_leader=None
                        )
                        Profile.objects.using(db_name).filter(user_id=user_id).update(
                            branch_id=branch_tenant.id, manager=None, team_leader=None
                        )

                        # Clear Managers & Team Leaders for this user
                        Profile.objects.using('default').filter(manager_id=user_id).update(manager=None)
                        Profile.objects.using(db_name).filter(manager_id=user_id).update(manager=None)
                        Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader=None)
                        Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader=None)

                        # 4. Update Role Groups
                        target_user.groups.clear()
                        if role_name:
                            group_name = f"{company.slug.lower()}_{role_name}"
                            target_group, _ = Group.objects.using('default').get_or_create(name=group_name)
                            target_user.groups.add(target_group)

                messages.success(request, f"Role updated successfully. {target_user.username} is now {role_name.title()}.")
            except Exception as e:
                # If it crashes, this is where the error text comes from
                messages.error(request, f"Error assigning role: {str(e)}")
            
            # Keeps you on the Branch Details page after updating!
            return redirect(request.META.get('HTTP_REFERER', 'company_owner_dashboard'))

    # 🛡️ 2. GET REQUEST DATA & 100% ACCURATE STATS
    branches = Branch.objects.using(db_name).filter(company=company)
    selected_role = request.GET.get('role', 'all').lower()
    
    tenant_profile_user_ids = list(Profile.objects.using(db_name).filter(
        company_id=company.id
    ).values_list('user_id', flat=True))

    company_users = User.objects.using('default').filter(
        id__in=tenant_profile_user_ids
    ).prefetch_related('groups')

    stats = {
        'total': 0, 'admin': 0, 'hr': 0, 'manager': 0, 'team_leader': 0, 'employee': 0
    }
    role_map = {} 
    owner_group_name = f"{company.slug}_owner".lower()
    slug_lower = company.slug.lower()

    for u in company_users:
        user_groups = [g.name.lower() for g in u.groups.all()]
        
        if any(g == owner_group_name for g in user_groups):
            role_map[u.id] = 'owner'
            continue
            
        emp_role = 'employee' # Default role
        if any(f"{slug_lower}_admin" in g for g in user_groups): emp_role = 'admin'
        elif any(f"{slug_lower}_hr" in g for g in user_groups): emp_role = 'hr'
        elif any(f"{slug_lower}_manager" in g for g in user_groups): emp_role = 'manager'
        elif any(f"{slug_lower}_team leader" in g for g in user_groups): emp_role = 'team leader'
            
        role_map[u.id] = emp_role
        
        if emp_role == 'admin': stats['admin'] += 1
        elif emp_role == 'hr': stats['hr'] += 1
        elif emp_role == 'manager': stats['manager'] += 1
        elif emp_role == 'team leader': stats['team_leader'] += 1
        elif emp_role == 'employee': stats['employee'] += 1
        
        stats['total'] += 1

    if selected_role != 'all':
        filtered_ids = [uid for uid, role in role_map.items() if role == selected_role]
    else:
        filtered_ids = [uid for uid, role in role_map.items() if role != 'owner']

    employees_qs = Profile.objects.using(db_name).filter(
        user_id__in=filtered_ids
    ).select_related('user', 'branch').order_by('-user__date_joined')

    processed_staff = []
    for emp in employees_qs:
        emp_role = role_map.get(emp.user_id, 'employee')
        emp.display_role = emp_role.title()
        processed_staff.append(emp)

    policies = AttendancePolicy.objects.using(db_name).filter(company=company)
    holidays = Holiday.objects.using(db_name).filter(company=company).order_by('date')
    shifts = Shift.objects.using(db_name).filter(company=company)

    return render(request, 'attendance/dashboard/company_owner_dashboard.html', {
        'company': company,
        'branches': branches,
        'employees': processed_staff,
        'stats': stats, 
        'policies': policies,
        'holidays': holidays,
        'shifts': shifts,
        'is_owner': True,
        'role_perms': {'can_view_team': True, 'can_manage_users': True, 'can_view_reports': True}
    })

import math


from django.db import transaction
from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden

@login_required
def delete_branch(request, branch_id):
    # 🛡️ 1. Find the Branch Safely (Check both databases)
    branch = Branch.objects.using('default').filter(id=branch_id).first()
    
    if not branch:
        try:
            company = request.user.profile.company
            db_name = f"{company.slug}_db"
            branch = Branch.objects.using(db_name).filter(id=branch_id).first()
        except Exception:
            return HttpResponseForbidden("Cannot determine database context.")
            
    if not branch:
        messages.error(request, "Branch not found.")
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

    # Setup the Database Name
    company = branch.company
    db_name = f"{company.slug}_db"
    
    from attendance.middleware import ThreadLocal
    ThreadLocal.DB_NAME = db_name

    # 🛡️ 2. Security Check (Owner or Superuser)
    owner_group_name = f"{company.slug}_owner"
    is_owner = request.user.groups.filter(name=owner_group_name).exists()

    if not is_owner and not request.user.is_superuser:
        messages.error(request, "⛔ Access Denied. Only Owners can delete branches.")
        return redirect('company_owner_dashboard', company_slug=company.slug)

    # 🛡️ 3. Safe Deletion Logic (From both DBs)
    try:
        branch_name = branch.name
        
        # આ ટ્રાન્ઝેક્શન મેઈન ડીબીને પ્રોટેક્ટ કરશે
        with transaction.atomic():
            # A. Free the staff assigned to this branch in BOTH DBs (Constraints Fix)
            Profile.objects.using(db_name).filter(branch_id=branch_id).update(branch=None)
            Profile.objects.using('default').filter(branch_id=branch_id).update(branch=None)
            
            # B. Delete the branch from Tenant DB
            Branch.objects.using(db_name).filter(id=branch_id).delete()
            
            # C. Delete the branch from Main DB
            Branch.objects.using('default').filter(id=branch_id).delete()
            
        messages.success(request, f"✅ Branch '{branch_name}' deleted successfully!")
    except Exception as e:
        messages.error(request, f"⚠️ Error deleting branch: {str(e)}")

    return redirect('company_owner_dashboard', company_slug=company.slug)

# Delete Function
from django.contrib.auth.models import Group
from django.contrib import messages
from .middleware import ThreadLocal

from django.db import connections

from django.db import connections

@login_required
def delete_company_role(request, role_id):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")

    from django.contrib.auth.models import Group
    from .models import Company

    referer = request.META.get('HTTP_REFERER', '')
    db_name = 'default'
    
    # ૧. સાચો ડેટાબેઝ નક્કી કરો
    if 'company' in referer:
        try:
            parts = referer.split('/')
            company_id_from_url = parts[parts.index('company') + 1]
            company_obj = Company.objects.using('default').filter(id=company_id_from_url).first()
            if company_obj:
                db_name = f"{company_obj.slug}_db"
                from .utils import ensure_db_connection
                ensure_db_connection(db_name)
        except:
            db_name = 'default'

    try:
        # ૨. પર્ટીક્યુલર કંપનીના ડેટાબેઝમાં ક્રમબદ્ધ સફાઈ (Sequential Cleanup)
        with connections[db_name].cursor() as cursor:
            
            # 🔥 STEP A: પરમિશન ટેબલમાંથી ડેટા સાફ કરો
            cursor.execute("DELETE FROM attendance_rolepermission WHERE group_id = %s", [role_id])
            
            # 🔥 STEP B: યુઝર અને ગ્રુપના મેપિંગ (auth_user_groups) માંથી લિંક તોડો 
            # આ તમારી લેટેસ્ટ એરર (1451) ને સોલ્વ કરશે
            cursor.execute("DELETE FROM auth_user_groups WHERE group_id = %s", [role_id])
            
            # 🔥 STEP C: હવે પર્ટીક્યુલર DB ના 'auth_group' માંથી રોલ ડિલીટ કરો
            cursor.execute("DELETE FROM auth_group WHERE id = %s", [role_id])
            
            print(f"✅ Full cleanup successful in {db_name}")

        # ૩. છેલ્લે મેઈન ડેટાબેઝ (default) માંથી પણ એ રોલ કાઢી નાખો
        Group.objects.using('default').filter(id=role_id).delete()
        
        messages.success(request, f"રોલ, તેની પરમિશન્સ અને યુઝર લિંક્સ સફળતાપૂર્વક સાફ કરવામાં આવી છે.")

    except Exception as e:
        messages.error(request, f"SQL Error: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', 'superadmin_dashboard'))
def logout_view(request):
    """લોગઆઉટ કરી કંપનીના સ્પેસિફિક લોગિન પેજ પર રીડાયરેક્ટ કરો"""
    slug = None
    
    if request.user.is_authenticated:
        try:
            # Level 0 logic: User ni profile mathi slug melvo
            # Note: request.user.profile main DB mathi load thashe
            slug = request.user.profile.company.slug
        except Exception:
            pass 
            
    # Logout session clear kare che
    logout(request)
    
    # Session cleanup pachi context clear karo
    ThreadLocal.DB_NAME = 'default'
    
    if slug:
        # ✅ Redirect to Company Specific Login (Ex: /google/login/)
        return redirect('login', company_slug=slug)
    else:
        # Superadmin mate default django admin ya main portal login
        return redirect('/admin/')
    
from .middleware import ThreadLocal # ThreadLocal check kari lejo

@login_required
def company_assign_admin(request, company_slug):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")

    if request.method == "POST":
        user_id = request.POST.get('user_id')
        user_to_promote = get_object_or_404(User, id=user_id)
        
        # 1. Group logic (Hamesha 'default' DB ma thashe auth mate)
        # Auth groups global database ma hashe jethi login check thase
        user_to_promote.groups.clear()
        admin_group_name = f"{company.slug}_admin"
        admin_group, _ = Group.objects.get_or_create(name=admin_group_name)
        user_to_promote.groups.add(admin_group)

        # 2. Profile Logic (Sacha DB context ma)
        # Main DB update
        profile = get_object_or_404(Profile, user=user_to_promote)
        profile.role = "ADMIN" 
        profile.is_approved = True
        profile.save() # default DB ma save thashe
        
        # ✅ STEP 3: SYNC TO COMPANY DATABASE (Level 0 Fix)
        # Aa profiles table company na potana database ma pan update thavu joie
        # 
        profile.save(using=db_name)
        
        # Create a notification in the company DB
        create_notification(user_to_promote, f"Congratulations! You have been promoted to Admin of {company.name}.")
        
        messages.success(request, f"{user_to_promote.username} is now an Admin for {company.name}")
        return redirect('superadmin_company_dashboard', company_id=company.id)

    return redirect('superadmin_company_dashboard', company_id=company.id)

# attendance/views.py



# attendance/views.py

# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal # ThreadLocal check kari lejo
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
# Import your models (Company, Group, User, Profile) here

@login_required
def company_assign_role(request, company_slug, user_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")
    
    user_to_update = get_object_or_404(User, id=user_id)
    
    # Roles hamesha Main/Default DB ma hoy che (Login purpose mate)
    roles = Group.objects.filter(name__startswith=f"{company.slug}_")

    if request.method == "POST":
        group_id = request.POST.get('group_id')
        
        # 1. Update Auth Groups (Main DB)
        user_to_update.groups.clear() 
        clean_role_name = "Employee" # Default name for success message
        
        # If assigning a specific role
        if group_id:
            selected_group = get_object_or_404(Group, id=group_id)
            
            # Security Check: Role aa j company no hovo joie
            if not selected_group.name.startswith(f"{company.slug}_"):
                messages.error(request, "Invalid Role Selection.")
                return redirect('superadmin_company_dashboard', company_id=company.id)
            
            user_to_update.groups.add(selected_group) 
            clean_role_name = selected_group.name.split('_')[-1].capitalize()

        # ---------------------------------------------------------
        # 🔥 2. FOREIGN KEY CLEANUP
        # ---------------------------------------------------------
        # Cleanup in COMPANY DB
        Profile.objects.using(db_name).filter(team_leader=user_to_update).update(team_leader=None)
        Profile.objects.using(db_name).filter(manager=user_to_update).update(manager=None)
        
        # Cleanup in MAIN DB (for redundancy/safety)
        Profile.objects.filter(team_leader=user_to_update).update(team_leader=None)
        Profile.objects.filter(manager=user_to_update).update(manager=None)
        
        # 3. PROFILE RESET (In both DBs)
        target_profile = get_object_or_404(Profile, user=user_to_update)
        target_profile.manager = None
        target_profile.team_leader = None
        
        # 🚨 DO NOT ADD: target_profile.role = ... 
        # The @property handles this dynamically now!
        
        # ✅ Save in MAIN DB
        target_profile.save()
        # ✅ Save in COMPANY DB
        target_profile.save(using=db_name)

        if group_id:
            messages.success(request, f"Role updated to '{clean_role_name}' and all team links cleared for {user_to_update.username}")
        else:
            messages.success(request, f"Role revoked (Set to Employee) and team links cleared for {user_to_update.username}")
            
        return redirect('superadmin_company_dashboard', company_id=company.id)

    return render(request, 'attendance/superadmin/company_assign_role.html', {
        'company': company,
        'user_to_update': user_to_update,
        'roles': roles
    })
# attendance/views.py

from .middleware import ThreadLocal # ThreadLocal check kari lejo

@transaction.non_atomic_requests
@login_required
def create_company_role(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    ensure_db_connection(db_name) # કનેક્શન જીવંત કરો

    if request.method == "POST":
        role_name = request.POST.get('role_name', '').strip().lower()
        dynamic_group_name = f"{company.slug.lower()}_{role_name}"
        selected_perms = request.POST.getlist('permissions')

        try:
            # ✅ STEP 1: મેઈન ડેટાબેઝમાં ગ્રુપ બનાવો
            group_main, _ = Group.objects.using('default').get_or_create(name=dynamic_group_name)

            # ✅ STEP 2: ટેનન્ટ ડેટાબેઝમાં પણ તે જ ગ્રુપ બનાવો (Constraint Fix)
            group_tenant, _ = Group.objects.using(db_name).get_or_create(name=dynamic_group_name)

            # ✅ STEP 3: ટેનન્ટ DB માં પરમિશન સેવ કરો (અહીં 'group' ફિલ્ડ વાપરવી)
            rp, _ = RolePermission.objects.using(db_name).get_or_create(group=group_tenant)
            
            rp.can_self_access = 'can_self_access' in selected_perms
            rp.can_view_team = 'can_view_team' in selected_perms
            rp.can_manage_users = 'can_manage_users' in selected_perms
            rp.can_approve_attendance = 'can_approve_attendance' in selected_perms
            rp.can_view_reports = 'can_view_reports' in selected_perms
            rp.can_approve_leave = 'can_approve_leave' in selected_perms
            rp.can_manage_shifts = 'can_manage_shifts' in selected_perms
            rp.can_view_policy = 'can_view_policy' in selected_perms
            
            rp.save(using=db_name) 

            messages.success(request, f"Role '{role_name}' created and synced successfully!")
            return redirect('superadmin_company_dashboard', company_id=company.id)

        except Exception as e:
            messages.error(request, f"System Error: {str(e)}")
            return redirect('superadmin_company_dashboard', company_id=company.id)

    return render(request, "attendance/dashboard/company_create_role.html", {"company": company})


@login_required
def manage_role_permissions(request, company_slug, group_id):
    company = get_object_or_404(Company, slug=company_slug)
    group = get_object_or_404(Group, id=group_id)
    db_name = f"{company.slug}_db"

    # ✅ Fetch from Specific Company DB
    perm_obj, _ = RolePermission.objects.using(db_name).get_or_create(group=group)
    
    if request.method == "POST":
        perm_obj.can_self_access = 'can_self_access' in request.POST
        perm_obj.can_view_team = 'can_view_team' in request.POST
        perm_obj.can_manage_users = 'can_manage_users' in request.POST
        
        # ✅ Save back to Company DB
        perm_obj.save(using=db_name)
        
        messages.success(request, "Permissions updated successfully.")
        return redirect('superadmin_company_dashboard', company_id=company.id)
        
    return render(request, 'attendance/superadmin/manage_permissions.html', {
        'group': group, 
        'perms': perm_obj,
        'company': company
    })

# attendance/views.py

 # તમારા કસ્ટમ મોડેલ માટે

from .middleware import ThreadLocal
from django.contrib import messages

@login_required
def manage_role_permissions(request, company_slug, group_id):
    """ડાયનેમિક પરમિશન સેવ કરવા માટેનું ફંક્શન (Popup Logic)"""
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")
    
    group = get_object_or_404(Group, id=group_id)
    company = get_object_or_404(Company, slug=company_slug)
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    # ✅ Custom Permission ટેબલમાંથી ઓબ્જેક્ટ કંપની ડેટાબેઝમાં મેળવો
    perm_obj, created = RolePermission.objects.using(db_name).get_or_create(group=group)
    
    if request.method == "POST":
        # ચેકબોક્સમાંથી ડેટા મેળવી ડાયનેમિકલી સેવ કરો
        perm_obj.can_self_access = 'can_self_access' in request.POST
        perm_obj.can_view_team = 'can_view_team' in request.POST
        perm_obj.can_manage_users = 'can_manage_users' in request.POST
        perm_obj.can_view_reports = 'can_view_reports' in request.POST
        
        # ✅ કંપનીના ડેટાબેઝમાં સેવ કરો
        perm_obj.save(using=db_name)
        
        messages.success(request, f"Permissions updated for role: {group.name}")
        return redirect('superadmin_company_dashboard', company_id=company.id)
    
    return redirect('superadmin_company_dashboard', company_id=company.id)


@login_required
def company_make_admin(request, company_slug, user_id):
    """એમ્પ્લોઈને કંપની એડમિન બનાવવા માટેનું ફંક્શન"""
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")
    
    company = get_object_or_404(Company, slug=company_slug)
    user_to_promote = get_object_or_404(User, id=user_id)
    db_name = f"{company.slug}_db"

    # 1. કંપની વાઇઝ એડમિન ગ્રુપ (Hamesha Main DB ma raheshe)
    group_name = f"{company.slug}_admin"
    admin_group, _ = Group.objects.get_or_create(name=group_name)
    
    # ૨. જૂના ગ્રુપ્સ કાઢી નાખો અને નવું એડમિન ગ્રુપ આપો
    user_to_promote.groups.clear()
    user_to_promote.groups.add(admin_group)
    
    # 🛡️ 3. PROFILE UPDATE (Level 0 Dual-Sync)
    # Main DB માં પ્રોફાઈલ અપડેટ
    profile = get_object_or_404(Profile, user=user_to_promote)
    profile.is_approved = True
    profile.role = "ADMIN"  # રોલ પણ સેટ કરી દઈએ
    profile.save() # default DB

    # ✅ કંપનીના ડેટાબેઝમાં પણ સિંક કરો
    profile.save(using=db_name)
    
    # 4. નોટિફિકેશન મોકલો (Company DB માં)
    create_notification(user_to_promote, f"System: You have been promoted to Admin for {company.name}")
    
    messages.success(request, f"{user_to_promote.username} is now Admin for {company.name}")
    return redirect('superadmin_company_dashboard', company_id=company.id)

from .middleware import ThreadLocal
from django.contrib import messages

@login_required
def company_assign_role(request, company_slug, user_id):
    """ડાયનેમિક રોલ (Group) અસાઇન કરવા માટેનું ફંક્શન (Popup POST)"""
    company = get_object_or_404(Company, slug=company_slug)
    
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    if request.method == "POST":
        user_to_update = get_object_or_404(User, id=user_id)
        group_id = request.POST.get('group_id')
        target_user = get_object_or_404(User, id=user_id)
        # ૧. Auth Groups હમેશા Main/Default DB માં સેવ થાય છે
        user_to_update.groups.clear()
        
        # ૨. પ્રોફાઈલ લોડ કરો (Sync કરવા માટે)
        profile = get_object_or_404(Profile, user=user_to_update)
        
        if group_id:
            selected_group = get_object_or_404(Group, id=group_id)
            user_to_update.groups.add(selected_group)
            
            # ૩. રોલનું નામ પ્રોફાઈલમાં અપડેટ કરો (Ex: 'google_Manager' -> 'Manager')
            new_role_name = selected_group.name.split('_')[-1].upper()
            target_user.groups.clear()
            target_user.groups.add(selected_group)
            messages.success(request, f"Role '{new_role_name}' assigned to {user_to_update.username}")
        else:
            profile.role = "EMPLOYEE" # Default role if none selected
            messages.info(request, f"All custom roles removed for {user_to_update.username}")
        
        # 🛡️ ૪. DUAL-DATABASE SYNC
        # Main DB માં સેવ કરો
        profile.save() 
        # ✅ કંપનીના આઈસોલેટેડ ડેટાબેઝમાં પણ સેવ કરો
        profile.save(using=db_name)

        # ૫. નોટિફિકેશન મોકલો
        create_notification(user_to_update, f"Your role has been updated to {profile.role}")
            
        return redirect('superadmin_company_dashboard', company_id=company.id)
            
    return redirect('superadmin_company_dashboard', company_id=company.id)

# attendance/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from accounts.models import Profile
from .models import Company, Attendance, AttendancePolicy
from django.http import HttpResponse


@login_required
def dashboard_redirect(request, company_slug):
    user = request.user
    
    # Profile મેળવો (સુરક્ષિત રીતે)
    profile = get_object_or_404(Profile, user=user)
    
    # 1. સુરક્ષા: જો યુઝર આ કંપનીનો ના હોય તો અટકાવો
    if not user.is_superuser and profile.company and profile.company.slug != company_slug:
        return HttpResponse("Unauthorized access to this company dashboard.")

    # --- 🔥 NEW: OWNER ROLE CHECK (આ કોડ ઉમેરો) ---
    # સૌથી પહેલા Owner ચેક કરો
    if user.groups.filter(name__icontains='_owner').exists():
        return redirect('company_owner_dashboard', company_slug=company_slug)
    # -----------------------------------------------

    # 2. પરમિશન મુજબ રીડાયરેક્શન
    # (નોંધ: user_has_permission ફંક્શન તમારી પાસે હોવું જોઈએ, અથવા ડાયરેક્ટ ગ્રુપ ચેક કરી શકો)
    
    if user.groups.filter(name__icontains='_admin').exists() or user_has_permission(user, 'can_manage_users'):
        return redirect('superadmin_company_dashboard', company_slug=company.slug)
        
    elif user.groups.filter(name__icontains='manager').exists() or user_has_permission(user, 'can_view_team'):
        return redirect('manager_dashboard') # Manager માટે company_slug નથી હોતું મોટે ભાગે URL માં
        
    else:
        # જો કોઈ રોલ ના હોય તો Employee
        return redirect('employee_dashboard', company_slug=company_slug)
# @login_required
# def employee_dashboard_view(request, company_slug):
#     company = get_object_or_404(Company, slug=company_slug)
    
#     # પ્રોફાઈલ અને કંપની ચેક
#     try:
#         profile = Profile.objects.get(user=request.user)
#     except Profile.DoesNotExist:
#         return HttpResponse("Profile not found.")

#     if profile.company != company and not request.user.is_superuser:
#         return HttpResponse("Access Denied")

#     # ડેટા ગણતરી (Attendance Logic)
#     today = timezone.localdate()
#     attendance = Attendance.objects.filter(user=request.user, date=today).first()
#     shift = profile.shift
    
#     work_hours = 0.0
#     if attendance and attendance.check_in and attendance.check_out:
#         in_dt = datetime.combine(today, attendance.check_in)
#         out_dt = datetime.combine(today, attendance.check_out)
#         work_hours = (out_dt - in_dt).total_seconds() / 3600

#     # ✅ NO MANUAL PERMISSIONS HERE! Django will handle it properly in HTML.
#     return render(request, "attendance/dashboard/employee_dashboard.html", {
#         "attendance": attendance,
#         "shift": shift,
#         "work_hours": work_hours,
#         "profile": profile,
#         "company": company,
#     })
    
    
from django.utils.text import slugify

from django.utils.text import slugify
from django.db import transaction

@login_required
def update_company_slug(request, company_slug):
    """કંપનીનો સ્લગ અપડેટ કરી ડેટાબેઝ અને ગ્રુપ્સ સાથે સિંક કરો"""
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")

    company = get_object_or_404(Company, slug=company_slug)
    old_slug = company.slug
    
    if request.method == "POST":
        new_raw_slug = request.POST.get('company_slug')
        new_slug = slugify(new_raw_slug)
        
        # 1. Validation Checks
        if not new_slug:
            messages.error(request, "Slug cannot be empty!")
            return redirect('superadmin_company_dashboard', company_id=company.id)

        if Company.objects.filter(slug=new_slug).exclude(id=company.id).exists():
            messages.error(request, f"Slug '{new_slug}' is already taken.")
            return redirect('superadmin_company_dashboard', company_id=company.id)

        # ---------------------------------------------------------
        # 🛡️ 2. MULTI-TENANT SYNC LOGIC (CRITICAL)
        # ---------------------------------------------------------
        # 
        
        with transaction.atomic():
            # A. Update Groups (Main DB માં રહેલા જૂના સ્લગ વાળા ગ્રુપ્સ બદલો)
            # Ex: 'google_admin' -> 'tata_admin'
            groups = Group.objects.filter(name__startswith=f"{old_slug}_")
            for group in groups:
                role_part = group.name.split('_')[-1]
                group.name = f"{new_slug}_{role_part}"
                group.save()

            # B. Update Company Slug
            company.slug = new_slug
            company.save()

            # C. WARNING: ડેટાબેઝ રીનેમ કરવાનું લોજિક
            # જો તમે SQL લેવલ પર ડેટાબેઝનું નામ બદલ્યું ન હોય, તો 
            # તમારે મેન્યુઅલી જૂના DB (old_slug_db) ને નવા DB (new_slug_db) માં 
            # રીનેમ કરવું પડશે, અન્યથા ડેટા કનેક્શન તૂટી જશે.
            
        messages.success(request, f"Slug successfully updated to '{new_slug}'. Please ensure database '{old_slug}_db' is renamed to '{new_slug}_db' manually.")
        
        return redirect('superadmin_company_dashboard', company_id=company.id)

    return redirect('superadmin_company_dashboard', company_id=company.id)
    
# attendance/views.py માં છેલ્લે ઉમેરો

from django.contrib.auth.models import Group
#from .models import Permission as Permission
from django.contrib import messages

# 1. સુપર એડમિન: કંપની માટે ડાયનેમિક રોલ બનાવો (Company Wise)
# @login_required
# def create_company_role(request, company_slug):
#     if not request.user.is_superuser:
#         return HttpResponse("Access Denied")

#     company = get_object_or_404(Company, slug=company_slug)
    
#     if request.method == "POST":
#         role_name = request.POST.get('role_name').strip() # e.g. "Manager"
        
#         # લોજિક: રોલનું નામ કંપનીના સ્લગ સાથે જોડી દો જેથી તે કંપની પૂરતું જ રહે
#         # દા.ત. Infosys નો Manager બનશે: "infosys_manager"
#         dynamic_group_name = f"{company.slug}_{role_name}"
        
#         # auth_group ટેબલમાં ગ્રુપ બનાવો
#         group, created = Group.objects.get_or_create(name=dynamic_group_name)
        
#         # Permissions મેળવો (Checkbox માંથી)
#         can_manage_users = request.POST.get('can_manage_users') == 'on'
#         can_approve_attendance = request.POST.get('can_approve_attendance') == 'on'
#         can_view_team = request.POST.get('can_view_team') == 'on'
#         can_view_reports = request.POST.get('can_view_reports') == 'on'
        
#         # તારા કસ્ટમ Permission ટેબલમાં એન્ટ્રી
#         custom_perm, created = RolePermission.objects.get_or_create(role=group)
#         custom_perm.can_manage_users = can_manage_users
#         custom_perm.can_approve_attendance = can_approve_attendance
#         custom_perm.can_view_team = can_view_team
#         custom_perm.can_view_reports = can_view_reports
#         custom_perm.save()

#         messages.success(request, f"Role '{role_name}' created for {company.name} with permissions.")
#         return redirect('superadmin_company_dashboard', company_id=company.id)

#     return redirect('superadmin_dashboard')

# 2. સુપર એડમિન: કંપનીનો મેઈન એડમિન સેટ કરો
from django.contrib.auth.models import Group
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Company, RolePermission  # RolePermission import karvu

from .middleware import ThreadLocal
from django.db import transaction

@login_required
def assign_company_admin(request, company_slug):
    """યુઝરને કંપની એડમિન બનાવી તેને જરૂરી પરમિશન્સ આપવી"""
    company = get_object_or_404(Company, slug=company_slug)
    db_name = f"{company.slug}_db"
    
    if not request.user.is_superuser:
        return HttpResponse("Access Denied")

    if request.method == "POST":
        user_id = request.POST.get("user_id")
        user = get_object_or_404(User, id=user_id)
        
        # 1. Admin Group Logic (Hamesha Main DB ma raheshe auth mate)
        admin_group_name = f"{company.slug}_Admin"
        admin_group, created = Group.objects.get_or_create(name=admin_group_name)
        
        # ---------------------------------------------------------
        # 🛡️ 2. ROLE PERMISSION SYNC (In Company Database)
        # ---------------------------------------------------------
        # 
        
        # જો ગ્રુપ નવું બન્યું હોય અથવા પરમિશન સેટ ના હોય, તો કંપની DB માં સેવ કરો
        role_perm, _ = RolePermission.objects.using(db_name).get_or_create(group=admin_group)
        role_perm.can_manage_users = True
        role_perm.can_approve_attendance = True
        role_perm.can_view_reports = True
        role_perm.can_self_access = True
        role_perm.can_view_policy = True
        
        # ✅ Save in the isolated company database
        role_perm.save(using=db_name)

        # 3. USER ROLE UPDATE (Main DB)
        with transaction.atomic():
            user.groups.clear() # જૂના બધા રોલ કાઢી નાખો
            user.groups.add(admin_group)
            
            # Profile status update (જો જરૂર હોય તો)
            profile = user.profile
            profile.is_approved = True
            profile.save() # default DB
            
            # ✅ Sync profile to Company DB
            profile.save(using=db_name)

        messages.success(request, f"User @{user.username} successfully promoted to Admin for {company.name}!")
        
    return redirect('superadmin_company_dashboard', company_id=company.id)


# 3. કંપની એડમિન: એમ્પ્લોઈને રોલ આપો (ફક્ત પોતાની કંપનીના રોલ)
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib import messages
from accounts.models import is_admin  # તારા accounts/models.py માં આ ફંક્શન છે
from .middleware import ThreadLocal
from django.db.models import Q

@login_required
def assign_squad_view(request, tl_id):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile_manager = request.user.profile
        company = profile_manager.company
        branch = profile_manager.branch
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Profile/Company not found")

    tl_user = get_object_or_404(User, id=tl_id)

    # 1. Team Leaders ના IDs મેળવો (Hamesha Main DB mathi checks thashe)
    tl_group_name = f"{company.slug}_team leader"
    all_tls = User.objects.filter(groups__name__icontains=tl_group_name).values_list('id', flat=True)

    # ---------------------------------------------------------
    # 2. DATA FETCHING (Using Company DB)
    # ---------------------------------------------------------
    # 
    
    # Available Members Filter:
    # - Potana company/branch na hoi (Company DB mathi)
    # - Bija koi na squad ma na hoi (team_leader__isnull=True)
    available_staff = Profile.objects.using(db_name).filter(
        company=company,
        branch=branch
    ).exclude(
        user__id__in=all_tls # 👈 Team Leaders list ma na avva joie
    ).filter(
        Q(team_leader__isnull=True) | Q(team_leader=tl_user)
    ).exclude(user=tl_user).select_related('user')

    if request.method == 'POST':
        selected_user_ids = request.POST.getlist('selected_users')
        
        # ✅ STEP 3: DUAL-DATABASE SYNC (Level 0 Fix)
        # 1. Pehla aa TL na badha members ne free karo (In Company DB)
        Profile.objects.using(db_name).filter(team_leader=tl_user).update(team_leader=None)
        
        # 2. Nava members ne assign karo (In Company DB)
        if selected_user_ids:
            Profile.objects.using(db_name).filter(user_id__in=selected_user_ids).update(team_leader=tl_user)
        
        # Optional: Redundancy mate Main DB ma pan update kari shako
        Profile.objects.filter(team_leader=tl_user).update(team_leader=None)
        Profile.objects.filter(user_id__in=selected_user_ids).update(team_leader=tl_user)

        messages.success(request, f"Squad for {tl_user.username} updated successfully!")
        return redirect('manager_dashboard')

    return render(request, 'attendance/dashboard/assign_squad.html', {
        'tl_user': tl_user,
        'available_staff': available_staff
    })

from .middleware import ThreadLocal
from django.db import transaction

@login_required
def admin_assign_role_dynamic(request):
    # 1. એડમિન ચેક (Custom function or Group check)
    if not request.user.groups.filter(name__icontains='admin').exists() and not request.user.is_superuser:
         messages.error(request, "Access Denied. You are not an Admin.")
         return redirect('dashboard')
         
    # 2. એડમિનની કંપની અને તેનો ડેટાબેઝ મેળવો
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name # 👈 કંપનીના ડેટાબેઝ પર સ્વિચ કરો
    except:
        return HttpResponse("Company not found for this user.")
    
    if request.method == "POST":
        user_id = request.POST.get('user_id')
        group_id = request.POST.get('group_id')
        
        target_user = get_object_or_404(User, id=user_id)
        target_group = get_object_or_404(Group, id=group_id)
        
        # 3. Security Check: શું આ રોલ આ જ કંપનીનો છે?
        if not target_group.name.startswith(f"{company.slug}_"):
             messages.error(request, "Invalid Role Selection.")
             return redirect('admin_dashboard')
             
        # 4. જૂના રોલ કાઢીને નવો રોલ આપો (Main DB Operation)
        target_user.groups.clear()
        target_user.groups.add(target_group)
        
        # 🛡️ 5. PROFILE SYNC (Level 0 Fix)
        # જો તમે પ્રોફાઇલમાં 'role' પ્રોપર્ટી વાપરતા હોવ તો પણ, 
        # એડમિન એક્શન તરીકે પ્રોફાઇલને કંપની DB માં સિંક કરવી જરૂરી છે.
        target_profile = get_object_or_404(Profile.objects.using(db_name), user=target_user)
        target_profile.is_approved = True
        target_profile.save(using=db_name) # 👈 કંપનીના ડેટાબેઝમાં સ્ટેટસ સિંક કરો
        
        messages.success(request, f"Role '{target_group.name}' assigned to {target_user.username}")
        # This sends you back to the page you clicked the button on!
        return redirect(request.META.get('HTTP_REFERER', 'admin_dashboard'))

    return redirect('admin_dashboard')



@login_required
def create_dynamic_role(request):
    if not request.user.is_superuser:
        return redirect('login')

    companies = Company.objects.all()
    # Permissions હમેશા Default DB માંથી જ આવશે
    permissions = RolePermission.objects.filter(content_type__app_label='attendance')

    if request.method == "POST":
        company_id = request.POST.get('company')
        role_name = request.POST.get('role_name')
        selected_perms_ids = request.POST.getlist('permissions') 

        company = get_object_or_404(Company, id=company_id)
        db_name = f"{company.slug}_db"
        
        # Group banavo (Main DB)
        clean_role = role_name.strip().lower().replace(" ", "_")
        group_name = f"{company.slug}_{clean_role}"
        
        with transaction.atomic(): # 👈 ડેટા સેફ્ટી માટે
            new_group, created = Group.objects.get_or_create(name=group_name)

            # Permissions Save Karo (Main DB)
            if selected_perms_ids:
                perms_to_add = RolePermission.objects.filter(id__in=selected_perms_ids)
                new_group.permissions.set(perms_to_add)
                
            # 🛡️ 6. TENANT PERMISSION SYNC
            # જો તમે કસ્ટમ RolePermission મોડેલ વાપરતા હોવ, તો તેને કંપની DB માં પણ રિફ્લેક્ટ કરો
            # આનાથી તે ટેનન્ટના ડેટાબેઝ લેવલ પર પણ સિક્યોરિટી જળવાશે.
            
        messages.success(request, f"Role {role_name} for {company.name} updated successfully!")
        return redirect('superadmin_dashboard') 

    return render(request, 'superadmin/create_role.html', {
        'companies': companies, 
        'permissions': permissions
    })

# attendance/views.py
# attendance/views.py

# attendance/views.py

# attendance/views.py

# attendance/views.py

from .middleware import ThreadLocal
from django.utils import timezone
from django.core.exceptions import PermissionDenied

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from accounts.models import Profile
from .models import Attendance, Task
from .middleware import ThreadLocal

@login_required
def team_leader_dashboard(request, company_slug=None):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        # Profile હંમેશા default DB માંથી લાવવી
        user_profile = Profile.objects.using('default').select_related('company').get(user=request.user)
        company = user_profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name 
    except Exception:
        return HttpResponse("Profile or Company not found.")

    # 1. PERMISSION CHECK
    role_perms = None
    all_groups = request.user.groups.using('default').select_related('role_permissions').all()
    
    allowed_entry = False
    for group in all_groups:
        perms = getattr(group, 'role_permissions', None)
        if perms and perms.can_self_access:
            allowed_entry = True
            if perms.can_view_team:
                role_perms = perms
                break
            if not role_perms:
                role_perms = perms

    if not allowed_entry:
        raise PermissionDenied("You do not have permission to view this dashboard.")

    # ---------------------------------------------------------
    # 2. SQUAD LOGIC (Using Company DB context)
    # ---------------------------------------------------------
    today = timezone.localdate()
    current_month = today.month
    current_year = today.year

    # 🔥 FIX: Squad ફેચ કરતી વખતે user_id વાપરવી અને default DB માંથી profiles લેવી
    # કારણ કે TL અને Employee નું મેપિંગ મેઈન DB માં જ સચવાયેલું હોય છે
    my_squad = Profile.objects.using('default').filter(
        team_leader_id=request.user.id, 
        company=company
    ).select_related('user')
    
    squad_status = []
    present_count = 0
    total_tasks_pending = 0

    # Squad મેમ્બર્સના IDs ની લિસ્ટ (Optimized Query માટે)
    squad_user_ids = [member.user_id for member in my_squad]

    # લૂપની બહાર જ આજની એટેન્ડન્સ અને ટાસ્ક ડેટા લાવી દઈએ (Performance Fix)
    today_att_records = Attendance.objects.using(db_name).filter(
        user_id__in=squad_user_ids, 
        date=today
    )
    att_map = {a.user_id: a for a in today_att_records}

    for member in my_squad:
        # Today's Status (From map)
        att = att_map.get(member.user_id)
        is_present = att.check_in is not None if att else False
        if is_present:
            present_count += 1
            
        # Monthly Attendance Count (Company DB)
        monthly_days_present = Attendance.objects.using(db_name).filter(
            user_id=member.user_id, 
            date__month=current_month, 
            date__year=current_year,
            check_in__isnull=False
        ).count()

        # Task Management (Company DB)
        member_tasks_query = Task.objects.using(db_name).filter(assigned_to_id=member.user_id)
        member_tasks = member_tasks_query.order_by('-created_at')
        pending_tasks = member_tasks_query.filter(status__in=['Pending', 'In Progress']).count()
        
        total_tasks_pending += pending_tasks
        
        squad_status.append({
            'profile': member,
            'user': member.user, 
            'is_present': is_present,
            'all_tasks': member_tasks,
            'total_tasks_count': member_tasks.count(),
            'check_in': att.check_in if att else None,
            'pending_tasks': pending_tasks,
            'monthly_attendance': monthly_days_present, 
        })

    # Tasks assigned by TL (Company DB)
    all_recent_tasks = Task.objects.using(db_name).filter(assigned_by_id=request.user.id).order_by('-created_at')

    return render(request, 'attendance/dashboard/tl_dashboard.html', {
        'company': company,
        'profile': user_profile,
        'squad': squad_status,        
        'present_count': present_count,
        'total_squad': my_squad.count(),
        'recent_tasks': all_recent_tasks[:5],
        'tasks': all_recent_tasks,
        'role_perms': role_perms, 
        'total_tasks_pending': total_tasks_pending,
        'current_month': today.strftime('%B')
    })
# attendance/views.py
from django.utils import timezone
from .models import Task  # Task મોડેલ ઈમ્પોર્ટ હોવું જોઈએ

from .middleware import ThreadLocal
from django.db import transaction

@login_required
def tl_assign_task(request):
    """ટીમ મેમ્બરને સિંગલ ટાસ્ક અસાઇન કરવા માટે"""
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Profile Error")

    if request.method == "POST":
        member_id = request.POST.get('member_id')
        title = request.POST.get('title')
        description = request.POST.get('description')
        due_date = request.POST.get('due_date')
        priority = request.POST.get('priority')

        member_user = get_object_or_404(User, id=member_id)

        # ✅ ટાસ્ક બનાવો (In Company DB)
        Task.objects.using(db_name).create(
            title=title,
            description=description,
            assigned_to=member_user,
            assigned_by=request.user,
            due_date=due_date,
            priority=priority,
            status='Pending'
        )
        
        # ✅ નોટિફિકેશન બનાવો (In Company DB)
        Notification.objects.using(db_name).create(
            user=member_user, 
            message=f"New Task Assigned: {title}. Target date is {due_date}."
        )

        messages.success(request, "Task assigned successfully!")
        return redirect(request.META.get('HTTP_REFERER', 'tl_dashboard'))





@login_required
def tl_broadcast_task(request):
    """આખા સ્ક્વોડને એકસાથે ટાસ્ક બ્રોડકાસ્ટ કરવા માટે"""
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING
    # ---------------------------------------------------------
    db_name = f"{request.user.profile.company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    if request.method == "POST":
        member_ids = request.POST.getlist('member_ids') 
        title = request.POST.get('title')
        description = request.POST.get('description')
        due_date = request.POST.get('due_date')
        priority = request.POST.get('priority')

        if not member_ids:
            messages.error(request, "No team members were selected!")
            return redirect(request.META.get('HTTP_REFERER'))

        # ✅ ATOMIC TRANSACTION (કંપની ડેટાબેઝ માટે)
        with transaction.atomic(using=db_name):
            for m_id in member_ids:
                target_user = get_object_or_404(User, id=m_id)
                
                # ટાસ્ક બનાવો
                Task.objects.using(db_name).create(
                    title=title,
                    description=description,
                    assigned_to=target_user,
                    assigned_by=request.user,
                    due_date=due_date,
                    priority=priority,
                    status='Pending'
                )
                
                # નોટિફિકેશન મોકલો
                Notification.objects.using(db_name).create(
                    user=target_user,
                    message=f"New Team Task: {title}. Deadline: {due_date}."
                )

        messages.success(request, f"Task successfully Assigned to {len(member_ids)} members!")
        return redirect(request.META.get('HTTP_REFERER'))
    
from .middleware import ThreadLocal
from django.db import transaction

@login_required
def delete_team_mission(request):
    """ટીમ લીડર દ્વારા આખા મિશન (તમામ મેમ્બર્સના ટાસ્ક) ડિલીટ કરવા માટે"""
    if request.method == "POST":
        mission_title = request.POST.get('mission_title')
        
        # 🛡️ 0. DATABASE SWITCHING
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name

        # ૧. આ TL દ્વારા આપવામાં આવેલા તે Title વાળા તમામ ટાસ્ક શોધો (In Company DB)
        tasks_to_delete = Task.objects.using(db_name).filter(
            assigned_by=request.user, 
            title=mission_title
        )
        
        deleted_count = tasks_to_delete.count()
        tasks_to_delete.delete() # કંપની ડેટાબેઝમાંથી કાયમી ડિલીટ થશે
        
        messages.warning(request, f"Mission '{mission_title}' and tasks for {deleted_count} members deleted.")
        
    return redirect(request.META.get('HTTP_REFERER', 'tl_dashboard'))

@login_required
def delete_single_task(request, task_id):
    if request.method == "POST":
        # Get the specific company database
        db_name = f"{request.user.profile.company.slug}_db"
        
        try:
            # Explicitly query using the specific database
            task = Task.objects.using(db_name).get(id=task_id, assigned_by=request.user)
            task_title = task.title
            
            # Delete using the same database connection
            task.delete(using=db_name)
            
            messages.warning(request, f"Task '{task_title}' has been removed.")
        except Task.DoesNotExist:
            messages.error(request, "Task not found or you don't have permission.")
            
        return redirect(request.META.get('HTTP_REFERER', 'tl_dashboard'))
    
    return redirect('tl_dashboard')
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.models import User, Group
from accounts.models import Profile
from .models import RolePermission

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User, Group
from accounts.models import Profile

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.models import User, Group
from django.contrib import messages
from accounts.models import Profile

@login_required
def make_team_leader(request, user_id):
    """
    યુઝરને ટીમ લીડર બનાવવો અથવા જો તે પહેલેથી હોય તો તેનો રોલ રિવોક (Revoke) કરવો.
    સાથે જ જો રોલ જાય, તો તેની ટીમ પણ ઓટોમેટિક ફ્રી થઈ જશે.
    """
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
    except Exception:
        return HttpResponse("Profile Error: તમારી પ્રોફાઇલ અથવા કંપનીની વિગત મળી નથી.")

    # ૧. ટાર્ગેટ યુઝરને મેળવો (Main DB)
    target_user = get_object_or_404(User.objects.using('default'), id=user_id)
    slug_lower = company.slug.lower()
    tl_group_name = f"{slug_lower}_team leader"

    # ૨. રોલ અસાઇનમેન્ટ લોજિક
    tl_group, _ = Group.objects.using('default').get_or_create(name=tl_group_name)
    
    if target_user.groups.filter(id=tl_group.id).exists():
        # 🔥 REVOKE LOGIC: જો તે પહેલેથી TL હોય, તો રોલ હટાવો
        target_user.groups.remove(tl_group)
        
        # 🧹 CASCADE CLEANUP: તેની અંડરના તમામ મેમ્બર્સને ફ્રી કરો
        Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader_id=None)
        Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader_id=None)
        
        messages.warning(request, f"{target_user.username} નો TL રોલ રદ કર્યો છે અને તેની ટીમ ફ્રી થઈ ગઈ છે.")
    else:
        # 🔥 MAKE TL: જો તે TL ના હોય, તો રોલ આપો
        target_user.groups.add(tl_group)
        
        # પ્રોફાઇલ અપડેટ (બંને DB માં)
        Profile.objects.using('default').filter(user_id=user_id).update(is_approved=True)
        Profile.objects.using(db_name).filter(user_id=user_id).update(is_approved=True)
        
        messages.success(request, f"અભિનંદન! {target_user.username} હવે ટીમ લીડર છે.")

    return redirect(request.META.get('HTTP_REFERER', 'admin_manage_roles'))
    

# attendance/views.py

from .middleware import ThreadLocal
from django.db import transaction

@login_required
def assign_team_to_leader(request):
    """ટીમ લીડરને સ્ક્વોડ મેમ્બર્સ અસાઇન કરવા માટે (Manager Level)"""
    
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        manager_profile = request.user.profile
        company = manager_profile.company
        manager_branch = manager_profile.branch 
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name # 👈 બધી ક્વેરી હવે આ DB માં જશે
    except:
        messages.error(request, "Profile Error")
        return redirect('manager_dashboard')

    if request.method == "POST":
        leader_id = request.POST.get('leader_id')
        employee_ids = request.POST.getlist('employee_ids')

        if not leader_id:
            messages.error(request, "Error: Team Leader ID missing.")
            return redirect('manager_dashboard')

        try:
            leader_user = get_object_or_404(User, id=leader_id)
            
            # 🛡️ 1. ATOMIC SYNC LOGIC (કંપની ડેટાબેઝમાં)
            # 
            
            with transaction.atomic(using=db_name):
                # A. જૂના મેમ્બર્સને ક્લિયર કરો (In Company DB)
                Profile.objects.using(db_name).filter(team_leader=leader_user).update(team_leader=None)

                # B. નવા સિલેક્ટ થયેલા સભ્યોને અસાઇન કરો
                count = 0
                if employee_ids:
                    # સિક્યોરિટી સાથે બલ્ક અપડેટ (Performance Optimization)
                    profiles_to_update = Profile.objects.using(db_name).filter(
                        user__id__in=employee_ids,
                        manager=request.user,
                        company=company
                    )
                    
                    if manager_branch:
                        profiles_to_update = profiles_to_update.filter(branch=manager_branch)
                    
                    count = profiles_to_update.update(team_leader=leader_user)

            messages.success(request, f"Squad updated! {count} members assigned to {leader_user.username}.")
            
            # C. એમ્પ્લોઈઝને નોટિફિકેશન મોકલો (Company DB)
            create_notification(leader_user, f"Your squad has been updated with {count} members.")
            
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            
        return redirect('manager_dashboard')
        
    return redirect('manager_dashboard')

# attendance/views.py
def ajax_check_username(request):
    username = request.GET.get('username', '').strip()
    company_slug = request.GET.get('company_slug', None)

    if not username:
        return JsonResponse({'is_taken': False})

    # ૧. આખા ડેટાબેઝમાં યુઝર શોધો (Case Insensitive)
    user_obj = User.objects.filter(
        Q(username__iexact=username) | Q(email__iexact=username)
    ).first()

    if user_obj:
        # ૨. જો તે સુપરએડમિન હોય -> ૧૦૦% ગ્રીન બોર્ડર
        if user_obj.is_superuser:
            return JsonResponse({'is_taken': True})

        # ૩. જો સામાન્ય યુઝર હોય તો કંપની ચેક કરો
        try:
            if company_slug and user_obj.profile.company.slug == company_slug:
                return JsonResponse({'is_taken': True})
        except Exception:
            pass

    return JsonResponse({'is_taken': False})

# attendance/views.py

from .middleware import ThreadLocal
from django.db.models import Max
from django.utils import timezone

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.db.models import Max, Count
from accounts.models import Profile
from .models import Attendance, Task
from .middleware import ThreadLocal

from django.db.models import Count, Max, Q
from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
# Make sure your other imports (ThreadLocal, Task, Attendance, Profile) are here

@login_required
def tl_dashboard_view(request, company_slug=None):
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        profile = request.user.profile
        company = profile.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name 
    except Exception:
        return HttpResponse("Profile Error: Company database not found.")

    # 1. SECURITY CHECK
    if not request.user.groups.filter(name__icontains="team leader").exists():
        return HttpResponse("Access Denied: Team Leader Access Only.")

    # 🛡️ 2. PERMISSION LOGIC
    all_groups = request.user.groups.all()
    role_perms = None
    has_access = False
    for group in all_groups:
        perms = getattr(group, 'role_permissions', None)
        if perms and perms.can_self_access:
            has_access = True
            if perms.can_view_team:
                role_perms = perms
                break
            if not role_perms: role_perms = perms

    if not has_access:
        return HttpResponse("Access Denied: Restricted by Admin.")

    # 3. BASIC DATA SETUP
    today = timezone.localdate()
    month_start = today.replace(day=1)
    
    # ---------------------------------------------------------
    # 🔥 4. UNIQUE TASK LOGIC (Fix for Orange Card Count)
    # ---------------------------------------------------------
    # અહીં આપણે એવા પેન્ડિંગ ટાસ્ક શોધીએ છીએ જે આ TL એ આપ્યા હોય
    all_pending_query = Task.objects.using(db_name).filter(
        assigned_by_id=request.user.id, 
        status__in=['Pending', 'In Progress']
    )
    
    # MASTER FIX: ટાઇટલ મુજબ ડિસ્ટિંક્ટ કાઉન્ટ (૩ જણાને ૧ ટાસ્ક = ૧ કાઉન્ટ)
    total_tasks_pending = all_pending_query.values('title').distinct().count()

    # 👥 5. SQUAD LOGIC (Using Company DB)
    my_squad = Profile.objects.using('default').filter(
        team_leader_id=request.user.id, 
        company=company
    ).select_related('user')
    
    squad_data = []
    present_today_count = 0
    
    for member in my_squad:
        # Today's Status
        att_today = Attendance.objects.using(db_name).filter(user_id=member.user_id, date=today).first()
        is_present = att_today.check_in is not None if att_today else False
        if is_present: present_today_count += 1
            
        # Base query for member tasks
        member_tasks_qs = Task.objects.using(db_name).filter(
            assigned_to_id=member.user_id, 
            assigned_by_id=request.user.id
        ).order_by('-created_at')
        
        # Calculate counts BEFORE converting to a list
        total_count = member_tasks_qs.count()
        pending_count = member_tasks_qs.filter(status__in=['Pending', 'In Progress']).count()

        # 🔥 CRITICAL FIX: Prefetch attachments AND force memory evaluation with list()
        evaluated_tasks = list(member_tasks_qs.prefetch_related('attachments'))

        # Monthly Attendance
        monthly_presents = Attendance.objects.using(db_name).filter(
            user_id=member.user_id, 
            date__gte=month_start, 
            status__in=['Present', 'Late', 'Half Day']
        ).count()

        squad_data.append({
            'profile': member,
            'user': member.user, 
            'is_present': is_present,
            'check_in': att_today.check_in if att_today else None,
            'monthly_attendance': monthly_presents,
            'pending_tasks': pending_count,
            'total_tasks_count': total_count,
            'all_tasks': evaluated_tasks,  # <--- Now passing the pre-loaded memory list!
        })

    # ---------------------------------------------------------
    # 🚀 6. TEAM MISSION TRACKER LOGIC (GROUP TASKS ONLY)
    # ---------------------------------------------------------
    all_assigned_tasks = Task.objects.using(db_name).filter(assigned_by_id=request.user.id)
    unique_mission_titles = all_assigned_tasks.values('title').distinct()
    team_tasks_summary = []

    for mission in unique_mission_titles:
        title = mission['title']
        group_tasks = all_assigned_tasks.filter(title=title)
        
        # ✅ THE MASTER FIX: જો ટાસ્ક ૧ થી વધુ મેમ્બરને આપ્યો હોય તો જ ટ્રેકરમાં બતાવો
        if group_tasks.count() > 1:
            members_with_status = []
            for t in group_tasks:
                members_with_status.append({
                    'name': t.assigned_to.username.title(),
                    'is_verified': t.status == 'Verified'
                })
            
            team_tasks_summary.append({
                'title': title,
                'total_count': group_tasks.count(),
                'verified_count': group_tasks.filter(status='Verified').count(),
                'task_description': group_tasks.first().description,
                'latest_due_date': group_tasks.aggregate(Max('due_date'))['due_date__max'],
                'task_priority': group_tasks.first().priority,
                'assigned_members': members_with_status
            })

    unique_active_missions_count = len(team_tasks_summary)
    
    # Wrap recent tasks in list() to ensure they don't lazily query later
    recent_tasks = list(all_assigned_tasks.order_by('-created_at')[:5])

    # Safety reset ThreadLocal for next request
    ThreadLocal.DB_NAME = 'default'

    context = {
        "profile": profile,
        "company": company,
        "squad": squad_data,
        "team_tasks": team_tasks_summary,
        "total_squad": my_squad.count(),
        "present_count": present_today_count,
        "total_tasks_pending_unique": unique_active_missions_count,
        "total_tasks_pending": total_tasks_pending,
        "recent_tasks": recent_tasks,
        "current_month": today.strftime('%B'),
        "role_perms": role_perms,
    }
    return render(request, "attendance/dashboard/tl_dashboard.html", context)    
        
from .middleware import ThreadLocal
from django.utils import timezone

@login_required
def tl_task_action(request, task_id):
    """ટીમ લીડર દ્વારા ટાસ્ક અપ્રૂવ કે રિજેક્ટ કરવા માટે"""
    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        db_name = f"{request.user.profile.company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Profile/Company not found")

    # ૧. સુરક્ષા: ટાસ્ક મેળવો (Company DB mathi)
    task = get_object_or_404(Task.objects.using(db_name), id=task_id, assigned_by=request.user)
    
    if request.method == "POST":
        action = request.POST.get('action')
        
        if action == 'approve':
            task.status = 'Verified'
            task.rejection_note = None
            task.save(using=db_name) # ✅ Save in Company DB
            
            # 🔥 એમ્પ્લોઈને નોટિફિકેશન મોકલો (Company DB)
            Notification.objects.using(db_name).create(
                user=task.assigned_to, 
                message=f"Task Accomplished! Your task '{task.title}' has been Verified by TL."
            )
            messages.success(request, "Task Verified Successfully!")
        
        elif action == 'reject':
            reason = request.POST.get('rejection_note')
            task.status = 'Rejected'
            task.rejection_note = reason
            task.rejected_at = timezone.now()
            task.save(using=db_name) # ✅ Save in Company DB
            
            # 🔥 એમ્પ્લોઈને રિજેક્શન નોટિફિકેશન (Company DB)
            Notification.objects.using(db_name).create(
                user=task.assigned_to, 
                message=f"Action Required: Your Task '{task.title}' was Rejected. Reason: {reason}"
            )
            messages.warning(request, "Task Rejected and Employee notified.")
        
        return redirect(request.META.get('HTTP_REFERER', 'tl_dashboard'))

    return redirect('tl_dashboard')



from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from accounts.models import Profile
from attendance.models import Branch # 👈 તમારા બ્રાન્ચ મોડલનું સાચું લોકેશન ચેક કરજો
from django.contrib.auth.decorators import login_required

@login_required
def hr_assign_member(request):
    if request.method == "POST":
        tl_id = request.POST.get('tl_id')
        member_id = request.POST.get('member_id')
        
        try:
            hr_profile = Profile.objects.using('default').get(user=request.user)
            db_name = f"{hr_profile.company.slug}_db"
            
            # ૧. મેઈન DB માંથી બંનેની પ્રોફાઈલ મેળવો (ફક્ત ડેટા ચેક કરવા)
            tl_prof = Profile.objects.using('default').get(user_id=tl_id)
            mem_prof = Profile.objects.using('default').get(user_id=member_id)

            # ૨. અપડેટ કરવા માટેનો ડેટા (બ્રાન્ચ અહીં નથી, એટલે તે બદલાશે નહીં)
            update_vals = {
                'team_leader_id': tl_id,
                'manager_id': tl_prof.manager_id, # TL જે મેનેજરનો છે તે જ મેનેજર મેમ્બરને મળશે
                'is_approved': True
            }

            # ✅ FORCE UPDATE IN BOTH DATABASES
            Profile.objects.using(db_name).filter(user_id=member_id).update(**update_vals)
            Profile.objects.using('default').filter(user_id=member_id).update(**update_vals)

            messages.success(request, "સભ્યને ટીમમાં ઉમેરવામાં આવ્યો છે. મેનેજર અને બ્રાન્ચ સુરક્ષિત છે.")
        except Exception as e:
            messages.error(request, f"ભૂલ: {str(e)}")
            
    return redirect('hr_tl_list')


    # attendance/views.py

from .middleware import ThreadLocal

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import Group, User
from accounts.models import Profile
from attendance.models import RolePermission

@login_required
def hr_team_leader_list(request):
    # ---------------------------------------------------------
    # 0. PROFILE & DB SETUP
    # ---------------------------------------------------------
    try:
        # બધું જ 'default' DB માંથી લાવો જેથી JOIN ક્યારેય ફેલ ના થાય
        hr_profile = Profile.objects.using('default').select_related('company', 'branch').get(user=request.user)
        company = hr_profile.company
        hr_branch = hr_profile.branch # આ બ્રાન્ચ ઓબ્જેક્ટ અથવા None હોઈ શકે
        db_name = f"{company.slug}_db"
    except Profile.DoesNotExist:
        return HttpResponse("Profile Error: Profile not found.")

    # 1. SECURITY CHECK
    if not (request.user.groups.filter(name__icontains='hr').exists() or request.user.is_superuser):
        return HttpResponse("Access Denied")

    # 🔥 2. PROXY HACK (HTML બટનો માટે)
    tenant_perms = None
    real_group = request.user.groups.first()
    if real_group:
        try:
            tenant_perms = RolePermission.objects.using(db_name).filter(group_id=real_group.id).first()
        except: pass

    class ProxyGroup:
        def __init__(self, obj): self._obj = obj; self.role_permissions = tenant_perms
        def __getattr__(self, name): return getattr(self._obj, name)
    class ProxyManager:
        def __init__(self, m): self._m = m
        def first(self): obj = self._m.first(); return ProxyGroup(obj) if obj else None
        def __getattr__(self, name): return getattr(self._m, name)
    class ProxyUser:
        def __init__(self, u): self._u = u; self.groups = ProxyManager(u.groups); self.id = u.id
        def __getattr__(self, name): return getattr(self._u, name)
    class ProxyRequest:
        def __init__(self, r): self._r = r; self.user = ProxyUser(r.user)
        def __getattr__(self, name): return getattr(self._r, name)

    # ---------------------------------------------------------
    # 📋 3. FETCH TEAM LEADERS (Branch Specific Filtering)
    # ---------------------------------------------------------
    slug_lower = company.slug.lower()
    tl_group_name = f"{slug_lower}_team leader"
    
    # 🔥 તમારી મેઈન કન્ડિશન: HR ની બ્રાન્ચ મુજબ જ TLs ફિલ્ટર થશે
    tls_base_qs = Profile.objects.using('default').filter(
        company=company,
        branch=hr_branch, # જો hr_branch=None હશે તો Head Office ના જ લોકો આવશે
        user__groups__name__iexact=tl_group_name
    ).select_related('user', 'manager', 'branch')

    tls_profiles = list(tls_base_qs)

    # ---------------------------------------------------------
    # 🔍 4. ELIGIBLE EMPLOYEES POOL (Branch Specific)
    # ---------------------------------------------------------
    # સ્ટાફ પણ એ જ બ્રાન્ચનો હોવો જોઈએ જેનો HR છે
    staff_base_qs = Profile.objects.using('default').filter(
        company=company,
        branch=hr_branch # 👈 સ્ટાફ ફિલ્ટરિંગમાં પણ બ્રાન્ચ કન્ડિશન
    ).exclude(
        user__groups__name__iregex=fr'({slug_lower}_admin|{slug_lower}_hr|{slug_lower}_manager|{slug_lower}_team leader|{slug_lower}_owner)'
    )

    # હવે કંપનીના ડીબી (Tenant DB) માં ચેક કરો કે જેમને હજુ TL મળ્યો નથી
    all_eligible_ids = list(staff_base_qs.values_list('user_id', flat=True))
    
    free_staff_ids = list(Profile.objects.using(db_name).filter(
        user_id__in=all_eligible_ids,
        team_leader__isnull=True
    ).values_list('user_id', flat=True))

    # ફાઇનલ સ્ટાફ પૂલ
    eligible_staff_pool = list(Profile.objects.using('default').filter(
        user_id__in=free_staff_ids
    ).select_related('user', 'manager'))

    # ---------------------------------------------------------
    # 📦 5. DATA PACKING (Strict Manager Alignment Fix)
    # ---------------------------------------------------------
    tl_data = []
    for tl in tls_profiles:
        # A. Current Squad (જે ઓલરેડી આ TL નીચે છે)
        squad_members = Profile.objects.using('default').filter(
            team_leader_id=tl.user_id,
            company=company,
            branch=hr_branch # 👈 સ્ક્વોડ લિસ્ટમાં પણ બ્રાન્ચ મેચ હોવી જોઈએ
        ).select_related('user', 'branch')
        
        # B. MANAGER ALIGNMENT
        current_tl_available = [
            staff for staff in eligible_staff_pool 
            if staff.manager_id == tl.manager_id and staff.user_id != tl.user_id
        ]

        tl_data.append({
            'user': tl.user,
            'profile': tl,
            'manager': tl.manager,
            'squad_count': squad_members.count(),
            'squad_list': squad_members,
            'available_staff': current_tl_available,
            'id': tl.user_id
        })

    # છેલ્લે ThreadLocal રીસેટ (Safety)
    from .middleware import ThreadLocal
    ThreadLocal.DB_NAME = 'default'

    return render(request, "attendance/hr/hr_tl_list.html", {
        "request": ProxyRequest(request),
        "team_leaders": tl_data,
        "company": company,
        "current_branch": hr_branch
    })    
            # attendance/views.py
@login_required
def hr_remove_member(request):
    """ટીમ મેમ્બરને સ્ક્વોડમાંથી દૂર કરવા પણ બ્રાન્ચ અને મેનેજર સાચવી રાખવા માટે"""
    if request.method == "POST":
        member_id = request.POST.get('member_id')
        
        try:
            # ૧. HR ની પ્રોફાઇલ પરથી DB નામ મેળવો
            hr_profile = Profile.objects.using('default').get(user=request.user)
            db_name = f"{hr_profile.company.slug}_db"
            
            # ---------------------------------------------------------
            # 🔥 THE CRITICAL FIX: Only set team_leader to None
            # ---------------------------------------------------------
            # આપણે મેનેજર (manager_id) ને None નથી કરતા કારણ કે મેમ્બર હજુ પણ 
            # એ જ બ્રાન્ચ અને મેનેજરનો ભાગ છે, ફક્ત તે TL નીચેથી નીકળે છે.
            
            update_data = {
                'team_leader_id': None,
                # 'manager_id': None  <-- આ લાઇન કાઢી નાખી છે જેથી મેનેજર સચવાય
            }

            # A. મેઈન ડેટાબેઝ (default) માં અપડેટ (બ્રાન્ચને અડ્યા વગર)
            Profile.objects.using('default').filter(user_id=member_id).update(**update_data)

            # B. ટેનન્ટ ડેટાબેઝ (Company DB) માં અપડેટ
            Profile.objects.using(db_name).filter(user_id=member_id).update(**update_data)

            messages.warning(request, "સભ્યને સ્ક્વોડમાંથી મુક્ત કર્યો છે. બ્રાન્ચ અને મેનેજર યથાવત છે.")
            
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            
    return redirect('hr_tl_list')
    
from .middleware import ThreadLocal
from django.db import transaction

@login_required
def update_weekoff_policy(request, company_slug):
    """કંપનીની વીક-ઓફ પોલિસી અપડેટ કરી કંપની ડેટાબેઝમાં સિંક કરો"""
    
    # 1. Company મેળવો
    company = get_object_or_404(Company, slug=company_slug)
    
    # 🛡️ SECURITY CHECK
    if not is_owner_or_superuser(request.user, company):
        messages.error(request, "Access Denied: You don't have permission to change policy.")
        return redirect('dashboard')

    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    db_name = f"{company.slug}_db"
    ThreadLocal.DB_NAME = db_name

    if request.method == "POST":
        # HTML માંથી સિલેક્ટ કરેલા દિવસો મેળવો
        selected_days = request.POST.getlist('week_off')
        week_off_string = ",".join(selected_days)

        # 🛡️ ATOMIC SYNC: Main DB ane Company DB banne ma policy update karo
        with transaction.atomic():
            # A. Company Database context ma save karo
            policy_local, _ = AttendancePolicy.objects.using(db_name).get_or_create(company=company)
            policy_local.week_off_days = week_off_string
            policy_local.save(using=db_name)

            # B. Default Database ma pan redundancy mate save kari shako
            policy_main, _ = AttendancePolicy.objects.get_or_create(company=company)
            policy_main.week_off_days = week_off_string
            policy_main.save()

        messages.success(request, f"Week-off policy for {company.name} updated successfully!")

        # ---------------------------------------------------------
        # DYNAMIC REDIRECTION (Tamaru logic perfect che)
        # ---------------------------------------------------------
        if request.user.is_superuser:
            return redirect('superadmin_company_dashboard', company_id=company.id)
        
        owner_group_name = f"{company.slug}_owner"
        if request.user.groups.filter(name=owner_group_name).exists():
            return redirect('company_owner_dashboard', company_slug=company.slug)
        
        return redirect('dashboard')

    return redirect('dashboard')    
    # return redirect('superadmin_company_dashboard', company_id=company.id)


def is_owner_or_superuser(user, company):
    """ચેક કરે છે કે યુઝર સુપરયુઝર છે અથવા તે કંપનીનો ઓનર છે."""
    if user.is_superuser:
        return True
    owner_group_name = f"{company.slug}_owner"
    return user.groups.filter(name=owner_group_name).exists()


# --- HELPER FUNCTION FOR DYNAMIC REDIRECT ---
def get_dashboard_redirect(user, company):
    """કંપની ઓનર અને સુપર એડમિન મુજબ સાચો રીડાયરેક્ટ પાથ આપે છે."""
    if user.is_superuser:
        return redirect('superadmin_company_dashboard', company_id=company.id)
    
    owner_group_name = f"{company.slug}_owner"
    if user.groups.filter(name=owner_group_name).exists():
        return redirect('company_owner_dashboard', company_slug=company.slug)
    
    return redirect('dashboard')



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User, Group
from .models import Company, Branch # Branch model import karvanu bhulta nai
from accounts.models import Profile

# Company Owner Dashboard View (Full Updated)
from .middleware import ThreadLocal
from django.db import transaction

# attendance/views.py

from django.db import transaction

from django.db import transaction

# @login_required
# def company_owner_dashboard(request, company_slug):
#     # 🛡️ 0. DATABASE & SECURITY
#     from .middleware import ThreadLocal
#     company = get_object_or_404(Company.objects.using('default'), slug=company_slug)
#     db_name = f"{company.slug}_db"
    
#     from attendance.utils import ensure_db_connection
#     ensure_db_connection(db_name)
#     ThreadLocal.DB_NAME = db_name

#     owner_group_name = f"{company.slug}_owner"
#     if not (request.user.is_superuser or request.user.groups.filter(name=owner_group_name).exists()):
#         ThreadLocal.DB_NAME = 'default'
#         return HttpResponseForbidden("Access Denied.")

#     # 🛡️ 1. POST HANDLING
#     if request.method == "POST":
        
#         # ----------------------------------------------------
#         # ACTION A: CREATE NEW BRANCH 
#         # ----------------------------------------------------
#         if 'create_branch' in request.POST:
#             branch_name = request.POST.get('branch_name')
#             location = request.POST.get('location')
#             branch_admin_id = request.POST.get('branch_admin_id') 
            
#             if branch_name:
#                 branch_main = None
#                 branch_tenant = None
#                 try:
#                     with transaction.atomic(using='default'):
#                         with transaction.atomic(using=db_name):
                            
#                             branch_main = Branch.objects.using('default').create(
#                                 company_id=company.id, name=branch_name, location=location
#                             )
#                             branch_tenant = Branch.objects.using(db_name).create(
#                                 company_id=company.id, name=branch_name, location=location
#                             )
                            
#                             msg = f"✅ Branch '{branch_name}' created successfully!"

#                             if branch_admin_id:
#                                 # 🔥 FIX: 'role' કાઢી નાખ્યું છે. 
#                                 Profile.objects.using('default').filter(user_id=branch_admin_id).update(
#                                     branch_id=branch_main.id, manager=None, team_leader=None
#                                 )
#                                 Profile.objects.using(db_name).filter(user_id=branch_admin_id).update(
#                                     branch_id=branch_tenant.id, manager=None, team_leader=None
#                                 )

#                                 # તેના જુના સ્ટાફને ફ્રી કરો (Cleanup)
#                                 Profile.objects.using('default').filter(manager_id=branch_admin_id).update(manager=None)
#                                 Profile.objects.using(db_name).filter(manager_id=branch_admin_id).update(manager=None)
#                                 Profile.objects.using('default').filter(team_leader_id=branch_admin_id).update(team_leader=None)
#                                 Profile.objects.using(db_name).filter(team_leader_id=branch_admin_id).update(team_leader=None)

#                                 # રોલ અપડેટ (Main DB) - સાચું રોલ મેનેજમેન્ટ અહીં થાય છે!
#                                 user = User.objects.using('default').get(id=branch_admin_id)
#                                 user.groups.clear()
#                                 admin_group_name = f"{company.slug.lower()}_admin"
#                                 admin_group, _ = Group.objects.using('default').get_or_create(name=admin_group_name)
#                                 user.groups.add(admin_group)

#                                 msg += f" And {user.username} is now the Branch Admin."
                                
#                     messages.success(request, msg)
                    
#                 except Exception as e:
#                     messages.error(request, f"❌ Error creating branch: {str(e)}")
                    
#             return redirect('company_owner_dashboard', company_slug=company.slug)

#         # ----------------------------------------------------
#         # ACTION B: ASSIGN STAFF ROLE
#         # ----------------------------------------------------
#         # ----------------------------------------------------
#         # ACTION B: ASSIGN STAFF ROLE (🔥 AUTO-SYNC ID FIX)
#         # ----------------------------------------------------
#         elif 'assign_staff_role' in request.POST:
#             user_id = request.POST.get('user_id')
#             tenant_branch_id = request.POST.get('branch_id')
#             role_name = request.POST.get('role_name', '').strip().lower() 

#             try:
#                 with transaction.atomic(using='default'):
#                     with transaction.atomic(using=db_name):
#                         target_user = User.objects.using('default').get(id=user_id)
                        
#                         # ૧. ટેનન્ટ ડીબીમાંથી બ્રાન્ચ મેળવો (જેનું ID ફોર્મમાંથી આવ્યું છે)
#                         branch_tenant = Branch.objects.using(db_name).get(id=tenant_branch_id)

#                         # ૨. મેઈન ડીબીમાં તે જ નામની બ્રાન્ચ શોધો અથવા ના હોય તો બનાવો!
#                         # આનાથી ID મિસમેચ નો પ્રોબ્લેમ કાયમ માટે ગાયબ થઈ જશે
#                         branch_main, _ = Branch.objects.using('default').get_or_create(
#                             company=company, 
#                             name=branch_tenant.name,
#                             defaults={'location': branch_tenant.location}
#                         )

#                         # ૩. હવે બંને ડીબી માટે પોત-પોતાના સાચા IDs વાપરો
#                         Profile.objects.using('default').filter(user_id=user_id).update(
#                             branch_id=branch_main.id, manager=None, team_leader=None
#                         )
#                         Profile.objects.using(db_name).filter(user_id=user_id).update(
#                             branch_id=branch_tenant.id, manager=None, team_leader=None
#                         )

#                         # ૪. Clean up subordinate links (આ યુઝર નીચે કામ કરતા લોકોને ફ્રી કરો)
#                         Profile.objects.using('default').filter(manager_id=user_id).update(manager=None)
#                         Profile.objects.using(db_name).filter(manager_id=user_id).update(manager=None)
#                         Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader=None)
#                         Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader=None)

#                         # ૫. Group assignment (રોલ એસાઈનમેન્ટ - જે હમેશા Main DB માં થાય છે)
#                         target_user.groups.clear()
#                         if role_name:
#                             group_name = f"{company.slug.lower()}_{role_name}"
#                             target_group, _ = Group.objects.using('default').get_or_create(name=group_name)
#                             target_user.groups.add(target_group)

#                 messages.success(request, f"✅ Role updated successfully. {target_user.username} is now {role_name.title()}.")
#             except Exception as e:
#                 messages.error(request, f"Error assigning role: {str(e)}")
            
#             return redirect('company_owner_dashboard', company_slug=company.slug)
#     # 🛡️ 2. GET REQUEST DATA 
#     # ----------------------------------------------------
#     # 🛡️ 2. GET REQUEST DATA & 100% ACCURATE STATS
#     # ----------------------------------------------------
#     branches = Branch.objects.using(db_name).filter(company=company)
#     selected_role = request.GET.get('role', 'all').lower()
    
#     # 1. Get ALL user IDs that belong to this company from Tenant DB
#     tenant_profile_user_ids = list(Profile.objects.using(db_name).filter(
#         company_id=company.id
#     ).values_list('user_id', flat=True))

#     # 2. Fetch those users from Main DB to check their exact groups
#     company_users = User.objects.using('default').filter(
#         id__in=tenant_profile_user_ids
#     ).prefetch_related('groups')

#     # 3. Categorize Each User in Python (NO CROSS-DB JOINS!)
#     stats = {
#         'total': 0, 'admin': 0, 'hr': 0, 'manager': 0, 'team_leader': 0, 'employee': 0
#     }
#     role_map = {} # Dictionary to store: user_id -> role_name
#     owner_group_name = f"{company.slug}_owner".lower()
#     slug_lower = company.slug.lower()

#     for u in company_users:
#         user_groups = [g.name.lower() for g in u.groups.all()]
        
#         # A. Check if owner (અમે માલિકને સ્ટાફની ગણતરીમાંથી કાઢી નાખીએ છીએ)
#         if any(g == owner_group_name for g in user_groups):
#             role_map[u.id] = 'owner'
#             continue
            
#         # B. Determine highest role for THIS company
#         emp_role = 'employee' # Default role
#         if any(f"{slug_lower}_admin" in g for g in user_groups):
#             emp_role = 'admin'
#         elif any(f"{slug_lower}_hr" in g for g in user_groups):
#             emp_role = 'hr'
#         elif any(f"{slug_lower}_manager" in g for g in user_groups):
#             emp_role = 'manager'
#         elif any(f"{slug_lower}_team leader" in g for g in user_groups):
#             emp_role = 'team leader'
            
#         # નકશો સેવ કરો
#         role_map[u.id] = emp_role
        
#         # C. Increment stats
#         if emp_role == 'admin': stats['admin'] += 1
#         elif emp_role == 'hr': stats['hr'] += 1
#         elif emp_role == 'manager': stats['manager'] += 1
#         elif emp_role == 'team leader': stats['team_leader'] += 1
#         elif emp_role == 'employee': stats['employee'] += 1
        
#         stats['total'] += 1

#     # 4. Filter the Table Data based on the Role Map
#     if selected_role != 'all':
#         # જો UI માંથી કોઈ ખાસ રોલ પસંદ કર્યો હોય (e.g., HR)
#         filtered_ids = [uid for uid, role in role_map.items() if role == selected_role]
#     else:
#         # 'All' માં ઓનર સિવાયના બધા બતાવો
#         filtered_ids = [uid for uid, role in role_map.items() if role != 'owner']

#     # 5. ફાઇનલ ડેટાબેઝ ક્વેરી (ફક્ત વેલિડ IDs સાથે)
#     employees_qs = Profile.objects.using(db_name).filter(
#         user_id__in=filtered_ids
#     ).select_related('user', 'branch').order_by('-user__date_joined')

#     # 6. Prepare Final List for Template
#     processed_staff = []
#     for emp in employees_qs:
#         # પેલા બનાવેલા નકશા (role_map) માંથી જ રોલનું નામ આપો 
#         # જેથી ટેબલ અને કાર્ડના આંકડા ક્યારેય મિસમેચ ના થાય!
#         emp_role = role_map.get(emp.user_id, 'employee')
#         emp.display_role = emp_role.title()
#         processed_staff.append(emp)

#     policies = AttendancePolicy.objects.using(db_name).filter(company=company)
#     holidays = Holiday.objects.using(db_name).filter(company=company).order_by('date')
#     shifts = Shift.objects.using(db_name).filter(company=company)

#     return render(request, 'attendance/dashboard/company_owner_dashboard.html', {
#         'company': company,
#         'branches': branches,
#         'employees': processed_staff,
#         'stats': stats, # ✅ The absolute source of truth
#         'policies': policies,
#         'holidays': holidays,
#         'shifts': shifts,
#         'is_owner': True,
#         'role_perms': {'can_view_team': True, 'can_manage_users': True, 'can_view_reports': True}
#     })    
from .middleware import ThreadLocal

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .middleware import ThreadLocal
from .models import Branch
from accounts.models import Profile
from django.contrib.auth.models import User

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .middleware import ThreadLocal
from .models import Branch
from accounts.models import Profile
from django.contrib.auth.models import User

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .middleware import ThreadLocal
from .models import Branch
from accounts.models import Profile
from django.contrib.auth.models import User

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from .models import Branch, Profile
from django.contrib.auth.models import User
from .middleware import ThreadLocal

@login_required
def branch_detail_view(request, branch_id):
    """View to see specific branch details and staff list (Fixed for Superadmin)"""

    # ---------------------------------------------------------
    # 0. DATABASE SWITCHING (Level 0 Logic)
    # ---------------------------------------------------------
    try:
        # 🔥 SUPERADMIN FIX: Superadmins don't have a specific company profile.
        # So, we fetch the company directly using the branch_id from the Main DB.
        if request.user.is_superuser:
            temp_branch = Branch.objects.using('default').filter(id=branch_id).first()
            if not temp_branch:
                messages.error(request, "Branch not found in Main Database.")
                return redirect(request.META.get('HTTP_REFERER', '/'))
            company = temp_branch.company
        else:
            # For regular users (Owner, HR, Manager)
            company = request.user.profile.company

        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except Exception as e:
        print(f"Error in branch_detail_view: {e}") # Prints to terminal for easy debugging
        return HttpResponse(f"Profile/Company Error: {e}")

    # 1. Get the Branch (From Company DB)
    branch = get_object_or_404(Branch.objects.using(db_name), id=branch_id)

    # 2. Get role from URL
    selected_role = request.GET.get('role', 'all').lower()

    # ---------------------------------------------------------
    # 🛡️ 3. ACCURATE STATS & DATA FETCHING
    # ---------------------------------------------------------
    branch_profile_user_ids = list(Profile.objects.using(db_name).filter(
        branch_id=branch.id,
        company_id=company.id
    ).values_list('user_id', flat=True))

    branch_users = User.objects.using('default').filter(
        id__in=branch_profile_user_ids
    ).prefetch_related('groups')

    stats = {
        'total': 0, 'admin': 0, 'hr': 0, 'manager': 0, 'team_leader': 0, 'employee': 0
    }

    role_map = {}
    slug_lower = company.slug.lower()
    owner_group_name = f"{slug_lower}_owner"

    for u in branch_users:
        user_groups = [g.name.lower() for g in u.groups.all()]

        # Skip the owner
        if any(g == owner_group_name for g in user_groups):
            role_map[u.id] = 'owner'
            continue

        emp_role = 'employee' # Default
        if any(f"{slug_lower}_admin" in g for g in user_groups): emp_role = 'admin'
        elif any(f"{slug_lower}_hr" in g for g in user_groups): emp_role = 'hr'
        elif any(f"{slug_lower}_manager" in g for g in user_groups): emp_role = 'manager'
        elif any(f"{slug_lower}_team leader" in g for g in user_groups): emp_role = 'team_leader'

        role_map[u.id] = emp_role

        # Increment stats
        if emp_role == 'admin': stats['admin'] += 1
        elif emp_role == 'hr': stats['hr'] += 1
        elif emp_role == 'manager': stats['manager'] += 1
        elif emp_role == 'team_leader': stats['team_leader'] += 1
        elif emp_role == 'employee': stats['employee'] += 1

        stats['total'] += 1

    # ---------------------------------------------------------
    # 4. FILTERING & PREPARING LIST
    # ---------------------------------------------------------
    if selected_role != 'all':
        role_filter = 'team_leader' if selected_role == 'team leader' else selected_role
        filtered_ids = [uid for uid, role in role_map.items() if role == role_filter]
    else:
        filtered_ids = [uid for uid, role in role_map.items() if role != 'owner']

    staff_qs = Profile.objects.using(db_name).filter(
        user_id__in=filtered_ids
    ).select_related('user').order_by('-user__date_joined')

    processed_staff = []
    for emp in staff_qs:
        display_role = role_map.get(emp.user_id, 'employee').replace('_', ' ').title()
        emp.display_role = display_role
        processed_staff.append(emp)

    # Reset ThreadLocal so it doesn't affect other pages
    ThreadLocal.DB_NAME = 'default'

    return render(request, 'attendance/dashboard/branch_detail.html', {
        'branch': branch,
        'company': company,
        'stats': stats,
        'staff_list': processed_staff,
    })

from django.contrib.auth.models import User
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden

@login_required
def safe_delete_user(request, user_id):
    # 🛡️ ૧. સિક્યોરિટી ચેક
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")
    
    # ૨. ટાર્ગેટ યુઝર મેળવો
    target_user = get_object_or_404(User, id=user_id)

    # ડીફોલ્ટ ડેટાબેઝ નામ સેટ કરો
    db_name = 'default'

    # 🛡️ ૩. પ્રોફાઇલ ચેક અને ડેટાબેઝ નામ સેટિંગ
    try:
        if hasattr(target_user, 'profile') and target_user.profile.company:
            company = target_user.profile.company
            db_name = f"{company.slug}_db"
            
            # પર્ટીક્યુલર DB કનેક્શન ખાતરી કરો
            from .utils import ensure_db_connection
            ensure_db_connection(db_name)
            
            # પર્ટીક્યુલર DB માંથી પ્રોફાઇલ ડિલીટ કરો
            from accounts.models import Profile
            Profile.objects.using(db_name).filter(user=target_user).delete()
    except Exception as e:
        print(f"Profile cleanup skipped: {e}")

    # ૪. ક્લીનઅપ ફંક્શન (🔥 અહી db_name પાસ કરવો જરૂરી છે!)
    cleanup_user_roles_and_links(target_user, db_name=db_name)
    
    # ૫. હવે યુઝરને સેફલી ડિલીટ કરો (Main DB માંથી)
    target_user.delete()
    
    messages.success(request, f"User '{target_user.username}' and all their links removed successfully.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType

# attendance/views.py

from .middleware import ThreadLocal
from django.db import transaction
def cleanup_user_roles_and_links(user_obj, db_name):
    """
    યુઝરનો રોલ બદલાતા પહેલા તેના જૂના તમામ લિંક્સ (Squad/Manager) સાફ કરવા માટે.
    """
    user_id = user_obj.id

    # ૧. જો આ યુઝર કોઈનો Team Leader હતો, તો તેની ટીમને ફ્રી કરો
    Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader_id=None)
    Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader_id=None)

    # ૨. જો આ યુઝર કોઈનો Manager હતો, તો તે એમ્પ્લોઈઝને મેનેજર વગરના કરો
    Profile.objects.using('default').filter(manager_id=user_id).update(manager_id=None)
    Profile.objects.using(db_name).filter(manager_id=user_id).update(manager_id=None)

    # ૩. આ યુઝર પોતે જો કોઈની અંડર (Squad/Manager) હોય, તો ત્યાંથી પણ તેને હટાવો
    Profile.objects.using('default').filter(user_id=user_id).update(team_leader_id=None, manager_id=None)
    Profile.objects.using(db_name).filter(user_id=user_id).update(team_leader_id=None, manager_id=None)

    # ૪. જૂના તમામ ગ્રુપ્સ (Roles) રિમૂવ કરો
    user_obj.groups.clear()
    
        
def assign_company_owner(request, company_slug):
    if request.method == "POST":
        # ૧. ડેટા ફેચ કરો
        company = get_object_or_404(Company.objects.using('default'), slug=company_slug)
        user_id = request.POST.get('user_id')
        user = get_object_or_404(User.objects.using('default'), id=user_id)
        
        db_name = f"{company.slug}_db"
        owner_group_name = f"{company.slug}_owner"

        # ૨. મેઈન (default) ડેટાબેઝમાં ગ્રુપ એસાઇન કરો
        owner_group_main, _ = Group.objects.using('default').get_or_create(name=owner_group_name)
        user.groups.clear() # default DB માંથી ક્લિયર થશે
        user.groups.add(owner_group_main)

        # ૩. કંપનીના પર્ટિક્યુલર ડેટાબેઝમાં ગ્રુપ એસાઇન કરો
        # આનાથી કંપનીના auth_group ટેબલમાં ડેટા જશે
        try:
            # તે કંપનીના DB માં ગ્રુપ બનાવો/મેળવો
            owner_group_company, _ = Group.objects.using(db_name).get_or_create(name=owner_group_name)
            
            # તે કંપનીના DB માં યુઝર ગ્રુપ ટેબલ અપડેટ કરવા માટે
            # અમે સીધું 'using(db_name)' વાપરીને યુઝર ઓબ્જેક્ટને તે DB માં સેવ કરાવીશું
            user.save(using=db_name) 
            
            # ખાસ નોંધ: મલ્ટી-DB માં user.groups.add(group) સીધું કામ નથી કરતું
            # એટલે આપણે મેન્યુઅલી Relationship સેટ કરવી પડશે
            user_groups_model = User.groups.through
            user_groups_model.objects.using(db_name).filter(user=user).delete() # જૂના ક્લિયર કરો
            user_groups_model.objects.using(db_name).create(user=user, group=owner_group_company)
            
        except Exception as e:
            print(f"Error updating company database: {e}")

        # ૪. પ્રોફાઈલ અપડેટ (તમારા કંપની DB માં)
        from accounts.models import Profile
        Profile.objects.using(db_name).update_or_create(
            user=user,
            defaults={
                'company': company,
                'is_approved': True,
                'is_verified': True,
            }
        )

        messages.success(request, f"{user.username} is now the OWNER in both Main and {company.name} DB")
        return redirect('superadmin_company_dashboard', company_id=company.id)

@login_required
def delete_company_owner(request, company_id):
    """Owner ડિલીટ થાય ત્યારે આખી કંપની (Branches, Users, Roles) ને ફ્રી કરવા માટે"""
    
    if not request.user.is_superuser:
        return HttpResponseForbidden("Access Denied")

    company = get_object_or_404(Company.objects.using('default'), id=company_id)
    db_name = f"{company.slug}_db"

    try:
        with transaction.atomic(using='default'):
            with transaction.atomic(using=db_name):
                target_user = User.objects.using('default').get(id=user_id)
                branch_tenant = Branch.objects.using(db_name).get(id=tenant_branch_id)

                        # 🔥 THE FIX: Safely fetch the branch without crashing on duplicates
                branch_main = Branch.objects.using('default').filter(
                            company=company, 
                            name=branch_tenant.name
                        ).first()

                        # If it somehow doesn't exist in the main DB, create it manually
                if not branch_main:
                    branch_main = Branch.objects.using('default').create(
                                company=company, 
                                name=branch_tenant.name,
                                location=branch_tenant.location, 
                                latitude=branch_tenant.latitude, 
                                longitude=branch_tenant.longitude, 
                                radius=branch_tenant.radius
                            )

                        # Update Profiles
                Profile.objects.using('default').filter(user_id=user_id).update(
                            branch_id=branch_main.id, manager=None, team_leader=None
                        )
                Profile.objects.using(db_name).filter(user_id=user_id).update(
                            branch_id=branch_tenant.id, manager=None, team_leader=None
                        )

                Profile.objects.using('default').filter(manager_id=user_id).update(manager=None)
                Profile.objects.using(db_name).filter(manager_id=user_id).update(manager=None)
                Profile.objects.using('default').filter(team_leader_id=user_id).update(team_leader=None)
                Profile.objects.using(db_name).filter(team_leader_id=user_id).update(team_leader=None)

                        # Update Roles
                target_user.groups.clear()
                if role_name:
                    group_name = f"{company.slug.lower()}_{role_name}"
                    target_group, _ = Group.objects.using('default').get_or_create(name=group_name)
                    target_user.groups.add(target_group)

        messages.success(request, f"Role updated successfully. {target_user.username} is now {role_name.title()}.")
    except Exception as e:
        messages.error(request, f"Error assigning role: {str(e)}")
            
            # 🔥 BONUS FIX: This ensures you stay on the Branch Details page instead of being kicked back to the Dashboard!
    return redirect(request.META.get('HTTP_REFERER', 'company_owner_dashboard'))

import os
from django.conf import settings
from django.db import connections, transaction
from django.core.management import call_command

def setup_tenant_database(company_slug):
    """
    ૧. નવો ડેટાબેઝ ફાઈલ લેવલ પર બનાવે છે (SQLite mate)
    ૨. ટેબલ્સ બનાવે છે (Migrations run kare che)
    ૩. જરૂરી ડેટા સિંક કરે છે
    """
    db_name = f"{company_slug}_db"
    db_path = os.path.join(settings.BASE_DIR, f"{db_name}.sqlite3")

    # 🛡️ Step 1: Create Database File if not exists
    if not os.path.exists(db_path):
        open(db_path, 'w').close()
        print(f"📁 Created Database File: {db_path}")

    # 🛡️ Step 2: Dynamically Add to Django Connections
    # Aa logic vagar Django 'target_db' ne olkhshe nahi
    settings.DATABASES[db_name] = {
        'ENGINE': 'django.db.backends.sqlite3',
        'NAME': db_path,
    }

    try:
        # 🛡️ Step 3: Run Migrations on New DB
        # Aa badha tables (Attendance, Task, Profile) navi DB ma banavshe
        call_command('migrate', database=db_name, interactive=False)
        print(f"🚀 Migrations complete for {db_name}")

        # 🛡️ Step 4: Sync Initial Data
        sync_initial_data(db_name)

    except Exception as e:
        print(f"⚠️ Setup Error for {db_name}: {e}")


def redirect_to_tenant_login(request):
    # આ વ્યુ ત્યારે ચાલશે જ્યારે કોઈ સીધું /accounts/login/ પર જશે
    # અહીં તમે કૂકીઝ અથવા રેફરર યુઆરએલ માંથી સ્લગ શોધી શકો છો
    # અત્યારે આપણે તેને મેઈન રજીસ્ટ્રેશન અથવા ગ્લોબલ લોગિન પર મોકલી શકીએ
    return redirect('attendance_home') # અથવા કોઈ બેઝ પેજ

def sync_initial_data(target_db_name, company_slug):
    from django.contrib.auth.models import User, Group
    from accounts.models import Profile
    from attendance.models import Company
    
    # ૧. મેઈન ડેટાબેઝમાંથી કંપનીનો ઓરિજિનલ ડેટા મેળવો
    original_company = Company.objects.using('default').get(slug=company_slug)
    
    try:
        # ૨. આ કંપનીને નવા ટેનન્ટ ડેટાબેઝમાં કોપી કરો (Constraint Fix 🔥)
        # આનાથી 'company_id' વાળી ભૂલ કાયમ માટે જતી રહેશે
        original_company.save(using=target_db_name)
        print(f"✅ Company '{company_slug}' synced to {target_db_name}")

        # ૩. સુપરયુઝર્સ સિંક કરો (એડમિન એક્સેસ માટે)
        superusers = User.objects.using('default').filter(is_superuser=True)
        for user in superusers:
            user.save(using=target_db_name)
            
        # ૪. ગ્રુપ્સ/રોલ્સ સિંક કરો
        groups = Group.objects.using('default').all()
        for group in groups:
            group.save(using=target_db_name)

    except Exception as e:
        print(f"❌ Sync Error: {e}")

from django.db import connection, connections
from django.core.management import call_command
from django.conf import settings


# attendance/views.py માં આ ફંક્શન શોધીને આ મુજબ અપડેટ કરો

from django.db import connection, connections
from django.conf import settings
from django.core.management import call_command
from django.db.utils import OperationalError

from django.db import connection, connections

def setup_company_db(company_slug):
    db_name = f"{company_slug}_db"
    
    # 🛡️ ૧. પેલા મેન્યુઅલી ડેટાબેઝ બનાવો
    try:
        with connection.cursor() as cursor:
            # આ લાઇન MySQL માં ડેટાબેઝ બનાવશે
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {db_name}")
            print(f"✅ MySQL Database Created: {db_name}")
    except Exception as e:
        print(f"❌ Database Creation Failed: {e}")
        return False

    # ૨. નવા ડેટાબેઝનું કનેક્શન સેટિંગ્સમાં એડ કરો
    new_db_settings = settings.DATABASES['default'].copy()
    new_db_settings['NAME'] = db_name
    settings.DATABASES[db_name] = new_db_settings
    connections.databases[db_name] = new_db_settings

    # 🛡️ ૩. ટેબલ્સ બનાવો (Migrations રન કરો)
    try:
        from django.core.management import call_command
        # આ લાઇન નવા DB માં બધા જ ટેબલ્સ (User, Profile, Attendance) બનાવશે
        call_command('migrate', database=db_name, interactive=False)
        print(f"✅ Migrations Finished for {db_name}")
        
        # ૪. શરૂઆતનો ડેટા સિંક કરો (Superuser વગેરે)
        sync_initial_data(db_name, company_slug)
        return True
    except Exception as e:
        print(f"❌ Migration Error: {e}")
        return False        

from django.db import connections

def clear_new_db_data(db_alias):
    """
    નવા બનાવેલા ટેનન્ટ ડેટાબેઝમાંથી બિનજરૂરી ડેટા સાફ કરે છે (MySQL Edition).
    આનાથી નવા ક્લાયન્ટને એકદમ ફ્રેશ (ખાલી) ડેટાબેઝ મળશે.
    """
    connection = connections[db_alias]
    cursor = connection.cursor()
    
    # ૧. આ ટેબલ્સને ક્યારેય ખાલી ન કરવા (Core Django Tables)
    exclude_tables = [
        'django_migrations', 
        'django_content_type', 
        'auth_permission', # પરમિશન ક્લિયર ના કરવી નહીતર લોગિન નહિ થાય
        'django_session',
    ]
    
    try:
        # ૨. ફોરેન કી ચેક બંધ કરો (MySQL Syntax)
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
        
        # ૩. ડેટાબેઝના તમામ ટેબલ્સના નામ મેળવો
        cursor.execute("SHOW TABLES;")
        tables = cursor.fetchall()
        
        for table in tables:
            table_name = table[0]
            if table_name not in exclude_tables:
                # DELETE ને બદલે TRUNCATE વાપરવું વધુ સારું છે કારણ કે તે 
                # ID (Auto-increment) ને પણ 1 થી રીસેટ કરી દેશે.
                cursor.execute(f"TRUNCATE TABLE {table_name};")
                print(f"🧹 Cleaned table: {table_name}")
        
        # ૪. ફોરેન કી ચેક ફરી ચાલુ કરો
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
        connection.commit()
        
        print(f"✅ Database {db_alias} is now fresh and ready for the new client!")
        
    except Exception as e:
        print(f"❌ Cleanup error for {db_alias}: {e}")

# --- ખૂટતા ફંક્શન્સ જે URLs માં વપરાયેલા છે ---

@login_required
def company_admin_pending_users(request):
    """કંપની એડમિન માટે પેન્ડિંગ યુઝર્સ જોવા માટેનું વ્યુ"""
    try:
        profile_admin = request.user.profile
        company = profile_admin.company
        db_name = f"{company.slug}_db"
        ThreadLocal.DB_NAME = db_name
    except:
        return HttpResponse("Admin Profile Error")

    # કંપનીના ડેટાબેઝમાંથી એવા યુઝર્સ લાવો જે વેરિફાઈડ છે પણ અપ્રૂવ્ડ નથી
    users = Profile.objects.using(db_name).filter(
        company=company,
        is_verified=True,
        is_approved=False
    )
    return render(request, "attendance/admin/company_admin_pending_users.html", {"users": users})

@login_required
def company_admin_approve_user(request, user_id):
    """કંપની એડમિન દ્વારા યુઝરને અપ્રૂવ કરવા માટે"""
    profile_admin = request.user.profile
    db_name = f"{profile_admin.company.slug}_db"

    profile = get_object_or_404(Profile, user__id=user_id)
    profile.is_approved = True
    profile.save() # Main DB
    profile.save(using=db_name) # Company DB sync

    messages.success(request, f"User {profile.user.username} approved.")
    return redirect("company_admin_pending_users")

@login_required
def company_admin_reject_user(request, user_id):
    """કંપની એડમિન દ્વારા યુઝરને રિજેક્ટ કરવા માટે"""
    profile_admin = request.user.profile
    db_name = f"{profile_admin.company.slug}_db"

    profile = get_object_or_404(Profile, user__id=user_id)
    profile.company = None # લિંક તોડી નાખો
    profile.is_approved = False
    profile.save()
    
    # કંપની DB માંથી પણ કાઢી નાખો
    try:
        Profile.objects.using(db_name).filter(user__id=user_id).delete()
    except:
        pass

    messages.warning(request, "User request rejected.")
    return redirect("company_admin_pending_users")